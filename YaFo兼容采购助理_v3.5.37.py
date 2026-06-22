# -*- coding: utf-8 -*-
"""
YaFo兼容采购助理 v3.5.37
功能：统一处理工厂直发订单和分仓订单，生成XML文件
- 工厂直发模式：一次性上传多个供应商文件，自动识别供应商，批量生成XML和汇总表
- 分仓订单模式：一次性上传多个分仓文件，自动识别分仓，批量生成XML
- 存货成本回填模式：选择NC成本表，批量导入供应商成本来源表，暂存后统一导出Excel
- 销量预测与备货建议模式：导入销售订单和库存信息，生成预测销量、备货建议和Excel报告
更新：
  v3.5.37 - 修复主界面功能选择横向单排显示不全问题，改为多行网格布局并优化标题区适配
  v3.5.36 - 新增集成销量预测与备货建议生成器为功能选择子功能，并优化CSV/Excel自动读取与子窗口嵌入逻辑
  v3.5.35 - 新增集成存货成本回填生成器为功能选择子功能，选择后自动弹出成本回填子模块窗口
  v3.5.34 - 集成新品上新定价生成器为功能选择子功能，选择后自动弹出定价子模块窗口
  v3.5.33 - 修复界面中文字在高DPI/不同字体环境下显示不全的问题
  v3.5.32 - 界面美化，NC产品信息按钮文案由CSV导出调整为CSV导入
  v3.5.31 - 将供应商直发表格生成器作为功能选择项，选择后自动弹出子模块窗口
  v3.5.30 - 集成供应商直发表格生成器子模块：预处理预定单并按供应商拆分导出Excel
  v3.5.29 - 新增周结预定单统计表支持：按供应商名称自动拆分并分别生成工厂直发XML
  v3.5.28 - 集成NC存货信息CSV导入功能，可在主界面直接从CSV生成NC产品信息Excel并自动加载
  v3.5.27 - 新增：界面增加"维护日志"按钮，点击查看版本更新历史
  v3.5.26 - 修复：get_product_name_by_model使用regex=False避免正则特殊字符问题
  v3.5.25 - 修复：征图分仓格式同时设置product_code和product_name_keyword列
  v3.5.24 - 修复征图分仓格式检测：有"型号"+"产品名称"+"Unnamed"列时识别为征图分仓
  v3.5.23 - 修复征图分仓格式识别和处理，添加get_product_by_name_simple函数
  v3.5.22 - 修复德源分仓和征图分仓的区分：德源有"产品编号"无"型号"，征图有"型号"
  v3.5.21 - 修复德源分仓文件识别问题，通过列名特征识别德源分仓格式
  v3.5.20 - 修复德源分仓文件识别问题，添加迪研供应商识别
  v3.5.15 - 新增中山迪研电子有限公司(X15)供应商
  v3.5.13 - 修复绘强格式单价列匹配
  v3.5.12 - 修复header=1导致存货编码匹配失败
  v3.5.11 - 修复存货编码判断优先级
  v3.5.10 - 修复绘强格式存货编码匹配
  v3.5.9 - 重新修复三家供应商识别逻辑
        - 核心修改：优先检查"价格"列（不是"单价"列）来判断是否是三家供应商格式
        - 判断优先级：价格列 → 客户产品编码 → 对方货号 → 供应商+存货编码(绘强)
        - 三家供应商（征图/纳思达/博克）：有"价格"列，使用"存货编码"作为产品编码
        - 绘强格式：没有"价格"列，有"供应商"和"存货编码"列（列名错位）
  v3.5.8 - 修复三家供应商（征图、纳思达、博克）识别问题
        - 增加三家供应商格式识别：同时有"存货编码"+"供应商"+"存货名称"+"价格"列
        - 三家供应商格式使用"存货编码"列作为产品编码（值为GA5723、XD0036等格式）
        - 优先检查三家供应商特征，避免被误识别为绘强格式
  v3.5.7 - 修复三家供应商识别：优先检查"存货编码"+"供应商"+"存货名称"组合
  v3.5.5 - 关键发现：绘强发的"型号"列实际存的是存货名称，不是真正的型号
        - 修改get_product_name_by_model函数，匹配NC的"存货名称"列而非"型号"列
        - 这与compare_with_nc函数的匹配逻辑一致
  v3.5.4 - 关键bug修复：is_huiqiang_warehouse未添加到col_mapping返回值
        - 添加col_mapping['is_huiqiang_warehouse'] = is_huiqiang_warehouse
        - 确保绘强发分仓格式正确走型号匹配分支
  v3.5.3 - 最终修复绘强发分仓格式型号匹配问题
        - 确保get_product_name_by_model函数正确工作
        - 确保绘强发分仓格式走型号匹配分支
  v3.5.2 - 修复绘强发分仓格式型号匹配问题
        - 新增get_product_name_by_model函数，用型号匹配NC的"型号"列
        - 修改分仓订单匹配逻辑，绘强发格式使用型号匹配
  v3.5.1 - 修复绘强发分仓格式识别问题
        - 增加"型号"列识别条件
        - 新增绘强发分仓格式（用型号列）
  v3.5 - 新增绘强发北京仓自动与NC存货清单比对功能
       - 生成比对结果Excel文件
       - 显示匹配率
       - 其他分仓规则不变
  v3.4 - 分仓订单支持多文件批量处理
       - 删除分仓下拉选择，改为文件名自动识别分仓（北京/沈阳/南京）
       - 分仓订单界面与工厂直发统一风格
  v3.3 - 修复XML格式和汇总表格式
       - XML声明改为<?xml version="1.0" ?>（不带encoding）
       - <vmemo/>使用自闭合标签
       - 汇总表改为"单据号汇总"和"详细明细"两个工作表
       - 增加单据号、规格型号、收件人、快递单号等字段
       - 征图格式：用"型号"列匹配NC产品名称（而非条码）
       - 绘强格式：用"业务员"列作为单据号
  v3.2 - 修复XML格式，与NC系统完全兼容
       - 修正根节点属性：billtype="21", roottag="voucher"
       - 修正明细节点：<billbody><entry>结构
       - 修正字段名：cmangid, nordernum, cwarehouseid等
       - 扩展header尝试范围，支持征图文件(header=4)
  v3.1 - 合并两个生成器，支持多选上传
       - 通过文件内容自动识别供应商
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import csv
from dataclasses import dataclass
from typing import Dict, List, Optional
import os
import sys
from datetime import datetime
import xml.etree.ElementTree as ET
from xml.dom import minidom
import warnings
import re
warnings.filterwarnings('ignore')

# OCR相关导入（暂时禁用）
OCR_AVAILABLE = False
OCR_IMPORT_ERROR = "OCR功能暂时禁用，网络稳定后可重新启用"

# ===== 工厂直发供应商列表 =====
SUPPLIERS = [
    ("兴发公司", "中山澳兴发科技有限公司/中山市兴发电子科技有限公司"),
    ("绘强", "中山市绘强电子科技有限公司(X08)"),
    ("征图", "珠海征图打印耗材有限公司(X14)"),
    ("迪研", "中山迪研电子有限公司(X15)"),
    ("德源", "中山市德源影像科技有限公司(X21)"),
    ("纳思达", "珠海纳思达信息技术有限公司"),
    ("博克", "扬州市博克打印耗材有限公司"),
    ("联合天润", "珠海联合天润打印耗材有限公司（X53)"),
    ("中主世隆", "北京中主世隆科技发展有限公司"),
    ("华人智创", "珠海华人智创科技有限公司"),
    ("回收", "回收（硒鼓）供应商")
]

# 供应商关键词映射（用于自动识别）
SUPPLIER_KEYWORDS = {
    "兴发": "兴发公司",
    "绘强": "绘强",
    "征图": "征图",
    "迪研": "迪研",
    "德源": "德源",
    "纳思达": "纳思达",
    "博克": "博克",
    "联合天润": "联合天润",
    "中主世隆": "中主世隆",
    "华人智创": "华人智创",
    "回收": "回收",
    "周结": "周结汇总",
    "预定单统计": "周结汇总"
}

# ===== 分仓列表 =====
WAREHOUSES = [
    ("北京", "北京仓"),
    ("沈阳", "沈阳仓"),
    ("南京", "南京仓")
]

# ===== 订单头信息（统一）=====
ORDER_HEADER = {
    "cbiztype": "兼容采购",
    "ctermprotocolid": "60天",
    "vdef2": "不预提",
    "vdef3": "未付",
    "vdef20": "北京智通仁和科技发展有限公司",
    "cdeptid": "市场管理部",
    "cemployeeid": "汪浩",
    "coperator": "雷人杰",
    "cuserid": "雷人杰",
    "vmemo": ""
}


class PurchaseOrderGenerator:
    def __init__(self, root):
        self.root = root
        self.root.title("YaFo兼容采购助理 v3.5.37")
        self.root.geometry("1100x940")
        self.root.minsize(1000, 880)
        self.root.resizable(True, True)
        
        # 变量
        self.product_file = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.order_type = tk.StringVar(value="factory")
        
        # 工厂直发模式：已选择的文件列表 [(文件路径, 供应商名称), ...]
        self.selected_files = []
        
        # 分仓订单模式：已选择的文件列表 [(文件路径, 分仓名称), ...]
        self.warehouse_files = []

        # 供应商直发表格子模块窗口引用
        self.supplier_export_window = None
        # 新品定价子模块窗口引用
        self.pricing_window = None
        # 存货成本回填子模块窗口引用
        self.cost_fill_window = None
        self.cost_fill_app = None
        # 销量预测与备货建议子模块窗口引用
        self.sales_forecast_window = None
        
        # 产品数据
        self.product_df = None
        
        # OCR引擎
        self.ocr_engine = None
        
        # UI组件引用
        self.factory_frame = None
        self.warehouse_frame = None
        self.file_listbox = None
        
        self.setup_ui()
        
        # 初始化OCR
        if OCR_AVAILABLE:
            try:
                self.ocr_engine = easyocr.Reader(['ch_sim', 'en'], gpu=False)
                self.log("✓ OCR引擎加载成功（easyocr）")
            except Exception as e:
                self.log(f"⚠ OCR引擎加载失败：{str(e)}")
        else:
            self.log(f"⚠ OCR模块未启用：{OCR_IMPORT_ERROR}")
        
        # 加载保存的配置
        self.load_saved_config()
        
        # 初始状态
        self.on_order_type_change()
    
    def setup_ui(self):
        """设置界面"""
        # 统一视觉参数
        self.ui_bg = "#f4f7fb"
        self.card_bg = "#ffffff"
        self.primary = "#2563eb"
        self.success = "#16a34a"
        self.danger = "#dc2626"
        self.muted = "#64748b"
        self.text_color = "#0f172a"
        self.border = "#dbe3ef"
        self.root.configure(bg=self.ui_bg)

        def make_card(parent, title, **kwargs):
            """创建卡片容器。

            不再使用 tk.LabelFrame：在部分 Windows 高DPI/字体缩放环境下，
            LabelFrame 的标题文字容易被边框裁切。改为 Frame + 独立标题 Label，
            以保证中文标题完整显示。
            """
            frame = tk.Frame(
                parent,
                padx=16,
                pady=10,
                bg=self.card_bg,
                bd=0,
                highlightbackground=self.border,
                highlightcolor=self.border,
                highlightthickness=1,
                **kwargs
            )
            title_label = tk.Label(
                frame,
                text=title,
                font=("微软雅黑", 11, "bold"),
                bg=self.card_bg,
                fg=self.text_color,
                anchor="w"
            )
            title_label.pack(fill="x", anchor="w", pady=(0, 8))
            return frame

        def make_btn(parent, text, command, bg="#e2e8f0", fg="#0f172a", width=None, height=None, font_size=10, bold=False):
            font = ("微软雅黑", font_size, "bold" if bold else "normal")
            btn = tk.Button(
                parent,
                text=text,
                command=command,
                font=font,
                bg=bg,
                fg=fg,
                activebackground=bg,
                activeforeground=fg,
                relief="flat",
                bd=0,
                padx=10,
                pady=5,
                cursor="hand2",
                width=width,
                height=height
            )
            return btn

        # 顶部标题区
        header_frame = tk.Frame(self.root, bg=self.primary, height=88)
        header_frame.pack(fill="x")
        header_frame.pack_propagate(False)

        title_block = tk.Frame(header_frame, bg=self.primary)
        title_block.pack(side="left", fill="x", expand=True, padx=24, pady=10)

        title_label = tk.Label(
            title_block,
            text="YaFo兼容采购助理 v3.5.37",
            font=("微软雅黑", 20, "bold"),
            fg="white",
            bg=self.primary
        )
        title_label.pack(anchor="w")

        subtitle_label = tk.Label(
            title_block,
            text="工厂直发 / 分仓订单 / 供应商直发表格处理 / 新品定价 / 存货成本回填 / 销量预测与备货建议 / NC存货CSV导入",
            font=("微软雅黑", 9),
            fg="#dbeafe",
            bg=self.primary,
            anchor="w",
            justify="left",
            wraplength=820
        )
        subtitle_label.pack(fill="x", anchor="w", pady=(2, 0))

        make_btn(header_frame, "维护日志", self.show_changelog, bg="#1e40af", fg="white", width=10).pack(side="right", padx=24, pady=18)

        # 主内容容器
        main = tk.Frame(self.root, bg=self.ui_bg)
        main.pack(fill="both", expand=True, padx=18, pady=12)

        # ===== 功能选择 =====
        type_frame = make_card(main, "功能选择")
        type_frame.pack(fill="x", pady=(0, 10))

        type_radio_frame = tk.Frame(type_frame, bg=self.card_bg)
        type_radio_frame.pack(fill="x", anchor="w")

        radio_opts = [
            ("工厂直发订单", "factory"),
            ("分仓订单", "warehouse"),
            ("供应商直发表格处理", "supplier_export"),
            ("新品上新定价", "pricing"),
            ("存货成本回填", "cost_fill"),
            ("销量预测与备货建议", "sales_forecast"),
        ]

        # 功能项较多时，单排横向排布容易在小屏/高DPI环境下被截断。
        # 改为3列多行网格，后续继续增加功能时也能自动换行。
        max_cols = 3
        for col_idx in range(max_cols):
            type_radio_frame.grid_columnconfigure(col_idx, weight=1, uniform="function_cols")

        for idx, (label, value) in enumerate(radio_opts):
            row_idx = idx // max_cols
            col_idx = idx % max_cols
            rb = tk.Radiobutton(
                type_radio_frame,
                text=label,
                variable=self.order_type,
                value=value,
                font=("微软雅黑", 10),
                command=self.on_order_type_change,
                bg=self.card_bg,
                fg=self.text_color,
                activebackground=self.card_bg,
                activeforeground=self.primary,
                selectcolor=self.card_bg,
                anchor="w",
                justify="left",
                wraplength=220,
                padx=8,
                pady=4
            )
            rb.grid(row=row_idx, column=col_idx, sticky="w", padx=(0, 18), pady=3)

        # ===== 工厂直发模式区域 =====
        self.factory_frame = make_card(main, "工厂直发文件上传（支持多选，自动识别供应商）")

        btn_row = tk.Frame(self.factory_frame, bg=self.card_bg)
        btn_row.pack(fill="x", pady=(0, 8))

        make_btn(btn_row, "选择文件（支持多选）", self.select_multiple_files, bg=self.primary, fg="white", width=20, bold=True).pack(side="left", padx=(0, 8))
        make_btn(btn_row, "清空列表", self.clear_file_list, bg="#e2e8f0", fg=self.text_color, width=12).pack(side="left", padx=4)
        make_btn(btn_row, "移除选中", self.remove_selected_file, bg="#fee2e2", fg="#991b1b", width=12).pack(side="left", padx=4)

        list_frame = tk.Frame(self.factory_frame, bg=self.card_bg)
        list_frame.pack(fill="both", pady=4, expand=True)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side="right", fill="y")

        self.file_listbox = tk.Listbox(
            list_frame,
            height=8,
            font=("微软雅黑", 10),
            yscrollcommand=scrollbar.set,
            selectmode=tk.EXTENDED,
            bg="#f8fafc",
            fg=self.text_color,
            bd=0,
            highlightthickness=1,
            highlightbackground=self.border,
            selectbackground="#bfdbfe",
            selectforeground=self.text_color
        )
        self.file_listbox.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=self.file_listbox.yview)
        self.file_listbox.insert(tk.END, "  供应商          文件名")
        self.file_listbox.insert(tk.END, "─" * 72)

        # ===== 分仓订单模式区域 =====
        self.warehouse_frame = make_card(main, "分仓订单配置")

        warehouse_btn_frame = tk.Frame(self.warehouse_frame, bg=self.card_bg)
        warehouse_btn_frame.pack(fill="x", pady=(0, 8))
        make_btn(warehouse_btn_frame, "选择分仓文件（支持多选）", self.select_warehouse_files, bg=self.primary, fg="white", width=24, bold=True).pack(side="left", padx=(0, 8))
        make_btn(warehouse_btn_frame, "清空", self.clear_warehouse_files, bg="#e2e8f0", fg=self.text_color, width=8).pack(side="left", padx=4)

        warehouse_list_frame = tk.Frame(self.warehouse_frame, bg=self.card_bg)
        warehouse_list_frame.pack(fill="both", pady=4)

        scrollbar_w = tk.Scrollbar(warehouse_list_frame)
        scrollbar_w.pack(side="right", fill="y")

        self.warehouse_listbox = tk.Listbox(
            warehouse_list_frame,
            height=6,
            font=("微软雅黑", 10),
            yscrollcommand=scrollbar_w.set,
            bg="#f8fafc",
            fg=self.text_color,
            bd=0,
            highlightthickness=1,
            highlightbackground=self.border,
            selectbackground="#bfdbfe",
            selectforeground=self.text_color
        )
        self.warehouse_listbox.pack(side="left", fill="both", expand=True)
        scrollbar_w.config(command=self.warehouse_listbox.yview)
        self.warehouse_listbox.insert(tk.END, "  分仓            文件名")
        self.warehouse_listbox.insert(tk.END, "─" * 72)
        self.warehouse_files = []

        # ===== 公共配置区域 =====
        config_frame = make_card(main, "配置")
        config_frame.pack(fill="x", pady=10)

        row1 = tk.Frame(config_frame, bg=self.card_bg)
        row1.pack(fill="x", pady=4)
        tk.Label(row1, text="NC产品信息文件", font=("微软雅黑", 10), width=17, anchor="w", bg=self.card_bg, fg=self.text_color).pack(side="left")
        tk.Entry(row1, textvariable=self.product_file, font=("微软雅黑", 9), width=60, relief="solid", bd=1, highlightthickness=0).pack(side="left", padx=6, ipady=3)
        make_btn(row1, "选择", self.select_product_file, width=6).pack(side="left", padx=2)
        make_btn(row1, "记住", self.save_product_path, width=5).pack(side="left", padx=2)
        make_btn(row1, "CSV导入", self.export_nc_csv_to_excel, bg="#7c3aed", fg="white", width=8, bold=True).pack(side="left", padx=2)

        row2 = tk.Frame(config_frame, bg=self.card_bg)
        row2.pack(fill="x", pady=4)
        tk.Label(row2, text="输出目录", font=("微软雅黑", 10), width=17, anchor="w", bg=self.card_bg, fg=self.text_color).pack(side="left")
        tk.Entry(row2, textvariable=self.output_dir, font=("微软雅黑", 9), width=60, relief="solid", bd=1, highlightthickness=0).pack(side="left", padx=6, ipady=3)
        make_btn(row2, "选择", self.select_output_dir, width=6).pack(side="left", padx=2)
        make_btn(row2, "打开", self.open_output_dir, width=5).pack(side="left", padx=2)

        # ===== 操作按钮 =====
        btn_frame = tk.Frame(main, bg=self.ui_bg)
        btn_frame.pack(fill="x", pady=8)

        self.run_btn = make_btn(btn_frame, "生成XML和汇总表", self.run_process, bg=self.success, fg="white", width=20, height=2, font_size=12, bold=True)
        self.run_btn.pack(side="left", padx=(0, 10))
        make_btn(btn_frame, "清空", self.clear_all_files, bg="#e2e8f0", fg=self.text_color, width=10, height=2).pack(side="left", padx=6)

        # ===== 进度显示 =====
        progress_frame = make_card(main, "处理日志")
        progress_frame.pack(fill="both", pady=(8, 0), expand=True)

        scroll_frame = tk.Frame(progress_frame, bg=self.card_bg)
        scroll_frame.pack(fill="both", expand=True)

        scrollbar2 = tk.Scrollbar(scroll_frame)
        scrollbar2.pack(side="right", fill="y")

        self.progress_text = tk.Text(
            scroll_frame,
            height=12,
            font=("Consolas", 10),
            yscrollcommand=scrollbar2.set,
            bg="#0f172a",
            fg="#e2e8f0",
            insertbackground="#e2e8f0",
            bd=0,
            padx=10,
            pady=8
        )
        self.progress_text.pack(side="left", fill="both", expand=True)
        scrollbar2.config(command=self.progress_text.yview)

        # 默认输出目录为桌面
        self.output_dir.set(os.path.join(os.path.expanduser("~"), "Desktop"))

    def on_order_type_change(self):
        """功能选择切换事件"""
        order_type = self.order_type.get()

        if order_type == "factory":
            # 显示工厂直发模式，隐藏分仓模式
            self.factory_frame.pack(fill="x", padx=20, pady=8)
            self.warehouse_frame.pack_forget()
            self.run_btn.config(text="▶ 生成XML和汇总表", state="normal")
        elif order_type == "warehouse":
            # 显示分仓模式，隐藏工厂直发模式
            self.factory_frame.pack_forget()
            self.warehouse_frame.pack(fill="x", padx=20, pady=8)
            self.run_btn.config(text="▶ 生成分仓XML文件", state="normal")
        elif order_type == "supplier_export":
            # 供应商直发表格处理作为独立子模块弹窗
            self.factory_frame.pack_forget()
            self.warehouse_frame.pack_forget()
            self.run_btn.config(text="▶ 已打开直发表格子模块", state="disabled")
            self.open_supplier_export_module()
        elif order_type == "pricing":
            # 新品上新定价作为独立子模块弹窗
            self.factory_frame.pack_forget()
            self.warehouse_frame.pack_forget()
            self.run_btn.config(text="▶ 已打开新品定价子模块", state="disabled")
            self.open_pricing_module()
        elif order_type == "cost_fill":
            # 存货成本回填作为独立子模块弹窗
            self.factory_frame.pack_forget()
            self.warehouse_frame.pack_forget()
            self.run_btn.config(text="▶ 已打开成本回填子模块", state="disabled")
            self.open_inventory_cost_module()
        elif order_type == "sales_forecast":
            # 销量预测与备货建议作为独立子模块弹窗
            self.factory_frame.pack_forget()
            self.warehouse_frame.pack_forget()
            self.run_btn.config(text="▶ 已打开销量预测子模块", state="disabled")
            self.open_sales_forecast_module()
    
    def prepare_excel_for_pandas(self, file_path):
        """兼容读取Excel文件，尤其是老式.xls文件

        pandas读取.xls通常依赖xlrd。如果当前环境缺少xlrd，则优先尝试将.xls临时转换为.xlsx后再读取。
        转换方式：Windows Excel COM（如可用）→ LibreOffice/soffice（如可用）。
        """
        if not file_path:
            return file_path

        # 非xls文件直接返回；xlsx通常由openpyxl读取
        ext = os.path.splitext(file_path)[1].lower()
        if ext != '.xls':
            return file_path

        # 如果当前环境已能直接读取xls，则不做转换
        try:
            test_xls = pd.ExcelFile(file_path)
            test_xls.close()
            return file_path
        except Exception as original_error:
            converted_dir = os.path.join(os.path.expanduser('~'), '.po_xml_generator_cache')
            os.makedirs(converted_dir, exist_ok=True)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            timestamp = datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y%m%d%H%M%S')
            converted_path = os.path.join(converted_dir, f"{base_name}_{timestamp}.xlsx")

            if os.path.exists(converted_path):
                return converted_path

            # 方式1：Windows环境尝试调用本机Excel转换
            if os.name == 'nt':
                try:
                    import win32com.client  # type: ignore
                    excel = win32com.client.Dispatch('Excel.Application')
                    excel.Visible = False
                    excel.DisplayAlerts = False
                    wb = excel.Workbooks.Open(os.path.abspath(file_path))
                    wb.SaveAs(os.path.abspath(converted_path), FileFormat=51)  # 51 = xlsx
                    wb.Close(False)
                    excel.Quit()
                    self.log(f"   ✓ 已将xls临时转换为xlsx：{os.path.basename(converted_path)}")
                    return converted_path
                except Exception:
                    try:
                        excel.Quit()
                    except Exception:
                        pass

            # 方式2：尝试调用LibreOffice/soffice转换
            try:
                import shutil
                import subprocess
                soffice = shutil.which('soffice') or shutil.which('libreoffice')
                if soffice:
                    tmp_dir = converted_dir
                    subprocess.run(
                        [soffice, '--headless', '--convert-to', 'xlsx', '--outdir', tmp_dir, file_path],
                        check=True,
                        stdout=subprocess.DEVNULL,
                        stderr=subprocess.DEVNULL
                    )
                    libreoffice_output = os.path.join(tmp_dir, f"{base_name}.xlsx")
                    if os.path.exists(libreoffice_output):
                        os.replace(libreoffice_output, converted_path)
                        self.log(f"   ✓ 已将xls临时转换为xlsx：{os.path.basename(converted_path)}")
                        return converted_path
            except Exception:
                pass

            raise RuntimeError(
                "当前环境无法直接读取.xls文件。请安装 xlrd>=2.0.1，或把该文件另存为.xlsx后再选择。"
            ) from original_error

    def identify_supplier_from_content(self, file_path):
        """从文件内容中识别供应商"""
        try:
            # 读取Excel文件
            excel_path = self.prepare_excel_for_pandas(file_path)
            xls = pd.ExcelFile(excel_path)
            
            # 尝试读取第一个工作表
            for sheet in xls.sheet_names:
                try:
                    # 先读取header=0，检查第一行第一列是否是供应商名称
                    df0 = pd.read_excel(excel_path, header=0, sheet_name=sheet)
                    if len(df0.columns) > 0:
                        first_cell = str(df0.columns[0]).strip()
                        
                        # 检测第一列是否包含供应商名称
                        if '兴发' in first_cell:
                            return "兴发公司"
                        if '德源' in first_cell:
                            return "德源"
                        if '征图' in first_cell or 'X14' in first_cell:
                            return "征图"
                        if '绘强' in first_cell or 'X08' in first_cell:
                            return "绘强"
                        if '纳思达' in first_cell:
                            return "纳思达"
                        if '博克' in first_cell:
                            return "博克"
                        if '迪研' in first_cell:
                            return "迪研"
                        if '联合天润' in first_cell or 'X53' in first_cell:
                            return "联合天润"
                        if '中主世隆' in first_cell:
                            return "中主世隆"
                        if '华人智创' in first_cell:
                            return "华人智创"
                    
                    # 尝试不同header读取列名
                    for h in [0, 1, 2, 3, 4]:
                        try:
                            df = pd.read_excel(excel_path, header=h, sheet_name=sheet)
                            df.columns = df.columns.astype(str).str.strip()
                            cols = list(df.columns)
                            
                            # 周结预定单统计表：一个文件内混合多个供应商，后续按“供应商名称”拆分生成XML
                            if all(c in cols for c in ['供应商名称', '存货编码', '主数量', '单价', '单据号']):
                                return "周结汇总"
                            
                            # =====================================================
                            # 核心判断逻辑（按优先级顺序，与match_columns一致）
                            # =====================================================
                            
                            # 1. 首先检查是否有"价格"列 → 三家供应商格式（征图/纳思达/博克）
                            if '价格' in cols:
                                # 根据供应商列的值判断具体是哪家
                                if '供应商' in cols:
                                    supplier_values = df['供应商'].dropna().unique()
                                    for val in supplier_values:
                                        val_str = str(val).strip()
                                        # 德源
                                        if '德源' in val_str:
                                            return "德源"
                                        # 纳思达：S03, S03/S04
                                        if 'S03' in val_str or '纳思达' in val_str:
                                            return "纳思达"
                                        # 博克：S02
                                        if 'S02' in val_str or '博克' in val_str:
                                            return "博克"
                                        # 征图：S04
                                        if 'S04' in val_str or '征图' in val_str or 'X14' in val_str:
                                            return "征图"
                                return "征图"  # 默认三家供应商
                            
                            # 2. 检查是否有"客户产品编码"列 → 兴发格式
                            if '客户产品编码' in cols:
                                return "兴发公司"
                            
                            # 3. 检查是否有"对方货号"列 → 征图旧格式
                            if '对方货号' in cols:
                                return "征图"
                            
                            # 4. 检查是否有"供应商"和"存货编码"列
                            if '供应商' in cols and '存货编码' in cols:
                                # 根据供应商列的值判断具体是哪家
                                supplier_values = df['供应商'].dropna().unique()
                                for val in supplier_values:
                                    val_str = str(val).strip()
                                    # 德源
                                    if '德源' in val_str:
                                        return "德源"
                                    # 纳思达：S03, S03/S04
                                    if 'S03' in val_str or '纳思达' in val_str:
                                        return "纳思达"
                                    # 博克：S02
                                    if 'S02' in val_str or '博克' in val_str:
                                        return "博克"
                                    # 征图：S04
                                    if 'S04' in val_str or '征图' in val_str or 'X14' in val_str:
                                        return "征图"
                                # 如果供应商值是"绘强"或"X08"，则是绘强格式
                                for val in supplier_values:
                                    val_str = str(val).strip()
                                    if '绘强' in val_str or 'X08' in val_str:
                                        return "绘强"
                                # 默认三家供应商（如果供应商名是其他）
                                return "征图"
                            
                            # 5. 检查是否有"客户品号"或"客户货号"列 → 分仓订单格式
                            if '客户品号' in cols or '客户货号' in cols:
                                return "分仓订单"
                            
                        except:
                            continue
                            
                except:
                    continue
            
            # 如果都识别不出来，尝试从文件名识别作为后备
            filename = os.path.basename(file_path)
            for keyword, supplier_name in SUPPLIER_KEYWORDS.items():
                if keyword in filename:
                    return supplier_name
            
            return "未识别"
            
        except Exception as e:
            # 如果读取失败，尝试从文件名识别
            filename = os.path.basename(file_path)
            for keyword, supplier_name in SUPPLIER_KEYWORDS.items():
                if keyword in filename:
                    return supplier_name
            return "未识别"
    
    def select_multiple_files(self):
        """一次性选择多个供应商文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择供应商文件（支持多选）",
            filetypes=[
                ("所有支持的文件", "*.xlsx *.xls *.png *.jpg *.jpeg"),
                ("Excel文件", "*.xlsx *.xls"),
                ("图片文件", "*.png *.jpg *.jpeg"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_paths:
            self.log("正在识别供应商...")
            for file_path in file_paths:
                filename = os.path.basename(file_path)
                # 从文件内容识别供应商
                supplier = self.identify_supplier_from_content(file_path)
                self.selected_files.append((file_path, supplier))
                
                # 更新列表显示
                display_name = filename[:42] + "..." if len(filename) > 42 else filename
                self.file_listbox.insert(tk.END, f"  {supplier:10}    {display_name}")
                
                self.log(f"✓ {filename} → {supplier}")
    
    def clear_file_list(self):
        """清空文件列表"""
        self.selected_files = []
        self.file_listbox.delete(0, tk.END)
        self.file_listbox.insert(tk.END, "  供应商          文件名")
        self.file_listbox.insert(tk.END, "─" * 60)
        self.log("✓ 已清空文件列表")
    
    def remove_selected_file(self):
        """移除选中的文件"""
        selection = self.file_listbox.curselection()
        if not selection:
            messagebox.showinfo("提示", "请先选择要移除的文件")
            return
        
        # 从后往前删除，避免索引错乱
        for index in reversed(selection):
            if index >= 2:  # 跳过标题行
                self.file_listbox.delete(index)
                # 计算实际索引（减去标题行）
                actual_index = index - 2
                if actual_index < len(self.selected_files):
                    removed = self.selected_files.pop(actual_index)
                    self.log(f"✓ 已移除：{removed[1]} - {os.path.basename(removed[0])}")
    
    def identify_warehouse_from_filename(self, file_path):
        """从文件名识别分仓"""
        filename = os.path.basename(file_path)
        
        # 分仓关键词映射
        warehouse_keywords = {
            "北京": "北京",
            "沈阳": "沈阳",
            "南京": "南京"
        }
        
        for keyword, warehouse_name in warehouse_keywords.items():
            if keyword in filename:
                return warehouse_name
        
        return "未识别"
    
    def select_warehouse_files(self):
        """一次性选择多个分仓文件"""
        file_paths = filedialog.askopenfilenames(
            title="选择分仓文件（支持多选）",
            filetypes=[
                ("Excel文件", "*.xlsx *.xls"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_paths:
            self.log("正在识别分仓...")
            for file_path in file_paths:
                filename = os.path.basename(file_path)
                # 从文件名识别分仓
                warehouse = self.identify_warehouse_from_filename(file_path)
                self.warehouse_files.append((file_path, warehouse))
                
                # 更新列表显示
                display_name = filename[:42] + "..." if len(filename) > 42 else filename
                self.warehouse_listbox.insert(tk.END, f"  {warehouse:10}    {display_name}")
                
                self.log(f"✓ {filename} → {warehouse}")
    
    def clear_warehouse_files(self):
        """清空分仓文件列表"""
        self.warehouse_files = []
        self.warehouse_listbox.delete(0, tk.END)
        self.warehouse_listbox.insert(tk.END, "  分仓            文件名")
        self.warehouse_listbox.insert(tk.END, "─" * 60)
        self.log("✓ 已清空分仓文件列表")
    
    def select_product_file(self):
        """选择NC产品信息文件"""
        file_path = filedialog.askopenfilename(
            title="选择NC产品信息文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            self.product_file.set(file_path)
            self.load_product_data()
    
    def save_product_path(self):
        """保存产品文件路径配置"""
        config_file = os.path.join(os.path.dirname(__file__), "config_v3.txt")
        try:
            with open(config_file, 'w', encoding='utf-8') as f:
                f.write(f"product_file={self.product_file.get()}\n")
                f.write(f"output_dir={self.output_dir.get()}\n")
            self.log("✓ 配置已保存")
        except Exception as e:
            self.log(f"⚠ 保存配置失败：{str(e)}")
    
    def load_saved_config(self):
        """加载保存的配置"""
        config_file = os.path.join(os.path.dirname(__file__), "config_v3.txt")
        if os.path.exists(config_file):
            try:
                with open(config_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.startswith("product_file="):
                            path = line.strip().split("=", 1)[1]
                            if os.path.exists(path):
                                self.product_file.set(path)
                                self.load_product_data()
                        elif line.startswith("output_dir="):
                            path = line.strip().split("=", 1)[1]
                            if os.path.exists(path):
                                self.output_dir.set(path)
            except:
                pass
    
    def select_output_dir(self):
        """选择输出目录"""
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.output_dir.set(dir_path)
    
    def clear_all_files(self):
        """清空所有文件选择"""
        if self.order_type.get() == "factory":
            self.clear_file_list()
        else:
            self.clear_warehouse_files()
        self.progress_text.delete(1.0, "end")
        self.log("✓ 已清空")
    
    def open_output_dir(self):
        """打开输出目录"""
        output_path = self.output_dir.get()
        if os.path.exists(output_path):
            os.startfile(output_path)
        else:
            messagebox.showwarning("提示", "输出目录不存在")
    
    def show_changelog(self):
        """显示维护日志"""
        changelog_text = """【YaFo兼容采购助理 - 维护日志】

v3.5.37 (2026.06.22)
  - 修复：主界面“功能选择”功能项过多时横向单排显示不全的问题
  - 优化：功能选择改为3列多行网格布局，支持后续继续增加功能
  - 优化：顶部功能说明文字增加自动换行，窗口默认尺寸和最小尺寸小幅上调

v3.5.36 (2026.06.22)
  - 新增：集成销量预测与备货建议生成器作为功能选择子功能
  - 优化：销量预测工具支持CSV/Excel自动读取，销售订单和库存文件均可导入CSV或Excel
  - 优化：销量预测工具以子窗口方式嵌入主程序，避免创建第二个Tk主循环
  - 保留：原XML生成、供应商直发表格处理、新品定价、存货成本回填逻辑不变

v3.5.35 (2026.06.22)
  - 新增：集成存货成本回填生成器作为功能选择子功能
  - 新增：选择“存货成本回填”后自动弹出成本回填子模块窗口
  - 规则：后续每次新增功能时同步递增版本号
  - 保留：原工厂直发、分仓订单、供应商直发表格处理、新品上新定价逻辑不变

v3.5.34 (2026.06.22)
  - 集成：新品上新定价生成器作为功能选择子功能
  - 新增：选择“新品上新定价”后自动弹出定价子模块窗口
  - 保留：原工厂直发、分仓订单、供应商直发表格处理逻辑不变

v3.5.33 (2026.05.28)
  - 修复：界面标题/分区文字在高DPI或字体缩放环境下显示不完整
  - 调整：窗口允许拉伸，并增加最小窗口尺寸
  - 优化：卡片标题改为独立Label，避免LabelFrame边框裁切中文

v3.5.32 (2026.05.28)
  - 优化：主界面视觉样式升级，统一卡片、按钮、日志区样式
  - 调整：NC产品信息按钮文案由“CSV导出”改为“CSV导入”

v3.5.31 (2026.05.28)
  - 新增：将供应商直发表格处理放入功能选择，选择后自动弹出子模块窗口

v3.5.30 (2026.05.28)
  - 集成：供应商直发表格生成器作为子模块，可预处理并按供应商拆分导出Excel

v3.5.29 (2026.05.28)
  - 新增：支持“周结预定单统计”混合供应商直发汇总表
  - 新增：自动按“供应商名称”拆分征图/纳思达/博克等供应商，并分别生成工厂直发库XML
  - 新增：周结表支持字段：存货编码、主数量、单价、单据号、收货人、69码、存货名称

v3.5.28 (2026.05.28)
  - 集成：新增“CSV导入”按钮，可将NC导出的CSV转换为生成器可直接读取的NC产品信息Excel
  - 优化：导出后自动填入NC产品信息文件路径并加载产品数据
  - 优化：NC产品信息读取兼容Sheet3和首个工作表

v3.5.27 (2026.04.27)
  - 新增：界面增加"维护日志"按钮，点击查看版本更新历史

v3.5.26 (2026.04.27)
  - 修复：get_product_name_by_model使用regex=False避免正则特殊字符问题

v3.5.25 (2026.04.27)
  - 修复：征图分仓格式同时设置product_code和product_name_keyword列

v3.5.24 (2026.04.27)
  - 修复征图分仓格式检测：有"型号"+"产品名称"+"Unnamed"列时识别为征图分仓

v3.5.23 (2026.04.27)
  - 修复征图分仓格式识别和处理，添加get_product_by_name_simple函数

v3.5.22 (2026.04.27)
  - 修复德源分仓和征图分仓的区分：德源有"产品编号"无"型号"，征图有"型号"

v3.5.21 (2026.04.27)
  - 修复德源分仓文件识别问题，通过列名特征识别德源分仓格式

v3.5.20 (2026.04.27)
  - 修复德源分仓文件识别问题，添加迪研供应商识别

v3.5.19 (2026.04.27)
  - 修复德源供应商识别问题

v3.5.15 (2026.04.24)
  - 新增中山迪研电子有限公司(X15)供应商

v3.5.14 (2026.04.24)
  - 修复德源供应商识别问题

v3.5.13 (2026.04.24)
  - 修复绘强格式单价列匹配

v3.5.12 (2026.04.24)
  - 修复header=1导致存货编码匹配失败

v3.5.11 (2026.04.24)
  - 修复存货编码判断优先级

v3.5.10 (2026.04.24)
  - 修复绘强格式存货编码匹配

v3.5.9 (2026.04.24)
  - 重新修复三家供应商识别逻辑
  - 核心修改：优先检查"价格"列来判断是否是三家供应商格式
  - 三家供应商（征图/纳思达/博克）：有"价格"列，使用"存货编码"作为产品编码
  - 绘强格式：没有"价格"列，有"供应商"和"存货编码"列

v3.5.8 (2026.04.24)
  - 修复三家供应商（征图、纳思达、博克）识别问题
  - 增加三家供应商格式识别：同时有"存货编码"+"供应商"+"存货名称"+"价格"列

v3.5.7 (2026.04.24)
  - 修复三家供应商识别：优先检查"存货编码"+"供应商"+"存货名称"组合

v3.5.5 (2026.04.24)
  - 关键发现：绘强发的"型号"列实际存的是存货名称
  - 修改get_product_name_by_model函数，匹配NC的"存货名称"列

v3.5.4 (2026.04.24)
  - 关键bug修复：is_huiqiang_warehouse未添加到col_mapping返回值
  - 确保绘强发分仓格式正确走型号匹配分支

v3.5.3 (2026.04.24)
  - 最终修复绘强发分仓格式型号匹配问题

v3.5.2 (2026.04.24)
  - 修复绘强发分仓格式型号匹配问题
  - 新增get_product_name_by_model函数

v3.5.1 (2026.04.24)
  - 修复绘强发分仓格式识别问题
  - 新增绘强发分仓格式（用型号列）

v3.5 (2026.04.24)
  - 新增绘强发北京仓自动与NC存货清单比对功能
  - 生成比对结果Excel文件

v3.4 (2026.04.24)
  - 分仓订单支持多文件批量处理
  - 文件名自动识别分仓（北京/沈阳/南京）

v3.3 (2026.04.24)
  - 修复XML格式和汇总表格式
  - 汇总表改为"单据号汇总"和"详细明细"两个工作表

v3.2 (2026.04.24)
  - 修复XML格式，与NC系统完全兼容
  - 扩展header尝试范围，支持征图文件

v3.1 (2026.04.24)
  - 合并两个生成器，支持多选上传
  - 通过文件内容自动识别供应商"""
        
        # 创建日志窗口
        changelog_window = tk.Toplevel(self.root)
        changelog_window.title("维护日志")
        changelog_window.geometry("600x650")
        changelog_window.resizable(True, True)
        
        # 文本框
        text_frame = tk.Frame(changelog_window)
        text_frame.pack(fill="both", expand=True, padx=10, pady=10)
        
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side="right", fill="y")
        
        text_widget = tk.Text(text_frame, font=("微软雅黑", 10), 
                             yscrollcommand=scrollbar.set, wrap="word")
        text_widget.pack(side="left", fill="both", expand=True)
        scrollbar.config(command=text_widget.yview)
        
        text_widget.insert("1.0", changelog_text)
        text_widget.config(state="disabled")  # 只读
    
    def log(self, message):
        """记录日志"""
        timestamp = datetime.now().strftime('%H:%M:%S')
        self.progress_text.insert("end", f"[{timestamp}] {message}\n")
        self.progress_text.see("end")
        self.root.update()
    
    def load_product_data(self):
        """加载产品数据"""
        try:
            file_path = self.product_file.get()
            if not file_path:
                return

            excel_path = self.prepare_excel_for_pandas(file_path)

            # 优先读取原生成器约定的Sheet3；如果没有Sheet3，则自动读取第一个工作表
            try:
                self.product_df = pd.read_excel(excel_path, sheet_name='Sheet3', header=0)
                sheet_used = 'Sheet3'
            except Exception:
                xls = pd.ExcelFile(excel_path)
                if not xls.sheet_names:
                    raise ValueError("Excel文件中没有可读取的工作表")
                sheet_used = xls.sheet_names[0]
                self.product_df = pd.read_excel(excel_path, sheet_name=sheet_used, header=0)

            self.product_df.columns = self.product_df.columns.astype(str).str.strip()

            required_cols = ['存货编码', '存货名称']
            missing_cols = [c for c in required_cols if c not in self.product_df.columns]
            if missing_cols:
                raise ValueError(f"NC产品信息文件缺少必要列：{missing_cols}")

            # 条码列不是XML生成的强依赖字段，但保留兼容提示
            if '条码' not in self.product_df.columns:
                self.log("⚠ NC产品信息文件未发现'条码'列，不影响按存货编码/名称生成XML")

            # 去重：按存货编码和存货名称去重，保留第一个
            self.product_df = self.product_df.drop_duplicates(subset=['存货编码', '存货名称'], keep='first')
            self.log(f"✓ 加载产品数据成功，工作表：{sheet_used}，共 {len(self.product_df)} 条记录（去重后）")
        except Exception as e:
            self.product_df = None
            self.log(f"❌ 加载产品数据失败：{str(e)}")

    def export_nc_csv_to_excel(self):
        """将NC导出的CSV转换为生成器可直接读取的NC产品信息Excel"""
        csv_path = filedialog.askopenfilename(
            title="选择NC导出的CSV文件",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not csv_path:
            return

        output_dir = self.output_dir.get() or os.path.join(os.path.expanduser("~"), "Desktop")
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)

        try:
            self.log("=" * 50)
            self.log("开始从CSV导入NC产品信息Excel")
            self.log(f"CSV文件：{os.path.basename(csv_path)}")

            # 兼容常见NC导出编码：utf-8-sig / gbk / gb18030
            read_error = None
            df = None
            for encoding in ['utf-8-sig', 'gbk', 'gb18030']:
                try:
                    df = pd.read_csv(csv_path, encoding=encoding)
                    self.log(f"✓ CSV读取成功，编码：{encoding}，共 {len(df):,} 行")
                    break
                except Exception as e:
                    read_error = e
            if df is None:
                raise read_error

            df.columns = df.columns.astype(str).str.strip()
            cols_needed = ['存货编码', '存货名称', '条码']
            missing_cols = [c for c in cols_needed if c not in df.columns]
            if missing_cols:
                raise ValueError(f"CSV缺少必要列：{missing_cols}")

            output_df = df[cols_needed].copy()
            output_df = output_df.dropna(subset=['存货编码', '存货名称'], how='any')
            output_df = output_df.drop_duplicates(subset=['存货编码', '存货名称'], keep='first')

            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = 'Sheet3'

            # 保持原NC导出工具的横向重复格式，同时确保Sheet3可被主生成器直接读取
            headers = ['存货编码', '存货名称', '条码', '存货编码', '存货名称', '条码']
            ws.append(headers)
            for _, row in output_df.iterrows():
                ws.append([
                    row['存货编码'], row['存货名称'], row['条码'],
                    row['存货编码'], row['存货名称'], row['条码']
                ])

            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"NC存货信息_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            wb.save(filepath)

            self.product_file.set(filepath)
            self.load_product_data()
            self.save_product_path()

            self.log(f"✓ NC产品信息Excel导出成功：{filename}")
            self.log(f"✓ 已自动填入NC产品信息文件并加载：{len(self.product_df) if self.product_df is not None else 0:,} 条")
            messagebox.showinfo("完成", f"导出成功，并已自动加载为NC产品信息文件！\n{filepath}")
        except Exception as e:
            self.log(f"❌ CSV导入失败：{str(e)}")
            messagebox.showerror("错误", str(e))

    def get_product_name(self, product_code):
        """根据产品编码获取NC存货名称"""
        if self.product_df is None or pd.isna(product_code):
            return None
        
        product_code_str = str(product_code).strip()
        match = self.product_df[self.product_df['存货编码'].astype(str).str.strip() == product_code_str]
        if len(match) > 0:
            return match['存货名称'].values[0]
        return None
    
    def get_product_by_name(self, product_name_keyword):
        """根据产品名称关键词匹配NC产品（用于征图格式）"""
        if self.product_df is None or pd.isna(product_name_keyword):
            return None, None
        
        keyword = str(product_name_keyword).strip()
        # 在存货名称中搜索包含关键词的产品
        match = self.product_df[self.product_df['存货名称'].str.contains(keyword, na=False, regex=False)]
        if len(match) > 0:
            return match['存货名称'].values[0], match['存货编码'].values[0]
        return None, None
    
    def get_product_by_name_simple(self, product_name_keyword):
        """根据产品名称关键词匹配NC产品名称（返回单个值，用于征图分仓格式）"""
        if self.product_df is None or pd.isna(product_name_keyword):
            return None
        
        keyword = str(product_name_keyword).strip()
        # 在存货名称中搜索包含关键词的产品
        match = self.product_df[self.product_df['存货名称'].str.contains(keyword, na=False, regex=False)]
        if len(match) > 0:
            return match['存货名称'].values[0]
        return None
    
    def get_product_name_by_model(self, model_str):
        """根据型号获取NC存货名称（用于绘强发分仓格式）
        
        注意：绘强发的"型号"列实际存的是存货名称，不是真正的型号
        所以这个函数实际是用存货名称去匹配NC的"存货名称"列
        """
        if self.product_df is None or pd.isna(model_str):
            return None
        
        model_str_clean = str(model_str).strip()
        
        # 绘强发的"型号"列实际是存货名称，去匹配NC的"存货名称"列
        # 先精确匹配
        match = self.product_df[self.product_df['存货名称'].astype(str).str.strip() == model_str_clean]
        if len(match) > 0:
            return match['存货名称'].values[0]
        
        # 再模糊匹配（包含关系）- 使用regex=False避免正则特殊字符问题
        match = self.product_df[self.product_df['存货名称'].astype(str).str.contains(model_str_clean, na=False, regex=False)]
        if len(match) > 0:
            return match['存货名称'].values[0]
        
        return None
    
    def match_columns(self, columns, first_row_data=None):
        """智能匹配列名
        Args:
            columns: 列名列表
            first_row_data: 第一行数据字典（用于判断供应商列的值类型）
        """
        col_mapping = {}
        columns_list = list(columns)
        
        self.log(f"   文件列名：{columns_list[:10]}...")
        
        cols_lower = [str(c).strip() for c in columns_list]
        
        # 格式特征判断
        is_huiqiang = False
        is_zhengtu_new = False
        is_xingfa = False
        is_warehouse = False  # 分仓订单格式
        is_huiqiang_warehouse = False  # 绘强发分仓格式（用型号列）
        is_sanqian_supplier = False  # 三家供应商格式（征图、纳思达、博克）
        is_weekly_mixed_supplier = False  # 周结预定单统计表：按供应商名称拆分
        
        # =====================================================
        # 核心判断逻辑（按优先级顺序）
        # =====================================================
        
        # -1. 周结预定单统计表：同一文件中包含多个供应商，需要按“供应商名称”拆分
        # 典型列：仓库、单据号、存货编码、供应商名称、存货名称、供应商、69码、主数量、单价、收货人
        if all(c in cols_lower for c in ['供应商名称', '存货编码', '主数量', '单价', '单据号']):
            is_weekly_mixed_supplier = True
            self.log(f"   识别为周结预定单统计表（混合供应商直发汇总）")
        
        # 0. 首先检查第一列是否是供应商名称
        if first_row_data:
            first_cell = ''
            if '公司' in str(first_row_data):
                first_cell = str(first_row_data).strip()
            if '德源' in first_cell:
                is_warehouse = True
                self.log(f"   识别为德源格式")
            elif '兴发' in first_cell:
                is_xingfa = True
                self.log(f"   识别为兴发格式")
            elif '征图' in first_cell:
                is_zhengtu_new = True
                self.log(f"   识别为征图格式")
            elif '迪研' in first_cell:
                is_warehouse = True
                self.log(f"   识别为迪研格式")
        
        # 0.5. 优先检查德源分仓格式：同时有"产品编号"和"产品名称"列，且有"对方货号"列，但**没有"型号"列**
        # 德源分仓文件结构：序号, 产品编号, 产品名称, 单位, 数量, 单价, 金额, 对方货号, 备注
        # 征图分仓文件结构：序号, 产品名称, 型号, 数量, 单位, 单价, 金额, 对方货号, 备注
        if ('产品编号' in cols_lower and '产品名称' in cols_lower and '对方货号' in cols_lower and '型号' not in cols_lower):
            is_warehouse = True
            self.log(f"   识别为德源分仓格式（产品编号+产品名称，无型号列）")
        
        # 0.6. 征图分仓格式检测：有"型号"列和"产品名称"列
        # 征图发北京/南京：序号, 产品名称, 型号, 数量, 单位, 单价, 金额, Unnamed: 7(条码), 备注
        # 征图发沈阳：序号, 产品名称, 型号, 数量, 单位, 单价, 金额, 对方货号, 备注
        elif ('型号' in cols_lower and '产品名称' in cols_lower and 
              ('对方货号' in cols_lower or any('Unnamed' in c for c in cols_lower)) and '产品编号' not in cols_lower):
            is_zhengtu_new = True
            self.log(f"   识别为征图分仓格式（型号+产品名称）")
        
        # 1. 首先检查是否有"价格"列（注意：不是"单价"）→ 三家供应商格式
        # 三家供应商（征图、纳思达、博克）的特征：有"价格"列
        if '价格' in cols_lower and not is_xingfa and not is_warehouse and not is_zhengtu_new:
            is_sanqian_supplier = True
            self.log(f"   识别为三家供应商格式（征图/纳思达/博克）- 有价格列")
        
        # 2. 检查是否有"客户产品编码"列 → 兴发格式
        elif '客户产品编码' in cols_lower and not is_xingfa:
            is_xingfa = True
            self.log(f"   识别为兴发格式")
        
        # 3. 检查是否有"对方货号"列且没有"产品编号"列 → 征图旧格式
        elif '对方货号' in cols_lower and '产品编号' not in cols_lower and not is_zhengtu_new and not is_warehouse:
            is_zhengtu_new = True
            self.log(f"   识别为征图格式（对方货号匹配）")
        
        # 4. 检查是否有"客户品号"或"客户货号"列 → 分仓订单格式
        elif '客户品号' in cols_lower or '客户货号' in cols_lower:
            is_warehouse = True
            self.log(f"   识别为分仓订单格式")
        
        # 6. 检查是否有"供应商"和"存货编码"列
        # 需要根据供应商列的值判断是三家供应商还是绘强格式
        # 优先级提高：在有存货编码的情况下，优先用存货编码匹配
        elif not is_weekly_mixed_supplier and '供应商' in cols_lower and '存货编码' in cols_lower:
            # 通过供应商列第一行数据判断具体类型
            supplier_value = ''
            if first_row_data and '供应商' in first_row_data:
                supplier_value = str(first_row_data['供应商']).strip()
            
            # 三家供应商（征图/纳思达/博克）的供应商值特征
            # - 征图：值为"征图"、"X14"等
            # - 纳思达：值为"纳思达"等
            # - 博克：值为"博克"等
            if any(kw in supplier_value for kw in ['征图', '纳思达', '博克', 'X14']):
                is_sanqian_supplier = True
                self.log(f"   识别为三家供应商格式（征图/纳思达/博克）- 无价格列")
            else:
                # 绘强格式：供应商值为"绘强"、"X08"等
                is_huiqiang = True
                self.log(f"   识别为绘强格式（供应商列值：{supplier_value}）")
        
        # 6.5. 特殊情况：有"存货编码"但没有"供应商"列
        # 这种情况下（如header=1导致供应商列变成公司名），应该使用存货编码匹配
        elif not is_weekly_mixed_supplier and '存货编码' in cols_lower:
            is_huiqiang = True
            self.log(f"   识别为绘强格式（有存货编码，无供应商列，使用存货编码匹配）")
        
        # 7. 检查是否是绘强发分仓格式（用型号列）
        # 只在没有"存货编码"时才用型号列
        elif '型号' in cols_lower and '数量' in cols_lower:
            is_huiqiang_warehouse = True
            self.log(f"   识别为绘强发分仓格式（型号列）")
        
        # =====================================================
        # 匹配产品编码列
        # =====================================================
        
        # 周结预定单统计表：使用“存货编码”列，后续按“供应商名称”拆分
        if is_weekly_mixed_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '存货编码':
                    col_mapping['product_code'] = columns_list[i]
                    break
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '供应商名称':
                    col_mapping['supplier_name'] = columns_list[i]
                    break

        # 德源格式：使用"产品编号"列
        # 条件：first_cell包含"德源" 或 同时有"产品编号"+"产品名称"+"对方货号"，且没有"型号"列（德源分仓格式）
        is_deyuan_warehouse = ('德源' in first_cell or 
                              ('产品编号' in cols_lower and '产品名称' in cols_lower and 
                               '对方货号' in cols_lower and '型号' not in cols_lower))
        if is_deyuan_warehouse and not is_weekly_mixed_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '产品编号':
                    col_mapping['product_code'] = columns_list[i]
                    break
            # 德源格式使用产品名称匹配
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '产品名称':
                    col_mapping['product_name_field'] = columns_list[i]
                    break
        
        # 三家供应商格式：使用"存货编码"列（值为GA5723、XD0036、AE0093等）
        elif is_sanqian_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '存货编码':
                    col_mapping['product_code'] = columns_list[i]
                    break
        
        # 绘强格式：优先使用"存货编码"列（如果存在）
        # 兼容旧版本：如果不存在"存货编码"，则用"供应商"列（列名错位的情况）
        elif is_huiqiang:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '存货编码':
                    col_mapping['product_code'] = columns_list[i]
                    break
            # 如果没找到"存货编码"，才用"供应商"列（向后兼容）
            if 'product_code' not in col_mapping:
                for i, col in enumerate(columns_list):
                    col_str = str(col).strip()
                    if col_str == '供应商':
                        col_mapping['product_code'] = columns_list[i]
                        break
        
        # 征图格式：用"型号"列作为产品名称匹配
        elif is_zhengtu_new:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '型号':
                    col_mapping['product_name_field'] = columns_list[i]
                    break
        
        # 绘强发分仓格式：用"型号"列作为product_code
        elif is_huiqiang_warehouse:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '型号':
                    col_mapping['product_code'] = columns_list[i]
                    break
        
        # 其他格式：匹配客户产品编码、存货编码等
        else:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str in ['客户产品编码', '存货编码', '客户品号', '客户货号']:
                    col_mapping['product_code'] = columns_list[i]
                    break
        
        # =====================================================
        # 匹配数量列
        # =====================================================
        
        for i, col in enumerate(columns_list):
            col_str = str(col).strip()
            if col_str == '数量' or col_str == '主数量':
                col_mapping['quantity'] = columns_list[i]
                break
        
        # =====================================================
        # 匹配单价列
        # =====================================================
        
        # 周结预定单统计表：使用“单价”列
        if is_weekly_mixed_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '单价':
                    col_mapping['price'] = columns_list[i]
                    break

        # 三家供应商格式：使用"价格"列（注意：不是"单价"）
        elif is_sanqian_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '价格':
                    col_mapping['price'] = columns_list[i]
                    break
        
        # 绘强格式：优先使用"成本"列，如果没有则使用"单价"列
        elif is_huiqiang:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '成本':
                    col_mapping['price'] = columns_list[i]
                    break
            # 如果没找到"成本"，再找"单价"
            if 'price' not in col_mapping:
                for i, col in enumerate(columns_list):
                    col_str = str(col).strip()
                    if col_str == '单价':
                        col_mapping['price'] = columns_list[i]
                        break
        
        # 其他格式：使用"单价"列
        else:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '单价':
                    col_mapping['price'] = columns_list[i]
                    break
        
        # =====================================================
        # 匹配规格列
        # =====================================================
        
        for i, col in enumerate(columns_list):
            col_str = str(col).strip()
            if col_str in ['客户规格型号', '规格型号', '型号']:
                col_mapping['spec'] = columns_list[i]
                break
        
        # =====================================================
        # 匹配收件人列
        # =====================================================
        
        for i, col in enumerate(columns_list):
            col_str = str(col).strip()
            if col_str in ['收件人', '收货人']:
                col_mapping['receiver'] = columns_list[i]
                break
        
        # =====================================================
        # 匹配快递单号列
        # =====================================================
        
        for i, col in enumerate(columns_list):
            col_str = str(col).strip()
            if col_str in ['快递单号', '发货单号']:
                col_mapping['tracking'] = columns_list[i]
                break
        
        # =====================================================
        # 匹配单据号列
        # =====================================================
        
        # 周结预定单统计表：使用“单据号”列
        if is_weekly_mixed_supplier:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '单据号':
                    col_mapping['order_no'] = columns_list[i]
                    break

        # 三家供应商格式或兴发格式：使用"单据号"列
        elif is_sanqian_supplier or is_xingfa:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str in ['单据号', '订单号']:
                    col_mapping['order_no'] = columns_list[i]
                    break
        
        # 绘强格式：单据号在"业务员"列
        elif is_huiqiang:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str == '业务员':
                    col_mapping['order_no'] = columns_list[i]
                    break
        
        # 其他格式
        else:
            for i, col in enumerate(columns_list):
                col_str = str(col).strip()
                if col_str in ['单据号', '订单号']:
                    col_mapping['order_no'] = columns_list[i]
                    break
        
        col_mapping['is_zhengtu_new'] = is_zhengtu_new
        col_mapping['is_huiqiang'] = is_huiqiang
        col_mapping['is_huiqiang_warehouse'] = is_huiqiang_warehouse
        col_mapping['is_sanqian_supplier'] = is_sanqian_supplier  # 三家供应商格式标识
        col_mapping['is_weekly_mixed_supplier'] = is_weekly_mixed_supplier  # 周结预定单统计表标识
        return col_mapping
    
    def parse_excel_file(self, file_path, supplier_keyword):
        """解析Excel文件"""
        try:
            self.log(f"   识别为Excel文件")
            
            # 尝试不同header读取
            df = None
            header = None
            sheet_name = None
            
            # 读取所有工作表
            excel_path = self.prepare_excel_for_pandas(file_path)
            xls = pd.ExcelFile(excel_path)
            self.log(f"   工作表：{xls.sheet_names}")
            
            # 遍历所有工作表，寻找有效数据
            for sheet_idx, sheet in enumerate(xls.sheet_names):
                # 尝试不同的header值
                for h in [0, 1, 2, 3, 4, 5, 6]:
                    try:
                        test_df = pd.read_excel(excel_path, header=h, sheet_name=sheet)
                        test_df.columns = test_df.columns.astype(str).str.strip()
                        
                        # 检查是否包含关键列（增加"型号"列用于绘强发分仓）
                        if any('产品编码' in c or '存货编码' in c or '客户产品编码' in c or '客户品号' in c or '客户货号' in c or '对方货号' in c or '型号' in c for c in test_df.columns):
                            df = test_df
                            header = h
                            sheet_name = sheet
                            self.log(f"   识别为工作表'{sheet}'，header={h}")
                            break
                    except:
                        continue
                if df is not None:
                    break
            
            if df is None:
                self.log(f"   ❌ 无法识别文件格式")
                return None
            
            self.log(f"   原始数据行数：{len(df)}")
            self.log(f"   文件列名：{list(df.columns)[:10]}...")
            
            # 获取第一行数据（用于判断供应商列的值类型）
            first_row_data = df.iloc[0].to_dict() if len(df) > 0 else {}
            
            # 智能匹配列名（传入第一行数据用于判断供应商类型）
            col_mapping = self.match_columns(df.columns, first_row_data)
            
            is_zhengtu = col_mapping.get('is_zhengtu_new', False)
            
            # 检查必要列（征图用product_name_field，德源/其他用product_code）
            # 检查是否是德源格式
            # 德源格式：文件名包含"德源" 或者 列中有"产品编号"+"产品名称"+"型号列"（德源有"产品编号"，征图有"型号"）
            # 区分：德源没有"型号"列，征图有"型号"列
            is_deyuan = ('德源' in str(first_row_data) or 
                        ('产品编号' in df.columns and '产品名称' in df.columns and '型号' not in df.columns)) if first_row_data else False
            
            if is_zhengtu and not is_deyuan:
                if 'product_name_field' not in col_mapping or 'quantity' not in col_mapping:
                    self.log(f"   ❌ 缺少必要列（型号或数量）")
                    return None
            else:
                if 'product_code' not in col_mapping or 'quantity' not in col_mapping:
                    self.log(f"   ❌ 缺少必要列（产品编码或数量）")
                    return None
            
            # 提取数据
            # 德源格式使用product_code匹配，征图格式使用product_name_keyword
            if is_zhengtu and not is_deyuan:
                # 征图格式：用型号作为产品名称关键词
                result_df = df[[col_mapping['product_name_field'], col_mapping['quantity']]].copy()
                result_df.columns = ['product_name_keyword', 'quantity']
                # 同步复制到product_code列，方便后续处理
                result_df['product_code'] = result_df['product_name_keyword']
            else:
                # 德源和其他格式：使用product_code
                result_df = df[[col_mapping['product_code'], col_mapping['quantity']]].copy()
                result_df.columns = ['product_code', 'quantity']
                # 同步复制到product_name_keyword列
                result_df['product_name_keyword'] = result_df['product_code']
            
            # 添加单价
            if 'price' in col_mapping:
                result_df['price'] = df[col_mapping['price']]
            else:
                result_df['price'] = 0
            
            # 添加规格
            if 'spec' in col_mapping:
                result_df['spec'] = df[col_mapping['spec']]
            else:
                result_df['spec'] = ''
            
            # 添加收件人
            if 'receiver' in col_mapping:
                result_df['receiver'] = df[col_mapping['receiver']]
            else:
                result_df['receiver'] = ''
            
            # 添加快递单号
            if 'tracking' in col_mapping:
                result_df['tracking'] = df[col_mapping['tracking']]
            else:
                result_df['tracking'] = ''
            
            # 添加单据号
            if 'order_no' in col_mapping:
                result_df['order_no'] = df[col_mapping['order_no']]
            else:
                result_df['order_no'] = ''

            # 添加供应商名称（周结预定单统计表用于自动拆分供应商）
            if 'supplier_name' in col_mapping:
                result_df['supplier_name'] = df[col_mapping['supplier_name']].astype(str).str.strip()
            else:
                result_df['supplier_name'] = supplier_keyword
            
            result_df['is_zhengtu_new'] = col_mapping.get('is_zhengtu_new', False)
            result_df['is_huiqiang_warehouse'] = col_mapping.get('is_huiqiang_warehouse', False)
            result_df['is_sanqian_supplier'] = col_mapping.get('is_sanqian_supplier', False)  # 三家供应商格式标识
            result_df['is_weekly_mixed_supplier'] = col_mapping.get('is_weekly_mixed_supplier', False)  # 周结预定单统计表标识
            
            # 清洗数据
            if is_zhengtu:
                # 征图格式：清洗产品名称关键词
                result_df = result_df.dropna(subset=['product_name_keyword'])
                result_df['product_name_keyword'] = result_df['product_name_keyword'].astype(str).str.strip()
                result_df = result_df[result_df['product_name_keyword'] != '']
                result_df = result_df[~result_df['product_name_keyword'].isin(['合计', '复核：', '客户确认：'])]
            else:
                result_df = result_df.dropna(subset=['product_code'])
                result_df['product_code'] = result_df['product_code'].astype(str).str.strip()
                result_df = result_df[result_df['product_code'] != '']
            
            result_df['quantity'] = pd.to_numeric(result_df['quantity'], errors='coerce')
            result_df = result_df.dropna(subset=['quantity'])
            
            self.log(f"   有效数据行数：{len(result_df)}")
            
            return result_df
            
        except Exception as e:
            self.log(f"   ❌ 解析失败：{str(e)}")
            return None
    
    def generate_xml(self, supplier_name, nc_name, order_data, output_dir, order_date):
        """生成XML文件（按参考XML格式）"""
        try:
            # 创建根元素（按参考XML格式）
            root = ET.Element("ufinterface")
            root.set("billtype", "21")
            root.set("isexchange", "Y")
            root.set("operation", "req")
            root.set("proc", "add")
            root.set("receiver", "CG21")
            root.set("replace", "Y")
            root.set("roottag", "voucher")
            root.set("sender", "CG21")
            
            # 创建bill节点
            bill = ET.SubElement(root, "bill")
            
            # 创建billhead节点
            billhead = ET.SubElement(bill, "billhead")
            
            # 订单头信息（按参考XML顺序）
            ET.SubElement(billhead, "cbiztype").text = ORDER_HEADER["cbiztype"]
            ET.SubElement(billhead, "ctermprotocolid").text = ORDER_HEADER["ctermprotocolid"]
            ET.SubElement(billhead, "vdef2").text = ORDER_HEADER["vdef2"]
            ET.SubElement(billhead, "vdef3").text = ORDER_HEADER["vdef3"]
            ET.SubElement(billhead, "vdef20").text = ORDER_HEADER["vdef20"]
            ET.SubElement(billhead, "cdeptid").text = ORDER_HEADER["cdeptid"]
            ET.SubElement(billhead, "cemployeeid").text = ORDER_HEADER["cemployeeid"]
            ET.SubElement(billhead, "coperator").text = ORDER_HEADER["coperator"]
            ET.SubElement(billhead, "cuserid").text = ORDER_HEADER["cuserid"]
            # 使用自闭合标签 <vmemo/>
            ET.SubElement(billhead, "vmemo")
            ET.SubElement(billhead, "cvendormangid").text = nc_name
            ET.SubElement(billhead, "cgiveinvoicevendor").text = nc_name
            ET.SubElement(billhead, "dorderdate").text = order_date
            
            # 创建billbody节点
            billbody = ET.SubElement(bill, "billbody")
            
            # 计算预计到货日期（订单日期+4天）
            from datetime import timedelta
            order_dt = datetime.strptime(order_date, '%Y-%m-%d')
            arrive_dt = order_dt + timedelta(days=4)
            arrive_date = arrive_dt.strftime('%Y/%m/%d')
            
            # 添加明细行
            for idx, row in order_data.iterrows():
                entry = ET.SubElement(billbody, "entry")
                
                # 获取产品名称
                product_name = self.get_product_name(row['product_code'])
                if not product_name:
                    product_name = ""
                
                ET.SubElement(entry, "dplanarrvdate").text = arrive_date
                ET.SubElement(entry, "cwarehouseid").text = "工厂直发库"
                ET.SubElement(entry, "cmangid").text = product_name
                ET.SubElement(entry, "nordernum").text = str(int(row['quantity']))
                ET.SubElement(entry, "norgnettaxprice").text = str(row.get('price', 0))
            
            # 格式化XML
            xml_str = ET.tostring(root, encoding='unicode')
            dom = minidom.parseString(xml_str)
            pretty_xml = dom.toprettyxml(indent="    ")
            
            # 移除空行，保留XML声明
            lines = [line for line in pretty_xml.split('\n') if line.strip()]
            pretty_xml = '\n'.join(lines)
            
            # 替换XML声明为 <?xml version="1.0" ?>
            pretty_xml = pretty_xml.replace('<?xml version="1.0" ?>', '<?xml version="1.0" ?>')
            
            # 生成文件名
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"采购订单导入_{supplier_name}_{timestamp}.xml"
            filepath = os.path.join(output_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(pretty_xml)
            
            return filename
            
        except Exception as e:
            self.log(f"   ❌ XML生成失败：{str(e)}")
            return None
    
    def run_process(self):
        """执行处理"""
        # 检查产品数据
        if self.product_df is None:
            messagebox.showwarning("提示", "请先选择NC产品信息文件")
            return
        
        output_dir = self.output_dir.get()
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        
        order_date = datetime.now().strftime('%Y-%m-%d')
        
        if self.order_type.get() == "factory":
            self.run_factory_process(output_dir, order_date)
        else:
            self.run_warehouse_process(output_dir, order_date)
    
    def get_supplier_nc_name(self, supplier_keyword):
        """根据供应商简称获取NC供应商名称"""
        if supplier_keyword is None:
            return None
        supplier_keyword = str(supplier_keyword).strip()
        for name, nc in SUPPLIERS:
            if name == supplier_keyword:
                return nc
        return None

    def normalize_supplier_name(self, supplier_name):
        """标准化供应商名称，兼容周结表中的简称、编码和NC全称"""
        if supplier_name is None or pd.isna(supplier_name):
            return ''
        text = str(supplier_name).strip()
        if not text or text.lower() == 'nan':
            return ''

        code_map = {
            'S02': '博克',
            'S03': '纳思达',
            'S04': '征图',
            'X14': '征图',
            'X15': '迪研',
            'X21': '德源',
            'X53': '联合天润',
            'X08': '绘强'
        }
        if text in code_map:
            return code_map[text]

        for keyword, supplier in SUPPLIER_KEYWORDS.items():
            if keyword and keyword in text:
                return supplier

        for name, nc in SUPPLIERS:
            if name in text or text in nc:
                return name

        return text

    def process_matched_factory_order(self, supplier_keyword, nc_name, order_data, output_dir, order_date):
        """匹配NC产品并生成单个供应商的工厂直发XML"""
        self.log(f"   正在匹配NC产品名称...")
        is_zhengtu = order_data['is_zhengtu_new'].iloc[0] if len(order_data) > 0 and 'is_zhengtu_new' in order_data.columns else False

        if is_zhengtu:
            # 征图特殊格式：用型号/名称关键词匹配产品名称
            matched_rows = []
            for idx, row in order_data.iterrows():
                keyword = row.get('product_name_keyword', '')
                product_name, product_code = self.get_product_by_name(keyword)
                if product_name:
                    row['product_name'] = product_name
                    row['product_code'] = product_code
                    matched_rows.append(row)
                else:
                    self.log(f"   ⚠ 未匹配：{keyword}")
            if matched_rows:
                matched = pd.DataFrame(matched_rows)
            else:
                matched = pd.DataFrame()
        else:
            order_data = order_data.copy()
            order_data['product_name'] = order_data['product_code'].apply(self.get_product_name)
            matched = order_data[order_data['product_name'].notna()]

            # 输出少量未匹配编码，便于排查NC清单是否过旧
            unmatched = order_data[order_data['product_name'].isna()]
            if len(unmatched) > 0:
                sample_codes = unmatched['product_code'].astype(str).drop_duplicates().head(10).tolist()
                self.log(f"   ⚠ 未匹配编码 {len(unmatched)} 行，示例：{', '.join(sample_codes)}")

        self.log(f"   ✓ 匹配成功：{len(matched)} 条")

        if len(matched) == 0:
            self.log(f"   ❌ 没有匹配的产品")
            return None

        # 添加供应商NC名称
        matched = matched.copy()
        matched['supplier'] = nc_name

        # 生成XML
        xml_filename = self.generate_xml(supplier_keyword, nc_name, matched, output_dir, order_date)
        if xml_filename:
            self.log(f"   ✓ XML生成成功：{xml_filename}")
            self.log(f"   ✓ 包含 {len(matched)} 条明细")
            return {
                'supplier': supplier_keyword,
                'nc_name': nc_name,
                'data': matched
            }
        return None

    def run_factory_process(self, output_dir, order_date):
        """工厂直发订单处理"""
        if not self.selected_files:
            messagebox.showwarning("提示", "请先选择供应商文件")
            return
        
        self.log("=" * 50)
        self.log("🚀 开始处理工厂直发订单")
        self.log("=" * 50)
        self.log(f"📁 已选择 {len(self.selected_files)} 个供应商文件")
        self.log(f"📋 NC产品数据：{len(self.product_df)} 条")
        
        processed_suppliers = 0
        generated_xmls = 0
        all_orders = []
        
        for file_path, supplier_keyword in self.selected_files:
            self.log("=" * 40)
            self.log(f"📄 处理 {supplier_keyword} - {os.path.basename(file_path)}")
            self.log("=" * 40)
            
            # 解析文件
            if file_path.lower().endswith(('.png', '.jpg', '.jpeg')):
                self.log(f"   识别为图片文件，OCR功能暂未启用")
                continue
            else:
                order_data = self.parse_excel_file(file_path, supplier_keyword)
            
            if order_data is None or len(order_data) == 0:
                self.log(f"   ❌ 没有有效数据")
                continue

            # 周结预定单统计表：按“供应商名称”拆分，一份汇总表生成多个供应商XML
            is_weekly_mixed = (
                'is_weekly_mixed_supplier' in order_data.columns and
                bool(order_data['is_weekly_mixed_supplier'].iloc[0])
            )
            if is_weekly_mixed:
                if 'supplier_name' not in order_data.columns:
                    self.log("   ❌ 周结表缺少供应商名称列，无法按供应商拆分")
                    continue

                order_data = order_data.copy()
                order_data['supplier_group'] = order_data['supplier_name'].apply(self.normalize_supplier_name)
                order_data = order_data[order_data['supplier_group'] != '']

                supplier_groups = sorted(order_data['supplier_group'].dropna().unique().tolist())
                self.log(f"   识别到混合供应商：{', '.join(supplier_groups)}")

                for supplier_group, supplier_df in order_data.groupby('supplier_group'):
                    nc_name = self.get_supplier_nc_name(supplier_group)
                    if not nc_name:
                        self.log(f"   ❌ 未找到供应商NC名称：{supplier_group}，已跳过")
                        continue

                    self.log("-" * 30)
                    self.log(f"   拆分处理供应商：{supplier_group}，明细 {len(supplier_df)} 行")
                    result = self.process_matched_factory_order(
                        supplier_group, nc_name, supplier_df.copy(), output_dir, order_date
                    )
                    if result:
                        generated_xmls += 1
                        processed_suppliers += 1
                        all_orders.append(result)
                continue

            # 普通工厂直发文件：沿用原逻辑，一份文件对应一个供应商XML
            nc_name = self.get_supplier_nc_name(supplier_keyword)
            if not nc_name:
                self.log(f"   ❌ 未找到供应商NC名称：{supplier_keyword}")
                continue

            result = self.process_matched_factory_order(
                supplier_keyword, nc_name, order_data, output_dir, order_date
            )
            if result:
                generated_xmls += 1
                processed_suppliers += 1
                all_orders.append(result)
        
        # 生成汇总表
        if all_orders:
            self.generate_summary_excel(all_orders, output_dir, order_date)
        
        self.log("=" * 50)
        self.log("✅ 处理完成！")
        self.log("=" * 50)
        self.log(f"- 处理供应商数：{processed_suppliers}")
        self.log(f"- 生成XML文件数：{generated_xmls}")
        self.log(f"- 输出目录：{output_dir}")
    
    def generate_summary_excel(self, all_orders, output_dir, order_date):
        """生成汇总Excel（按参考格式）"""
        try:
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            filename = f"工厂直发汇总_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            wb = Workbook()
            
            # ===== 工作表1：单据号汇总（按供应商分组） =====
            ws_summary = wb.active
            ws_summary.title = '单据号汇总'
            
            row_idx = 1
            for order in all_orders:
                df = order['data']
                nc_name = order['nc_name']
                
                # 第1行：供应商名称
                ws_summary.cell(row=row_idx, column=1, value=nc_name)
                row_idx += 1
                # 第2行：空行
                row_idx += 1
                # 第3行：列名
                ws_summary.cell(row=row_idx, column=1, value='单据号')
                ws_summary.cell(row=row_idx, column=2, value='完成情况')
                ws_summary.cell(row=row_idx, column=3, value='备注')
                row_idx += 1
                
                # 收集该供应商的单据号
                if 'order_no' in df.columns and df['order_no'].notna().any():
                    grouped = df.groupby('order_no', dropna=False)
                    for order_no, group in grouped:
                        order_no_str = str(order_no) if not pd.isna(order_no) else ''
                        if order_no_str and order_no_str != '':
                            ws_summary.cell(row=row_idx, column=1, value=order_no_str)
                            ws_summary.cell(row=row_idx, column=2, value=1)
                            row_idx += 1
                
                # 供应商之间空一行
                row_idx += 1
            
            # ===== 工作表2：详细明细 =====
            ws_detail = wb.create_sheet('详细明细')
            
            # 第1行：列名
            detail_headers = ['供应商', '单据号', '客户产品编码', 'NC存货名称', '客户规格型号', '数量', '单价', '金额', '收件人', '快递单号']
            for col_idx, header in enumerate(detail_headers, 1):
                ws_detail.cell(row=1, column=col_idx, value=header)
            
            # 收集所有详细数据
            row_idx = 2
            for order in all_orders:
                df = order['data'].copy()
                nc_name = order['nc_name']
                
                for idx, row in df.iterrows():
                    # 计算金额
                    quantity = row.get('quantity', 0)
                    price = row.get('price', 0)
                    try:
                        amount = float(quantity) * float(price)
                    except:
                        amount = 0
                    
                    ws_detail.cell(row=row_idx, column=1, value=nc_name)
                    ws_detail.cell(row=row_idx, column=2, value=row.get('order_no', ''))
                    ws_detail.cell(row=row_idx, column=3, value=row.get('product_code', ''))
                    ws_detail.cell(row=row_idx, column=4, value=row.get('product_name', ''))
                    ws_detail.cell(row=row_idx, column=5, value=row.get('spec', ''))
                    ws_detail.cell(row=row_idx, column=6, value=quantity)
                    ws_detail.cell(row=row_idx, column=7, value=price)
                    ws_detail.cell(row=row_idx, column=8, value=amount)
                    ws_detail.cell(row=row_idx, column=9, value=row.get('receiver', ''))
                    ws_detail.cell(row=row_idx, column=10, value=row.get('tracking', ''))
                    row_idx += 1
            
            # ===== 工作表3：供应商原始数据 =====
            for order in all_orders:
                df = order['data'].copy()
                nc_name = order['nc_name']
                
                # 工作表名称（截取前30字符，替换非法字符）
                sheet_name = nc_name[:30] if len(nc_name) > 30 else nc_name
                # 替换Excel工作表名非法字符: \ / ? * [ ]
                for char in ['\\', '/', '?', '*', '[', ']']:
                    sheet_name = sheet_name.replace(char, '-')
                ws_supplier = wb.create_sheet(sheet_name)
                
                # 第1行：供应商名称
                ws_supplier.cell(row=1, column=1, value=nc_name)
                # 第2行：发票
                ws_supplier.cell(row=2, column=1, value='发票')
                # 第3行：列名
                supplier_headers = ['单据号', '客户产品编码', '客户规格型号', '数量', '单价', '金额', '客户产品条码', '收件人', '件数', '快递单号']
                for col_idx, header in enumerate(supplier_headers, 1):
                    ws_supplier.cell(row=3, column=col_idx, value=header)
                
                # 写入数据
                row_idx = 4
                for idx, row in df.iterrows():
                    # 计算金额
                    quantity = row.get('quantity', 0)
                    price = row.get('price', 0)
                    try:
                        amount = float(quantity) * float(price)
                    except:
                        amount = 0
                    
                    ws_supplier.cell(row=row_idx, column=1, value=row.get('order_no', ''))
                    ws_supplier.cell(row=row_idx, column=2, value=row.get('product_code', ''))
                    ws_supplier.cell(row=row_idx, column=3, value=row.get('spec', ''))
                    ws_supplier.cell(row=row_idx, column=4, value=quantity)
                    ws_supplier.cell(row=row_idx, column=5, value=price)
                    ws_supplier.cell(row=row_idx, column=6, value=amount)
                    ws_supplier.cell(row=row_idx, column=7, value=row.get('barcode', ''))
                    ws_supplier.cell(row=row_idx, column=8, value=row.get('receiver', ''))
                    ws_supplier.cell(row=row_idx, column=9, value='1件')  # 件数默认1件
                    ws_supplier.cell(row=row_idx, column=10, value=row.get('tracking', ''))
                    row_idx += 1
            
            # 保存文件
            wb.save(filepath)
            self.log(f"📊 汇总表生成成功：{filename}")
            
        except Exception as e:
            self.log(f"⚠ 汇总表生成失败：{str(e)}")
            import traceback
            self.log(f"错误详情：{traceback.format_exc()}")
    
    def run_warehouse_process(self, output_dir, order_date):
        """分仓订单处理（支持批量多文件）"""
        if not self.warehouse_files:
            messagebox.showwarning("提示", "请先选择分仓文件")
            return
        
        self.log("=" * 50)
        self.log("🚀 开始处理分仓订单")
        self.log("=" * 50)
        self.log(f"📁 已选择 {len(self.warehouse_files)} 个分仓文件")
        self.log(f"📋 NC产品数据：{len(self.product_df)} 条")
        
        warehouse_mapping = dict(WAREHOUSES)
        processed_count = 0
        generated_xmls = 0
        
        for file_path, warehouse in self.warehouse_files:
            self.log("=" * 40)
            self.log(f"📄 处理 {warehouse}")
            self.log("=" * 40)
            
            if warehouse == "未识别":
                self.log(f"   ❌ 无法识别分仓，跳过")
                continue
            
            warehouse_nc = warehouse_mapping.get(warehouse, "北京仓")
            self.log(f"   分仓映射：{warehouse} -> {warehouse_nc}")
            
            # 获取文件名作为备注
            filename = os.path.basename(file_path)
            filename_no_ext = os.path.splitext(filename)[0]
            
            # 解析Excel
            order_data = self.parse_excel_file(file_path, warehouse)
            if order_data is None or len(order_data) == 0:
                self.log(f"   ❌ 没有有效数据")
                continue
            
            # 识别供应商（从文件内容）
            supplier_name = self.identify_supplier_from_content(file_path)
            nc_name = None
            for name, nc in SUPPLIERS:
                if name == supplier_name:
                    nc_name = nc
                    break
            
            if not nc_name:
                self.log(f"   ⚠ 未识别供应商，使用默认值")
                nc_name = "中山澳兴发科技有限公司/中山市兴发电子科技有限公司"
            
            # 特殊处理：绘强发北京仓 - 和NC存货清单比对
            if supplier_name == "绘强" and warehouse == "北京":
                self.log(f"   【绘强发北京仓】正在与NC存货清单比对...")
                self.compare_with_nc(order_data, supplier_name, warehouse)
            
            # 匹配产品名称
            is_huiqiang_warehouse = order_data.get('is_huiqiang_warehouse', pd.Series([False]*len(order_data))).iloc[0] if len(order_data) > 0 else False
            is_zhengtu_warehouse = order_data.get('is_zhengtu_new', pd.Series([False]*len(order_data))).iloc[0] if len(order_data) > 0 else False
            
            if is_huiqiang_warehouse:
                # 绘强发分仓格式：用型号匹配NC产品
                self.log(f"   正在匹配NC产品名称（绘强发-型号匹配）...")
                order_data['product_name'] = order_data['product_code'].apply(self.get_product_name_by_model)
                matched = order_data[order_data['product_name'].notna()]
                self.log(f"   ✓ 匹配成功：{len(matched)} 条")
            elif is_zhengtu_warehouse:
                # 征图分仓格式：用型号列（实际是产品名称关键词）匹配
                self.log(f"   正在匹配NC产品名称（征图分仓-型号匹配）...")
                order_data['product_name'] = order_data['product_name_keyword'].apply(self.get_product_by_name_simple)
                matched = order_data[order_data['product_name'].notna()]
                self.log(f"   ✓ 匹配成功：{len(matched)} 条")
            else:
                # 其他格式：用product_code匹配
                self.log(f"   正在匹配NC产品名称...")
                order_data['product_name'] = order_data['product_code'].apply(self.get_product_name)
                matched = order_data[order_data['product_name'].notna()]
                self.log(f"   ✓ 匹配成功：{len(matched)} 条")
            
            if len(matched) == 0:
                self.log(f"   ❌ 没有匹配的产品")
                continue
            
            # 生成分仓XML
            xml_filename = self.generate_warehouse_xml(supplier_name, nc_name, matched, output_dir, order_date, warehouse_nc, filename_no_ext)
            
            if xml_filename:
                self.log(f"   ✓ XML生成成功：{xml_filename}")
                self.log(f"   ✓ 包含 {len(matched)} 条明细")
                generated_xmls += 1
            
            processed_count += 1
        
        self.log("=" * 50)
        self.log("✅ 处理完成！")
        self.log("=" * 50)
        self.log(f"- 处理分仓文件数：{processed_count}")
        self.log(f"- 生成XML文件数：{generated_xmls}")
        self.log(f"- 输出目录：{output_dir}")
    
    def compare_with_nc(self, order_data, supplier_name, warehouse):
        """和NC存货清单比对型号名称"""
        if self.product_df is None:
            self.log(f"   ⚠ NC产品数据未加载，跳过比对")
            return
        
        try:
            matched = []
            unmatched = []
            
            # 提取发货清单中的型号
            order_models = order_data['product_code'].dropna().tolist()
            
            # 准备NC数据（从Sheet3获取）
            nc_products = []
            for idx, row in self.product_df.iterrows():
                nc_products.append({
                    '存货编码': row['存货编码'],
                    '存货名称': row['存货名称']
                })
            
            # 比对逻辑
            for order_model in order_models:
                match = None
                for nc in nc_products:
                    # 完全匹配
                    if order_model == nc['存货名称']:
                        match = nc
                        break
                    # 包含匹配
                    elif nc['存货名称'] in order_model or order_model in nc['存货名称']:
                        match = nc
                        break
                
                if match:
                    matched.append({
                        '发货清单型号': order_model,
                        'NC存货编码': match['存货编码'],
                        'NC存货名称': match['存货名称']
                    })
                else:
                    unmatched.append({
                        '发货清单型号': order_model
                    })
            
            # 输出结果
            match_rate = len(matched) / len(order_models) * 100 if order_models else 0
            self.log(f"   ✓ 型号比对完成：{len(matched)}/{len(order_models)} 匹配 ({match_rate:.1f}%)")
            
            # 保存比对结果
            if matched or unmatched:
                from datetime import datetime
                timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
                output_file = f"发货清单_NC比对_{supplier_name}_{warehouse}_{timestamp}.xlsx"
                
                try:
                    df_matched = pd.DataFrame(matched)
                    df_unmatched = pd.DataFrame(unmatched)
                    
                    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                        df_matched.to_excel(writer, sheet_name='匹配成功', index=False)
                        df_unmatched.to_excel(writer, sheet_name='未匹配', index=False)
                    
                    self.log(f"   ✓ 比对结果已保存：{output_file}")
                except Exception as e:
                    self.log(f"   ⚠ 保存比对结果失败：{str(e)}")
        
        except Exception as e:
            self.log(f"   ⚠ 比对过程出错：{str(e)}")

    def generate_warehouse_xml(self, supplier_name, nc_name, order_data, output_dir, order_date, warehouse_nc, memo):
        """生成分仓订单XML（按参考XML格式）"""
        try:
            root = ET.Element("ufinterface")
            root.set("billtype", "21")
            root.set("isexchange", "Y")
            root.set("operation", "req")
            root.set("proc", "add")
            root.set("receiver", "CG21")
            root.set("replace", "Y")
            root.set("roottag", "voucher")
            root.set("sender", "CG21")
            
            bill = ET.SubElement(root, "bill")
            
            billhead = ET.SubElement(bill, "billhead")
            
            ET.SubElement(billhead, "cbiztype").text = ORDER_HEADER["cbiztype"]
            ET.SubElement(billhead, "ctermprotocolid").text = ORDER_HEADER["ctermprotocolid"]
            ET.SubElement(billhead, "vdef2").text = ORDER_HEADER["vdef2"]
            ET.SubElement(billhead, "vdef3").text = ORDER_HEADER["vdef3"]
            ET.SubElement(billhead, "vdef20").text = ORDER_HEADER["vdef20"]
            ET.SubElement(billhead, "cdeptid").text = ORDER_HEADER["cdeptid"]
            ET.SubElement(billhead, "cemployeeid").text = ORDER_HEADER["cemployeeid"]
            ET.SubElement(billhead, "coperator").text = ORDER_HEADER["coperator"]
            ET.SubElement(billhead, "cuserid").text = ORDER_HEADER["cuserid"]
            ET.SubElement(billhead, "vmemo").text = memo
            ET.SubElement(billhead, "cvendormangid").text = nc_name
            ET.SubElement(billhead, "cgiveinvoicevendor").text = nc_name
            ET.SubElement(billhead, "dorderdate").text = order_date
            
            billbody = ET.SubElement(bill, "billbody")
            
            # 计算预计到货日期（订单日期+4天）
            from datetime import timedelta
            order_dt = datetime.strptime(order_date, '%Y-%m-%d')
            arrive_dt = order_dt + timedelta(days=4)
            arrive_date = arrive_dt.strftime('%Y/%m/%d')
            
            for idx, row in order_data.iterrows():
                entry = ET.SubElement(billbody, "entry")
                
                product_name = row.get('product_name', '')
                
                ET.SubElement(entry, "dplanarrvdate").text = arrive_date
                ET.SubElement(entry, "cwarehouseid").text = warehouse_nc
                ET.SubElement(entry, "cmangid").text = product_name
                ET.SubElement(entry, "nordernum").text = str(int(row['quantity']))
                ET.SubElement(entry, "norgnettaxprice").text = str(row.get('price', 0))
            
            xml_str = ET.tostring(root, encoding='unicode')
            dom = minidom.parseString(xml_str)
            pretty_xml = dom.toprettyxml(indent="    ")
            
            lines = [line for line in pretty_xml.split('\n') if line.strip()]
            pretty_xml = '\n'.join(lines[1:])
            
            # 替换XML声明
            pretty_xml = '<?xml version="1.0" ?>\n' + pretty_xml
            
            # 文件名包含：分仓名称_供应商_时间戳
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            # 从warehouse_nc提取分仓名称（如"北京仓"->"北京"）
            warehouse_short = warehouse_nc.replace("仓", "")
            # 从nc_name提取供应商简称
            supplier_short = supplier_name if supplier_name else "未知供应商"
            filename = f"分仓订单_{warehouse_short}_{supplier_short}_{timestamp}.xml"
            filepath = os.path.join(output_dir, filename)
            
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(pretty_xml)
            
            return filename
            
        except Exception as e:
            self.log(f"❌ XML生成失败：{str(e)}")
            return None



# ===== 子模块：供应商直发表格生成器 =====
SUPPLIER_DIRECT_NAME_MAPPING = {
    'X04': '兴发',
    'X08': '绘强',
    'X14': '征图',
    'X64': '征图',
    'S02': '博克',
    'S02/S03': '博克',
    'S02S03': '博克',
    'S03': '纳思达',
    'S03/S04': '纳思达',
    'S03S04': '纳思达',
    'X21': '德源',
    'X15': '迪研',
    'X02': '华人',
    '华人': '华人',
}

class SupplierDirectTableCore:
    """供应商直发表格生成器子模块核心逻辑"""

    def __init__(self, parent_app=None):
        self.parent_app = parent_app
        self.df = None
        self.df_processed = None
        self.suppliers = []
        self.latest_date = None
        self.filtered_count = 0

    def _read_excel(self, file_path):
        excel_path = file_path
        if self.parent_app is not None and hasattr(self.parent_app, 'prepare_excel_for_pandas'):
            excel_path = self.parent_app.prepare_excel_for_pandas(file_path)
        return pd.read_excel(excel_path)

    def load_data(self, file_path):
        try:
            self.df = self._read_excel(file_path)
            self.df.columns = self.df.columns.astype(str).str.strip()
            if 'Unnamed: 0' in self.df.columns:
                self.df = self.df.drop(columns=['Unnamed: 0'])
            return True, f"成功加载 {len(self.df):,} 条记录"
        except Exception as e:
            return False, f"加载失败: {str(e)}"

    def preprocess_data(self, remove_price=True):
        if self.df is None:
            return False, "请先加载数据"

        try:
            df = self.df.copy()
            if '审批日期' in df.columns:
                date_series = df['审批日期'].dropna()
                if len(date_series) > 0:
                    # 兼容字符串日期和Excel日期类型
                    parsed = pd.to_datetime(date_series, errors='coerce')
                    if parsed.notna().any():
                        latest_ts = parsed.max()
                        all_parsed = pd.to_datetime(df['审批日期'], errors='coerce')
                        original_count = len(df)
                        df = df[all_parsed == latest_ts]
                        self.latest_date = latest_ts.strftime('%Y-%m-%d')
                        self.filtered_count = original_count - len(df)
                    else:
                        self.latest_date = max(date_series)
                        original_count = len(df)
                        df = df[df['审批日期'] == self.latest_date]
                        self.filtered_count = original_count - len(df)

            if remove_price and '单价' in df.columns:
                df = df.drop(columns=['单价'])

            if '供应商' in df.columns:
                supplier_code = df['供应商'].astype(str).str.strip()
                df['供应商名称'] = supplier_code.map(SUPPLIER_DIRECT_NAME_MAPPING).fillna(supplier_code)
            elif '供应商名称' not in df.columns:
                return False, "缺少'供应商'或'供应商名称'列，无法按供应商拆分"

            cols = df.columns.tolist()
            if '品牌' in cols and '存货编码' in cols and '供应商名称' in cols:
                cols.remove('供应商名称')
                brand_idx = cols.index('品牌')
                cols.insert(brand_idx + 1, '供应商名称')
                df = df[cols]

            supplier_col = '供应商名称' if '供应商名称' in df.columns else '供应商'
            self.df_processed = df
            self.suppliers = [x for x in df[supplier_col].dropna().unique().tolist() if str(x).strip()]
            return True, "预处理完成"
        except Exception as e:
            return False, f"预处理失败: {str(e)}"

    def export_by_supplier(self, output_dir, date_prefix):
        if self.df_processed is None:
            return False, "请先预处理数据"
        try:
            os.makedirs(output_dir, exist_ok=True)
            supplier_col = '供应商名称' if '供应商名称' in self.df_processed.columns else '供应商'
            results = []
            for supplier in self.suppliers:
                supplier_df = self.df_processed[self.df_processed[supplier_col] == supplier]
                if len(supplier_df) == 0:
                    continue
                safe_supplier = re.sub(r'[\\/:*?"<>|]+', '_', str(supplier).strip())
                filename = f"{date_prefix}{safe_supplier}直发.xlsx"
                filepath = os.path.join(output_dir, filename)
                supplier_df.to_excel(filepath, index=False, engine='openpyxl')
                results.append({'supplier': str(supplier), 'count': len(supplier_df), 'file': filepath})
            if not results:
                return False, "没有数据可导出，请检查预处理结果"
            return True, results
        except Exception as e:
            import traceback
            return False, f"导出失败: {str(e)}\n{traceback.format_exc()}"

class SupplierDirectTableWindow:
    """嵌入主生成器的供应商直发表格生成器子窗口"""

    def __init__(self, parent_app):
        self.parent_app = parent_app
        self.window = tk.Toplevel(parent_app.root)
        self.window.title("供应商直发表格生成器子模块 v2.0")
        self.window.geometry("760x620")
        self.window.resizable(True, True)
        self.window.transient(parent_app.root)

        self.generator = SupplierDirectTableCore(parent_app)
        self.file_path = tk.StringVar()
        self.output_dir = tk.StringVar(value=parent_app.output_dir.get() or os.path.join(os.path.expanduser('~'), 'Desktop'))
        self.date_prefix = tk.StringVar(value=datetime.now().strftime("%m-%d").lstrip('0'))
        self.status_var = tk.StringVar(value="就绪")
        self.remove_price = tk.BooleanVar(value=True)
        self.add_to_main = tk.BooleanVar(value=False)

        self.create_widgets()
        self.center_window()

    def create_widgets(self):
        main_frame = ttk.Frame(self.window, padding="16")
        main_frame.pack(fill=tk.BOTH, expand=True)

        ttk.Label(main_frame, text="供应商直发表格生成器子模块", font=('Microsoft YaHei', 17, 'bold')).pack(pady=4)
        ttk.Label(main_frame, text="预处理预定单数据，并按供应商拆分导出多个Excel文件", foreground='gray').pack(pady=2)

        file_frame = ttk.LabelFrame(main_frame, text="文件选择", padding="10")
        file_frame.pack(fill=tk.X, pady=8)

        input_frame = ttk.Frame(file_frame)
        input_frame.pack(fill=tk.X, pady=4)
        ttk.Label(input_frame, text="预定单文件:", width=12).pack(side=tk.LEFT)
        ttk.Entry(input_frame, textvariable=self.file_path, width=60).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(input_frame, text="浏览...", command=self.browse_input_file).pack(side=tk.LEFT)

        output_frame = ttk.Frame(file_frame)
        output_frame.pack(fill=tk.X, pady=4)
        ttk.Label(output_frame, text="输出目录:", width=12).pack(side=tk.LEFT)
        ttk.Entry(output_frame, textvariable=self.output_dir, width=60).pack(side=tk.LEFT, padx=5, fill=tk.X, expand=True)
        ttk.Button(output_frame, text="浏览...", command=self.browse_output_dir).pack(side=tk.LEFT)

        option_frame = ttk.LabelFrame(main_frame, text="导出选项", padding="10")
        option_frame.pack(fill=tk.X, pady=8)

        date_frame = ttk.Frame(option_frame)
        date_frame.pack(fill=tk.X, pady=4)
        ttk.Label(date_frame, text="日期前缀:", width=12).pack(side=tk.LEFT)
        ttk.Entry(date_frame, textvariable=self.date_prefix, width=12).pack(side=tk.LEFT, padx=5)
        ttk.Label(date_frame, text="文件名格式：{日期前缀}{供应商名称}直发.xlsx").pack(side=tk.LEFT, padx=6)

        check_frame = ttk.Frame(option_frame)
        check_frame.pack(fill=tk.X, pady=4)
        ttk.Checkbutton(check_frame, text="删除'单价'列（对外发供应商时建议勾选）", variable=self.remove_price).pack(anchor=tk.W)
        ttk.Checkbutton(check_frame, text="导出后加入主生成器的工厂直发待处理列表", variable=self.add_to_main).pack(anchor=tk.W)

        explain_frame = ttk.LabelFrame(main_frame, text="预处理规则", padding="10")
        explain_frame.pack(fill=tk.X, pady=8)
        explain_text = (
            "1. 按'审批日期'筛选，只保留最新日期数据；\n"
            "2. 根据'供应商'编码增加'供应商名称'列：X04→兴发、X08→绘强、X14/X64→征图、S02→博克、S03→纳思达、X21→德源、X15→迪研；\n"
            "3. 按'供应商名称'拆分预定单数据，分别导出Excel文件；\n"
            "4. 如勾选加入主生成器列表，可继续用主界面的工厂直发XML功能处理。"
        )
        ttk.Label(explain_frame, text=explain_text, justify=tk.LEFT, font=('Microsoft YaHei', 9)).pack(anchor=tk.W)

        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(pady=8)
        self.export_btn = ttk.Button(btn_frame, text="执行预处理并生成表格", command=self.export_files)
        self.export_btn.pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="预览数据", command=self.preview_data).pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="打开输出目录", command=self.open_output_dir).pack(side=tk.LEFT, padx=8)
        ttk.Button(btn_frame, text="清空", command=self.clear).pack(side=tk.LEFT, padx=8)

        status_frame = ttk.Frame(main_frame)
        status_frame.pack(fill=tk.X, pady=4)
        ttk.Label(status_frame, text="状态:").pack(side=tk.LEFT)
        ttk.Label(status_frame, textvariable=self.status_var, foreground='blue').pack(side=tk.LEFT, padx=5)

        log_frame = ttk.LabelFrame(main_frame, text="处理日志", padding="5")
        log_frame.pack(fill=tk.BOTH, expand=True, pady=8)
        log_container = ttk.Frame(log_frame)
        log_container.pack(fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(log_container)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.log_text = tk.Text(log_container, height=11, width=88, font=('Consolas', 9), yscrollcommand=scrollbar.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.log_text.yview)

    def center_window(self):
        self.window.update_idletasks()
        w, h = self.window.winfo_width(), self.window.winfo_height()
        x = (self.window.winfo_screenwidth() // 2) - (w // 2)
        y = (self.window.winfo_screenheight() // 2) - (h // 2)
        self.window.geometry(f'{w}x{h}+{x}+{y}')

    def log(self, message):
        self.log_text.insert(tk.END, f"{message}\n")
        self.log_text.see(tk.END)
        self.window.update()

    def browse_input_file(self):
        filename = filedialog.askopenfilename(
            title="选择预定单Excel文件",
            filetypes=[("Excel文件", "*.xls *.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            self.file_path.set(filename)
            if not self.output_dir.get():
                self.output_dir.set(self.parent_app.output_dir.get() or os.path.join(os.path.expanduser('~'), 'Desktop'))

    def browse_output_dir(self):
        dirname = filedialog.askdirectory(title="选择输出目录")
        if dirname:
            self.output_dir.set(dirname)

    def open_output_dir(self):
        path = self.output_dir.get()
        if not path or not os.path.exists(path):
            messagebox.showwarning("提示", "输出目录不存在")
            return
        try:
            if os.name == 'nt':
                os.startfile(path)
            elif sys.platform == 'darwin':
                os.system(f'open "{path}"')
            else:
                os.system(f'xdg-open "{path}"')
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def preview_data(self):
        if not self.file_path.get():
            messagebox.showwarning("警告", "请先选择预定单文件")
            return
        self.log("正在加载数据...")
        success, message = self.generator.load_data(self.file_path.get())
        self.log(message)
        if not success:
            messagebox.showerror("错误", message)
            return

        df = self.generator.df
        if '审批日期' in df.columns:
            self.log("\n【审批日期分布】")
            for date, count in df['审批日期'].value_counts(dropna=False).items():
                self.log(f"  {date}: {count} 条")
        if '供应商' in df.columns:
            self.log("\n【供应商分布】")
            for supplier, count in df['供应商'].astype(str).str.strip().value_counts(dropna=False).items():
                name = SUPPLIER_DIRECT_NAME_MAPPING.get(supplier, supplier)
                self.log(f"  {supplier} ({name}): {count} 条")
        elif '供应商名称' in df.columns:
            self.log("\n【供应商名称分布】")
            for supplier, count in df['供应商名称'].value_counts(dropna=False).items():
                self.log(f"  {supplier}: {count} 条")

    def export_files(self):
        if not self.file_path.get():
            messagebox.showwarning("警告", "请先选择预定单文件")
            return
        if not self.output_dir.get():
            messagebox.showwarning("警告", "请先选择输出目录")
            return
        try:
            self.export_btn.config(state=tk.DISABLED)
            self.log("=" * 60)
            self.log("开始处理...\n")

            self.log("[1/3] 加载数据...")
            success, message = self.generator.load_data(self.file_path.get())
            self.log(message)
            if not success:
                raise Exception(message)

            self.log("\n[2/3] 执行预处理...")
            success, message = self.generator.preprocess_data(remove_price=self.remove_price.get())
            self.log(message)
            if not success:
                raise Exception(message)

            if self.generator.latest_date:
                self.log(f"  ✓ 审批日期筛选：保留 {self.generator.latest_date}，删除 {self.generator.filtered_count} 条")
            self.log(f"  ✓ {'删除' if self.remove_price.get() else '保留'}'单价'列")
            self.log(f"  ✓ 增加/规范'供应商名称'列")
            self.log(f"  ✓ 预处理后数据：{len(self.generator.df_processed):,} 条")

            self.log("\n[3/3] 按供应商导出...")
            success, result = self.generator.export_by_supplier(self.output_dir.get(), self.date_prefix.get())
            if not success:
                raise Exception(result)

            self.log("\n【导出结果】")
            for item in result:
                self.log(f"  ✓ {item['supplier']}: {item['count']} 条 → {os.path.basename(item['file'])}")

            if self.add_to_main.get():
                self.parent_app.add_factory_files_from_submodule([item['file'] for item in result])
                self.log(f"\n已加入主生成器工厂直发待处理列表：{len(result)} 个文件")

            self.log(f"\n共导出 {len(result)} 个文件到: {self.output_dir.get()}")
            self.status_var.set("导出成功")
            self.parent_app.log(f"✓ 供应商直发表格子模块导出成功：{len(result)} 个文件")
            messagebox.showinfo("成功", f"成功导出 {len(result)} 个文件！")
        except Exception as e:
            self.log(f"\n错误: {str(e)}")
            messagebox.showerror("错误", str(e))
            self.status_var.set("导出失败")
        finally:
            self.export_btn.config(state=tk.NORMAL)

    def clear(self):
        self.file_path.set("")
        self.log_text.delete(1.0, tk.END)
        self.status_var.set("就绪")

def _open_supplier_export_module(self):
    """打开供应商直发表格生成器子模块"""
    existing = getattr(self, 'supplier_export_window', None)
    if existing is not None:
        try:
            if existing.window.winfo_exists():
                existing.window.lift()
                existing.window.focus_force()
                return
        except Exception:
            pass
    self.supplier_export_window = SupplierDirectTableWindow(self)

def _add_factory_files_from_submodule(self, file_paths):
    """把子模块导出的供应商文件加入主生成器工厂直发待处理列表"""
    if self.order_type.get() != 'factory':
        self.order_type.set('factory')
        self.on_order_type_change()
    for file_path in file_paths:
        supplier = self.identify_supplier_from_content(file_path)
        self.selected_files.append((file_path, supplier))
        filename = os.path.basename(file_path)
        display_name = filename[:42] + "..." if len(filename) > 42 else filename
        self.file_listbox.insert(tk.END, f"  {supplier:10}    {display_name}")
        self.log(f"✓ 子模块导出文件已加入：{filename} → {supplier}")



# =====================================================
# 新品上新定价生成器子模块（由独立GUI改为Toplevel子窗口）
# =====================================================
# =========================
# 1. 定价规则配置区
# =========================

@dataclass
class MarginRule:
    min_margin: float
    max_margin: float
    target_margin: float


@dataclass
class BrandRule:
    brand: str
    wholesale: MarginRule
    retail: MarginRule
    wholesale_tail: str = "auto"
    retail_tail: str = "auto"
    note: str = ""


CURRENT_PRIORITY_RULES: Dict[str, BrandRule] = {
    "智通": BrandRule(
        brand="智通",
        wholesale=MarginRule(0.35, 0.40, 0.38),
        retail=MarginRule(0.49, 0.57, 0.57),
        wholesale_tail="auto",
        retail_tail="9",
        note="当前优先规则：批发35%-40%，零售49%-57%"
    ),
    "盈佳上尊": BrandRule(
        brand="盈佳上尊",
        wholesale=MarginRule(0.33, 0.38, 0.35),
        retail=MarginRule(0.53, 0.57, 0.55),
        wholesale_tail="auto",
        retail_tail="auto",
        note="当前优先规则：批发33%-38%，零售约55%"
    ),
    "懿品佳": BrandRule(
        brand="懿品佳",
        wholesale=MarginRule(0.30, 0.35, 0.325),
        retail=MarginRule(0.50, 0.59, 0.55),
        wholesale_tail="auto",
        retail_tail="0",
        note="当前优先规则：批发30%-35%，零售50%-59%"
    ),
}


RULES_WITH_PRICE_DROP_ROOM: Dict[str, BrandRule] = {
    "懿智通": BrandRule("懿智通", MarginRule(0.20, 0.25, 0.21), MarginRule(0.43, 0.48, 0.45), "auto", "8"),
    "懿智通医疗版": BrandRule("懿智通医疗版", MarginRule(0.25, 0.30, 0.30), MarginRule(0.48, 0.53, 0.50), "auto", "8", "医疗版一般比正常版贵约5%"),
    "懿品佳": BrandRule("懿品佳", MarginRule(0.25, 0.30, 0.25), MarginRule(0.43, 0.50, 0.45), "auto", "0"),
    "盈佳": BrandRule("盈佳", MarginRule(0.35, 0.43, 0.39), MarginRule(0.42, 0.50, 0.48), "auto", "8"),
    "盈佳上尊": BrandRule("盈佳上尊", MarginRule(0.25, 0.28, 0.26), MarginRule(0.47, 0.51, 0.49), "auto", "auto"),
    "扬帆耐立": BrandRule("扬帆耐立", MarginRule(0.45, 0.55, 0.48), MarginRule(0.50, 0.55, 0.53), "auto", "9"),
    "智通": BrandRule("智通", MarginRule(0.30, 0.35, 0.30), MarginRule(0.44, 0.49, 0.45), "auto", "9"),
    "智通Plus": BrandRule("智通Plus", MarginRule(0.35, 0.42, 0.36), MarginRule(0.57, 0.68, 0.60), "auto", "9"),
}


RULES_WITHOUT_PRICE_DROP_ROOM: Dict[str, BrandRule] = {
    "懿智通": BrandRule("懿智通", MarginRule(0.24, 0.30, 0.26), MarginRule(0.48, 0.55, 0.49), "auto", "8"),
    "懿智通医疗版": BrandRule("懿智通医疗版", MarginRule(0.29, 0.35, 0.31), MarginRule(0.53, 0.60, 0.55), "auto", "8", "按普通版约上浮5%处理，需人工复核"),
    "懿品佳": BrandRule("懿品佳", MarginRule(0.27, 0.35, 0.29), MarginRule(0.50, 0.59, 0.50), "auto", "0"),
    "盈佳": BrandRule("盈佳", MarginRule(0.46, 0.53, 0.46), MarginRule(0.53, 0.63, 0.53), "auto", "8"),
    "盈佳上尊": BrandRule("盈佳上尊", MarginRule(0.31, 0.35, 0.33), MarginRule(0.53, 0.57, 0.55), "auto", "auto"),
    "扬帆耐立": BrandRule("扬帆耐立", MarginRule(0.52, 0.60, 0.53), MarginRule(0.55, 0.66, 0.55), "auto", "9"),
    "智通": BrandRule("智通", MarginRule(0.36, 0.42, 0.38), MarginRule(0.49, 0.57, 0.49), "auto", "9"),
    "智通Plus": BrandRule("智通Plus", MarginRule(0.39, 0.45, 0.41), MarginRule(0.64, 0.75, 0.69), "auto", "9"),
}


DEFAULT_BRANDS = ["懿智通", "懿品佳", "盈佳上尊", "智通", "智通Plus", "盈佳", "扬帆耐立"]
ALL_BRANDS = ["懿智通", "懿智通医疗版", "懿品佳", "盈佳上尊", "智通", "智通Plus", "盈佳", "扬帆耐立"]


# =========================
# 2. 计算函数
# =========================

def to_float(value, default=None):
    if value is None:
        return default
    text = str(value).strip().replace("元", "").replace(",", "")
    if text == "":
        return default
    try:
        return float(text)
    except ValueError:
        return default


def gross_margin(cost: float, price: float) -> float:
    if price <= 0:
        return 0.0
    return (price - cost) / price


def theoretical_price(cost: float, target_margin: float) -> float:
    return cost / (1 - target_margin)


def allowed_by_tail(price: int, tail: str) -> bool:
    if tail in ("auto", "", None):
        return True
    last = str(price)[-1]
    if tail == "05":
        return last in ("0", "5")
    return last == str(tail)


def candidate_prices(cost: float, rule: MarginRule, tail: str, min_price: int = 1, max_price: int = 99999) -> List[int]:
    low = max(min_price, int(cost / (1 - rule.min_margin)) - 5)
    high = min(max_price, int(cost / (1 - rule.max_margin)) + 10)
    result = []
    for price in range(max(1, low), high + 1):
        margin = gross_margin(cost, price)
        if rule.min_margin <= margin <= rule.max_margin and allowed_by_tail(price, tail):
            result.append(price)
    return sorted(set(result))


def choose_price(cost: float, rule: MarginRule, tail: str = "auto", price_type: str = "wholesale") -> int:
    target_price = theoretical_price(cost, rule.target_margin)

    if price_type == "wholesale":
        search_tail = "05" if tail == "05" else "auto"
    else:
        search_tail = tail if tail not in ("", None) else "auto"

    candidates = candidate_prices(cost, rule, search_tail)

    if not candidates and price_type == "retail":
        candidates = candidate_prices(cost, rule, "auto")

    if not candidates:
        price = max(1, round(target_price))
        while gross_margin(cost, price) < rule.min_margin:
            price += 1
        return price

    def score(p: int):
        margin_diff = abs(gross_margin(cost, p) - rule.target_margin)
        tail_bonus = 0
        if price_type == "wholesale" and str(p)[-1] in ("0", "5"):
            tail_bonus = -0.0001
        if price_type == "retail" and allowed_by_tail(p, tail):
            tail_bonus = -0.0001
        return (margin_diff + tail_bonus, p)

    return sorted(candidates, key=score)[0]


def format_percent(x: float) -> str:
    return f"{x * 100:.2f}%"


def resolve_rules(mode: str) -> Dict[str, BrandRule]:
    if mode == "有降价空间":
        return dict(RULES_WITH_PRICE_DROP_ROOM)
    if mode == "无降价空间":
        return dict(RULES_WITHOUT_PRICE_DROP_ROOM)
    rules = dict(RULES_WITHOUT_PRICE_DROP_ROOM)
    rules.update(CURRENT_PRIORITY_RULES)
    return rules


def price_one_item(model: str, cost: float, capacity: str = "", original_retail: Optional[float] = None,
                   brands: Optional[List[str]] = None, mode: str = "当前规则优先",
                   enforce_wholesale_05: bool = False) -> List[Dict[str, object]]:
    rules = resolve_rules(mode)
    selected_brands = brands or DEFAULT_BRANDS
    rows = []

    for brand in selected_brands:
        if brand not in rules:
            rows.append({
                "品牌": brand, "型号": model, "容量": capacity, "成本": cost,
                "建议批发价": "", "批发毛利率": "", "建议零售价": "", "零售毛利率": "",
                "原装零售价": original_retail if original_retail else "",
                "原装8折": round(original_retail * 0.8, 2) if original_retail else "",
                "原装7折": round(original_retail * 0.7, 2) if original_retail else "",
                "复核提示": "未配置该品牌规则", "规则备注": ""
            })
            continue

        rule = rules[brand]
        wholesale_tail = "05" if enforce_wholesale_05 else rule.wholesale_tail
        wholesale_price = choose_price(cost, rule.wholesale, wholesale_tail, "wholesale")
        retail_price = choose_price(cost, rule.retail, rule.retail_tail, "retail")
        w_margin = gross_margin(cost, wholesale_price)
        r_margin = gross_margin(cost, retail_price)

        warnings = []
        if not (rule.wholesale.min_margin <= w_margin <= rule.wholesale.max_margin):
            warnings.append("批发毛利率需复核")
        if not (rule.retail.min_margin <= r_margin <= rule.retail.max_margin):
            warnings.append("零售毛利率需复核")
        if original_retail:
            if retail_price >= original_retail * 0.8:
                warnings.append("零售价高于或接近原装8折")
            elif retail_price >= original_retail * 0.7:
                warnings.append("零售价接近原装7折")
        if brand == "懿智通医疗版":
            warnings.append("医疗版需确认是否较普通版上浮约5%")

        rows.append({
            "品牌": brand,
            "型号": model,
            "容量": capacity,
            "成本": round(cost, 2),
            "建议批发价": wholesale_price,
            "批发毛利率": format_percent(w_margin),
            "建议零售价": retail_price,
            "零售毛利率": format_percent(r_margin),
            "原装零售价": original_retail if original_retail else "",
            "原装8折": round(original_retail * 0.8, 2) if original_retail else "",
            "原装7折": round(original_retail * 0.7, 2) if original_retail else "",
            "复核提示": "；".join(warnings),
            "规则备注": rule.note,
        })
    return rows


def make_chat_table(rows: List[Dict[str, object]]) -> str:
    headers = ["品牌", "型号", "成本", "建议批发价", "批发毛利率", "建议零售价", "零售毛利率", "复核提示"]
    lines = ["｜".join(headers), "｜".join(["——"] * len(headers))]
    for row in rows:
        lines.append("｜".join(str(row.get(h, "")) for h in headers))
    return "\n".join(lines)


def make_text_detail(rows: List[Dict[str, object]]) -> str:
    parts = []
    for row in rows:
        parts.append(
            f"{row['品牌']}{row['型号']}，成本{row['成本']}元，"
            f"建议批发价{row['建议批发价']}元，批发毛利率{row['批发毛利率']}，"
            f"建议零售价{row['建议零售价']}元，零售毛利率{row['零售毛利率']}"
        )
    return "\n\n".join(parts)


# =========================
# 3. GUI界面
# =========================

class PricingToolWindow(tk.Toplevel):
    def __init__(self, parent_app):
        self.parent_app = parent_app
        super().__init__(parent_app.root)
        self.title("新品上新定价生成器 - 全品牌版")
        self.geometry("1260x760")
        self.minsize(1100, 680)
        self.transient(parent_app.root)
        self.protocol("WM_DELETE_WINDOW", self.destroy)

        self.input_items: List[Dict[str, object]] = []
        self.result_rows: List[Dict[str, object]] = []
        self.brand_vars: Dict[str, tk.BooleanVar] = {}

        self._build_ui()
        self._load_demo()

    def _build_ui(self):
        self.columnconfigure(0, weight=1)
        self.rowconfigure(2, weight=1)

        title = ttk.Label(self, text="新品上新定价生成器 - 全品牌版", font=("Microsoft YaHei UI", 16, "bold"))
        title.grid(row=0, column=0, sticky="w", padx=14, pady=(10, 4))

        top = ttk.Frame(self)
        top.grid(row=1, column=0, sticky="ew", padx=14, pady=6)
        top.columnconfigure(1, weight=1)

        input_box = ttk.LabelFrame(top, text="新品输入")
        input_box.grid(row=0, column=0, sticky="nsw", padx=(0, 10))
        for i in range(2):
            input_box.columnconfigure(i, weight=1)

        ttk.Label(input_box, text="型号").grid(row=0, column=0, sticky="w", padx=8, pady=(8, 2))
        self.model_var = tk.StringVar()
        ttk.Entry(input_box, textvariable=self.model_var, width=24).grid(row=1, column=0, padx=8, pady=2)

        ttk.Label(input_box, text="成本").grid(row=0, column=1, sticky="w", padx=8, pady=(8, 2))
        self.cost_var = tk.StringVar()
        ttk.Entry(input_box, textvariable=self.cost_var, width=12).grid(row=1, column=1, padx=8, pady=2)

        ttk.Label(input_box, text="容量").grid(row=2, column=0, sticky="w", padx=8, pady=(8, 2))
        self.capacity_var = tk.StringVar()
        ttk.Entry(input_box, textvariable=self.capacity_var, width=24).grid(row=3, column=0, padx=8, pady=2)

        ttk.Label(input_box, text="原装零售价，可空").grid(row=2, column=1, sticky="w", padx=8, pady=(8, 2))
        self.original_var = tk.StringVar()
        ttk.Entry(input_box, textvariable=self.original_var, width=12).grid(row=3, column=1, padx=8, pady=2)

        btn_row = ttk.Frame(input_box)
        btn_row.grid(row=4, column=0, columnspan=2, sticky="ew", padx=8, pady=8)
        ttk.Button(btn_row, text="添加到待定价", command=self.add_item).pack(side="left", padx=(0, 6))
        ttk.Button(btn_row, text="清空输入", command=self.clear_input_fields).pack(side="left")

        option_box = ttk.LabelFrame(top, text="规则与品牌")
        option_box.grid(row=0, column=1, sticky="nsew")
        option_box.columnconfigure(0, weight=1)

        rule_frame = ttk.Frame(option_box)
        rule_frame.grid(row=0, column=0, sticky="ew", padx=8, pady=(8, 4))

        ttk.Label(rule_frame, text="规则口径").pack(side="left")
        self.mode_var = tk.StringVar(value="当前规则优先")
        mode_combo = ttk.Combobox(rule_frame, textvariable=self.mode_var, width=18, state="readonly",
                                  values=["当前规则优先", "有降价空间", "无降价空间"])
        mode_combo.pack(side="left", padx=(8, 16))

        self.enforce_05_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(rule_frame, text="批发价强制0/5结尾", variable=self.enforce_05_var).pack(side="left")

        brand_frame = ttk.Frame(option_box)
        brand_frame.grid(row=1, column=0, sticky="ew", padx=8, pady=4)
        for i, brand in enumerate(ALL_BRANDS):
            var = tk.BooleanVar(value=(brand in DEFAULT_BRANDS))
            self.brand_vars[brand] = var
            ttk.Checkbutton(brand_frame, text=brand, variable=var).grid(row=i // 4, column=i % 4, sticky="w", padx=8, pady=2)

        action_frame = ttk.Frame(option_box)
        action_frame.grid(row=2, column=0, sticky="ew", padx=8, pady=(8, 8))
        ttk.Button(action_frame, text="全选品牌", command=self.select_all_brands).pack(side="left", padx=(0, 6))
        ttk.Button(action_frame, text="默认品牌", command=self.select_default_brands).pack(side="left", padx=(0, 6))
        ttk.Button(action_frame, text="导入CSV", command=self.import_csv).pack(side="left", padx=(0, 6))
        ttk.Button(action_frame, text="清空待定价", command=self.clear_items).pack(side="left", padx=(0, 6))
        ttk.Button(action_frame, text="生成定价", command=self.generate_pricing).pack(side="left", padx=(12, 6))
        ttk.Button(action_frame, text="规则说明", command=self.show_rules).pack(side="left")

        notebook = ttk.Notebook(self)
        notebook.grid(row=2, column=0, sticky="nsew", padx=14, pady=(4, 10))

        pending_tab = ttk.Frame(notebook)
        result_tab = ttk.Frame(notebook)
        chat_tab = ttk.Frame(notebook)
        notebook.add(pending_tab, text="待定价清单")
        notebook.add(result_tab, text="定价结果")
        notebook.add(chat_tab, text="聊天复制版")

        self.pending_tree = self._make_tree(pending_tab, ["型号", "成本", "容量", "原装零售价"])
        self.result_tree = self._make_tree(result_tab, [
            "品牌", "型号", "容量", "成本", "建议批发价", "批发毛利率", "建议零售价",
            "零售毛利率", "原装零售价", "原装8折", "原装7折", "复核提示"
        ])

        chat_frame = ttk.Frame(chat_tab)
        chat_frame.pack(fill="both", expand=True, padx=8, pady=8)
        chat_frame.rowconfigure(0, weight=1)
        chat_frame.columnconfigure(0, weight=1)

        self.chat_text = tk.Text(chat_frame, wrap="none", font=("Microsoft YaHei UI", 10))
        self.chat_text.grid(row=0, column=0, sticky="nsew")
        yscroll = ttk.Scrollbar(chat_frame, orient="vertical", command=self.chat_text.yview)
        xscroll = ttk.Scrollbar(chat_frame, orient="horizontal", command=self.chat_text.xview)
        self.chat_text.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")

        bottom = ttk.Frame(self)
        bottom.grid(row=3, column=0, sticky="ew", padx=14, pady=(0, 12))
        ttk.Button(bottom, text="复制聊天框表格", command=self.copy_chat_table).pack(side="left", padx=(0, 8))
        ttk.Button(bottom, text="导出CSV", command=self.export_csv).pack(side="left", padx=(0, 8))
        ttk.Button(bottom, text="导出TXT", command=self.export_txt).pack(side="left", padx=(0, 8))
        ttk.Button(bottom, text="删除选中待定价项", command=self.delete_selected_pending).pack(side="left", padx=(18, 8))
        ttk.Button(bottom, text="退出", command=self.destroy).pack(side="right")

        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(bottom, textvariable=self.status_var).pack(side="right", padx=(0, 20))

    def _make_tree(self, parent, columns):
        frame = ttk.Frame(parent)
        frame.pack(fill="both", expand=True, padx=8, pady=8)
        frame.rowconfigure(0, weight=1)
        frame.columnconfigure(0, weight=1)

        tree = ttk.Treeview(frame, columns=columns, show="headings", height=14)
        for col in columns:
            tree.heading(col, text=col)
            width = 110
            if col in ("型号", "复核提示"):
                width = 180
            if col == "容量":
                width = 80
            tree.column(col, width=width, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")

        yscroll = ttk.Scrollbar(frame, orient="vertical", command=tree.yview)
        xscroll = ttk.Scrollbar(frame, orient="horizontal", command=tree.xview)
        tree.configure(yscrollcommand=yscroll.set, xscrollcommand=xscroll.set)
        yscroll.grid(row=0, column=1, sticky="ns")
        xscroll.grid(row=1, column=0, sticky="ew")
        return tree

    def _load_demo(self):
        self.input_items = [
            {"型号": "W1580A-2.5K", "成本": 13.0, "容量": "2.5K", "原装零售价": ""},
            {"型号": "W1580X-5K", "成本": 16.0, "容量": "5K", "原装零售价": ""},
        ]
        self.refresh_pending_tree()

    def add_item(self):
        model = self.model_var.get().strip()
        cost = to_float(self.cost_var.get())
        capacity = self.capacity_var.get().strip()
        original = to_float(self.original_var.get(), "")

        if not model:
            messagebox.showwarning("提示", "型号不能为空。")
            return
        if cost is None:
            messagebox.showwarning("提示", "成本必须填写数字。")
            return

        self.input_items.append({"型号": model, "成本": cost, "容量": capacity, "原装零售价": original if original else ""})
        self.refresh_pending_tree()
        self.clear_input_fields()
        self.status_var.set(f"已添加：{model}")

    def clear_input_fields(self):
        self.model_var.set("")
        self.cost_var.set("")
        self.capacity_var.set("")
        self.original_var.set("")

    def refresh_pending_tree(self):
        self.pending_tree.delete(*self.pending_tree.get_children())
        for item in self.input_items:
            self.pending_tree.insert("", "end", values=[
                item.get("型号", ""), item.get("成本", ""), item.get("容量", ""), item.get("原装零售价", "")
            ])

    def refresh_result_tree(self):
        self.result_tree.delete(*self.result_tree.get_children())
        for row in self.result_rows:
            self.result_tree.insert("", "end", values=[
                row.get("品牌", ""), row.get("型号", ""), row.get("容量", ""), row.get("成本", ""),
                row.get("建议批发价", ""), row.get("批发毛利率", ""), row.get("建议零售价", ""),
                row.get("零售毛利率", ""), row.get("原装零售价", ""), row.get("原装8折", ""),
                row.get("原装7折", ""), row.get("复核提示", "")
            ])

    def selected_brands(self):
        return [brand for brand, var in self.brand_vars.items() if var.get()]

    def select_all_brands(self):
        for var in self.brand_vars.values():
            var.set(True)

    def select_default_brands(self):
        for brand, var in self.brand_vars.items():
            var.set(brand in DEFAULT_BRANDS)

    def clear_items(self):
        if messagebox.askyesno("确认", "确认清空待定价清单？"):
            self.input_items = []
            self.refresh_pending_tree()
            self.status_var.set("已清空待定价清单")

    def delete_selected_pending(self):
        selected = self.pending_tree.selection()
        if not selected:
            return
        indexes = sorted([self.pending_tree.index(item) for item in selected], reverse=True)
        for idx in indexes:
            if 0 <= idx < len(self.input_items):
                self.input_items.pop(idx)
        self.refresh_pending_tree()

    def generate_pricing(self):
        brands = self.selected_brands()
        if not brands:
            messagebox.showwarning("提示", "请至少选择一个品牌。")
            return
        if not self.input_items:
            messagebox.showwarning("提示", "请先添加或导入待定价产品。")
            return

        rows = []
        for item in self.input_items:
            rows.extend(price_one_item(
                model=str(item.get("型号", "")).strip(),
                cost=float(item.get("成本", 0)),
                capacity=str(item.get("容量", "")).strip(),
                original_retail=to_float(item.get("原装零售价")),
                brands=brands,
                mode=self.mode_var.get(),
                enforce_wholesale_05=self.enforce_05_var.get()
            ))
        self.result_rows = rows
        self.refresh_result_tree()

        self.chat_text.delete("1.0", "end")
        self.chat_text.insert("1.0", "【逐条文本】\n")
        self.chat_text.insert("end", make_text_detail(rows))
        self.chat_text.insert("end", "\n\n【聊天框表格】\n")
        self.chat_text.insert("end", make_chat_table(rows))

        self.status_var.set(f"已生成 {len(rows)} 条定价结果")

    def import_csv(self):
        path = filedialog.askopenfilename(
            title="选择CSV文件",
            filetypes=[("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not path:
            return
        imported = []
        try:
            with open(path, "r", encoding="utf-8-sig") as f:
                reader = csv.DictReader(f)
                for raw in reader:
                    model = (raw.get("型号") or raw.get("通用型号") or raw.get("model") or "").strip()
                    cost = to_float(raw.get("成本") or raw.get("工厂成本") or raw.get("cost"))
                    capacity = (raw.get("容量") or raw.get("capacity") or "").strip()
                    original = to_float(raw.get("原装零售价") or raw.get("original_retail"), "")
                    if model and cost is not None:
                        imported.append({"型号": model, "成本": cost, "容量": capacity, "原装零售价": original if original else ""})
            self.input_items.extend(imported)
            self.refresh_pending_tree()
            self.status_var.set(f"已导入 {len(imported)} 条")
        except Exception as e:
            messagebox.showerror("导入失败", str(e))

    def export_csv(self):
        if not self.result_rows:
            messagebox.showwarning("提示", "请先生成定价结果。")
            return
        path = filedialog.asksaveasfilename(
            title="保存CSV",
            defaultextension=".csv",
            filetypes=[("CSV文件", "*.csv")]
        )
        if not path:
            return
        try:
            fieldnames = list(self.result_rows[0].keys())
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(self.result_rows)
            self.status_var.set(f"已导出CSV：{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def export_txt(self):
        if not self.result_rows:
            messagebox.showwarning("提示", "请先生成定价结果。")
            return
        path = filedialog.asksaveasfilename(
            title="保存TXT",
            defaultextension=".txt",
            filetypes=[("TXT文件", "*.txt")]
        )
        if not path:
            return
        try:
            content = self.chat_text.get("1.0", "end").strip()
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            self.status_var.set(f"已导出TXT：{path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))

    def copy_chat_table(self):
        content = self.chat_text.get("1.0", "end").strip()
        if not content:
            messagebox.showwarning("提示", "请先生成定价结果。")
            return
        self.clipboard_clear()
        self.clipboard_append(content)
        self.status_var.set("已复制到剪贴板，可直接粘贴到聊天框")

    def show_rules(self):
        text = (
            "当前规则优先：\n"
            "智通：批发35%-40%，零售49%-57%。\n"
            "盈佳上尊：批发33%-38%，零售约55%。\n"
            "懿品佳：批发30%-35%，零售50%-59%。\n\n"
            "其他品牌默认按无降价空间口径：\n"
            "懿智通：批发24%-30%，零售48%-55%。\n"
            "盈佳：批发46%-53%，零售53%-63%。\n"
            "扬帆耐立：批发52%-60%，零售55%-66%。\n"
            "智通Plus：批发39%-45%，零售64%-75%。\n\n"
            "说明：低成本产品默认不强制批发价0/5结尾，优先保证毛利率区间。"
        )
        messagebox.showinfo("规则说明", text)




# ===== 存货成本回填生成器子模块（集成自批量暂存版）=====
# -*- coding: utf-8 -*-
"""
存货成本回填生成器 - 批量暂存版

核心变化：
1. 支持一次性选择多个成本来源表。
2. 不再每次生成后自动保存Excel。
3. 点击“开始回填/加入暂存”后，所有匹配结果暂存在当前程序内。
4. 可继续追加更多供应商文件，回填结果会累计保留。
5. 最后点击“导出结果”才生成Excel。
6. 支持绘强 / 征图 / 兴发：
   - 绘强：存货编码 + 成本
   - 征图：对方货号/条码 + 单价
   - 兴发：客户产品编码 + 客户产品条码 + 单价，编码优先，条码兜底
"""

import os
import re
import traceback
import tkinter as tk
from tkinter import filedialog, messagebox
from decimal import Decimal, InvalidOperation
from copy import copy
from openpyxl import load_workbook


CODE_ALIASES = [
    "存货编码", "物料编码", "商品编码", "产品编码", "编码", "nc存货编码", "nc编码",
    "产品代码", "产品长代码", "客户产品编码", "客户编码", "客户物料编码", "客户商品编码", "供应商产品编码", "供应商编码"
]

BARCODE_ALIASES = [
    "条码", "商品条码", "条形码", "69码", "69条码", "69编码", "最新69码", "新69码", "最新条码",
    "对方货号", "货号", "barcode", "bar code",
    "客户产品条码", "客户条码", "产品条码", "供应商条码"
]

COST_ALIASES = [
    "成本", "存货成本", "单价", "单价成本", "最新单价", "最新成本", "征图报价", "报价",
    "nc成本", "参考成本", "含税成本", "不含税成本", "采购单价", "销售单价", "含税单价"
]

NAME_ALIASES = [
    "存货名称", "产品名称", "型号", "产品型号", "规格型号", "客户规格型号", "名称"
]


def normalize_text(v):
    if v is None:
        return ""
    return str(v).strip().replace(" ", "").replace("\u3000", "").lower()


def clean_key(value):
    if value is None:
        return ""

    s = str(value).strip()
    if not s:
        return ""

    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]

    if re.fullmatch(r"[+-]?\d+(\.\d+)?[eE][+-]?\d+", s):
        try:
            s = format(Decimal(s), "f").rstrip("0").rstrip(".")
        except InvalidOperation:
            pass

    return s.strip()


def split_keys(value):
    s = clean_key(value)
    if not s:
        return []

    parts = re.split(r"[,，;；/、\n\r\t]+", s)
    result = []
    for p in parts:
        p = clean_key(p)
        if p and p not in result:
            result.append(p)
    return result


def clean_cost(value):
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value).strip().replace(",", "")
    if not s:
        return None
    s = s.replace("￥", "").replace("¥", "")
    try:
        return float(s)
    except Exception:
        return value


def read_xlsx_sheets(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for ws in wb.worksheets:
        values = []

        # K239成本文件里“销售订单”是历史明细大表，可能存在异常字符；
        # 成本池应取“按品牌产品分类/TO客户”等汇总表，所以这里只读取前40行用于识别，不整表解析。
        max_read = ws.max_row
        if "销售订单" in str(ws.title) and (ws.max_row or 0) > 5000:
            max_read = 40

        try:
            for row in ws.iter_rows(min_row=1, max_row=max_read, values_only=True):
                values.append(list(row))
        except Exception:
            # 单个sheet异常时不中断整个文件；其他可识别sheet继续参与。
            continue

        sheets.append((ws.title, values))
    return sheets


def read_xls_sheets(path):
    try:
        import xlrd
    except Exception:
        raise RuntimeError(
            "当前文件是 .xls 格式，需要安装 xlrd 后才能读取。\n"
            "请在命令行运行：pip install xlrd\n"
            "或者先用Excel另存为 .xlsx 后再导入。"
        )

    book = xlrd.open_workbook(path)
    sheets = []

    for sheet in book.sheets():
        values = []
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                v = cell.value
                if cell.ctype == xlrd.XL_CELL_EMPTY:
                    v = None
                elif cell.ctype == xlrd.XL_CELL_NUMBER:
                    if float(v).is_integer():
                        v = int(v)
                row.append(v)
            values.append(row)
        sheets.append((sheet.name, values))

    return sheets


def read_source_sheets(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in [".xlsx", ".xlsm"]:
        return read_xlsx_sheets(path)
    if ext == ".xls":
        return read_xls_sheets(path)
    raise ValueError("暂不支持该文件格式，请使用 .xlsx、.xlsm 或 .xls。")


def find_header_in_matrix(values, need_cost=True, max_scan_rows=35):
    code_aliases = [normalize_text(x) for x in CODE_ALIASES]
    barcode_aliases = [normalize_text(x) for x in BARCODE_ALIASES]
    cost_aliases = [normalize_text(x) for x in COST_ALIASES]
    name_aliases = [normalize_text(x) for x in NAME_ALIASES]

    best = None

    for r_idx, row in enumerate(values[:max_scan_rows]):
        normalized = [normalize_text(x) for x in row]

        code_col = None
        barcode_col = None
        cost_col = None
        name_col = None

        for c_idx, h in enumerate(normalized):
            if h in code_aliases and code_col is None:
                code_col = c_idx
            if h in barcode_aliases and barcode_col is None:
                barcode_col = c_idx
            if h in cost_aliases and cost_col is None:
                cost_col = c_idx
            if h in name_aliases and name_col is None:
                name_col = c_idx

        score = 0
        if code_col is not None:
            score += 3
        if barcode_col is not None:
            score += 3
        if cost_col is not None:
            score += 4
        if name_col is not None:
            score += 1

        if score:
            item = {
                "header_index": r_idx,
                "code_col": code_col,
                "barcode_col": barcode_col,
                "cost_col": cost_col,
                "name_col": name_col,
                "score": score,
            }
            best = item

            if need_cost and cost_col is not None and (code_col is not None or barcode_col is not None):
                return item
            if not need_cost and (code_col is not None or barcode_col is not None):
                return item

    return best


def source_candidate_priority(sheet_name, values, header):
    """
    选择来源表时的优先级。
    目的：K239 文件里同时有“销售订单”和“按品牌产品分类/TO客户”汇总表，
    应优先取“产品代码 + 最新单价”的汇总表，避免误取销售订单里的“含税单价”。
    """
    headers = []
    if 0 <= header.get("header_index", -1) < len(values):
        headers = [normalize_text(x) for x in values[header["header_index"]]]

    hset = set(headers)
    sname = normalize_text(sheet_name)
    priority = 0

    if "最新单价" in hset:
        priority += 80
    if "产品代码" in hset:
        priority += 60
    if "征图报价" in hset:
        priority += 80
    if "最新69码" in hset:
        priority += 60
    if "客户产品编码" in hset and "客户产品条码" in hset:
        priority += 50
    if "单价" in hset and ("69码" in hset or "客户产品条码" in hset):
        priority += 40

    if "按品牌产品分类" in sname:
        priority += 120
    if "to客户" in sname:
        priority += 70
    if "新系统导出来单价" in sname:
        priority += 70
    if "征图" in sname and "报价" in sname:
        priority += 90

    # 明细销售订单通常包含大量历史交易行，含税单价不是我们要回填的“成本池”优先来源。
    if "销售订单" in sname:
        priority -= 200
    if "含税单价" in hset and "最新单价" not in hset:
        priority -= 80

    return priority


def add_to_map(target_map, conflicts, duplicates, key, cost, source_row, key_type, source_file):
    if not key:
        return
    if key in target_map:
        duplicates.setdefault(key_type, {}).setdefault(key, []).append(source_row)
        old_cost, old_file = target_map[key]
        if str(old_cost) != str(cost):
            conflicts.setdefault(key_type, {}).setdefault(key, set()).update([
                f"{old_cost}（{old_file}）",
                f"{cost}（{source_file}）"
            ])
    else:
        target_map[key] = (cost, source_file)


def build_source_maps(source_path):
    sheets = read_source_sheets(source_path)

    candidates = []
    for sheet_name, values in sheets:
        header = find_header_in_matrix(values, need_cost=True)
        if not header:
            continue
        if header["cost_col"] is None:
            continue
        if header["code_col"] is None and header["barcode_col"] is None:
            continue
        candidates.append((sheet_name, values, header))

    if not candidates:
        raise ValueError(
            "成本来源表未识别到可用字段。\n"
            "至少需要：\n"
            "1）存货编码/客户产品编码 + 成本/单价；或\n"
            "2）条码/对方货号/客户产品条码 + 成本/单价。"
        )

    candidates.sort(key=lambda x: (source_candidate_priority(x[0], x[1], x[2]), x[2]["score"], len(x[1])), reverse=True)
    sheet_name, values, header = candidates[0]
    source_file = os.path.basename(source_path)

    code_map = {}
    barcode_map = {}
    duplicates = {}
    conflicts = {}
    source_rows = []

    for r_idx in range(header["header_index"] + 1, len(values)):
        row = values[r_idx]

        if header["cost_col"] >= len(row):
            continue

        cost = clean_cost(row[header["cost_col"]])
        if cost is None or cost == "":
            continue

        code_keys = []
        barcode_keys = []
        name = ""

        if header["code_col"] is not None and header["code_col"] < len(row):
            code_keys = split_keys(row[header["code_col"]])

        if header["barcode_col"] is not None and header["barcode_col"] < len(row):
            barcode_keys = split_keys(row[header["barcode_col"]])

        if header["name_col"] is not None and header["name_col"] < len(row):
            name = row[header["name_col"]]

        if not code_keys and not barcode_keys:
            continue

        for key in code_keys:
            add_to_map(code_map, conflicts, duplicates, key, cost, r_idx + 1, "编码", source_file)

        for key in barcode_keys:
            add_to_map(barcode_map, conflicts, duplicates, key, cost, r_idx + 1, "条码", source_file)

        source_rows.append({
            "row": r_idx + 1,
            "code_keys": " / ".join(code_keys),
            "barcode_keys": " / ".join(barcode_keys),
            "name": name,
            "cost": cost,
            "source_file": source_file,
        })

    if not code_map and not barcode_map:
        raise ValueError("成本来源表中没有读取到有效的编码/条码 + 成本数据。")

    return {
        "sheet_name": sheet_name,
        "header_index": header["header_index"],
        "has_code": bool(code_map),
        "has_barcode": bool(barcode_map),
        "code_map": code_map,
        "barcode_map": barcode_map,
        "source_key_count": len(code_map) + len(barcode_map),
        "source_code_count": len(code_map),
        "source_barcode_count": len(barcode_map),
        "duplicates": duplicates,
        "conflicts": conflicts,
        "source_rows": source_rows,
        "source_file": source_file,
    }


def find_nc_header_and_columns(ws):
    # NC表只认标准字段。不要把“客户产品编码”当成NC存货编码。
    code_aliases = [normalize_text(x) for x in ["存货编码", "物料编码", "商品编码", "产品编码", "编码", "nc存货编码", "nc编码"]]
    barcode_aliases = [normalize_text(x) for x in ["条码", "商品条码", "条形码", "69码", "69条码", "69编码", "最新69码", "新69码", "barcode", "bar code"]]
    cost_aliases = [normalize_text(x) for x in COST_ALIASES]

    best = None

    for row in range(1, min(ws.max_row, 35) + 1):
        code_col = None
        barcode_col = None
        cost_col = None

        for col in range(1, ws.max_column + 1):
            h = normalize_text(ws.cell(row=row, column=col).value)
            if h in code_aliases and code_col is None:
                code_col = col
            if h in barcode_aliases and barcode_col is None:
                barcode_col = col
            if h in cost_aliases and cost_col is None:
                cost_col = col

        score = 0
        if code_col:
            score += 3
        if barcode_col:
            score += 3
        if cost_col:
            score += 1

        if score:
            best = (row, code_col, barcode_col, cost_col, score)

        if code_col or barcode_col:
            return row, code_col, barcode_col, cost_col

    if best:
        return best[:4]

    return None, None, None, None


def get_or_create_cost_column(ws, header_row):
    cost_aliases = [normalize_text(x) for x in COST_ALIASES]

    for col in range(1, ws.max_column + 1):
        h = normalize_text(ws.cell(row=header_row, column=col).value)
        if h in cost_aliases:
            return col

    new_col = ws.max_column + 1
    ws.cell(row=header_row, column=new_col).value = "成本"

    if new_col > 1:
        src = ws.cell(row=header_row, column=new_col - 1)
        dst = ws.cell(row=header_row, column=new_col)
        if src.has_style:
            dst.font = copy(src.font)
            dst.fill = copy(src.fill)
            dst.border = copy(src.border)
            dst.alignment = copy(src.alignment)
            dst.number_format = src.number_format
            dst.protection = copy(src.protection)

    return new_col


def add_report_sheet(wb, title, headers, rows):
    if title in wb.sheetnames:
        del wb[title]

    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)

    for cell in ws[1]:
        cell.font = copy(cell.font)
        cell.font = cell.font.copy(bold=True)

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 26

    return ws


def flatten_conflicts(conflicts):
    rows = []
    for key_type, data in conflicts.items():
        for key, vals in data.items():
            rows.append(["成本冲突", key_type, key, " / ".join(sorted(list(vals)))])
    return rows


def flatten_duplicates(duplicates):
    rows = []
    for key_type, data in duplicates.items():
        for key, row_list in data.items():
            rows.append(["重复键", key_type, key, "来源表重复行：" + "、".join(map(str, row_list))])
    return rows


def merge_source_info(existing, incoming):
    """
    将多个来源表合并成一套编码/条码成本池。
    已有数据优先；新数据如与已有成本冲突，进入异常，不覆盖。
    """
    for key, val in incoming["code_map"].items():
        cost, source_file = val
        if key in existing["code_map"]:
            old_cost, old_file = existing["code_map"][key]
            if str(old_cost) != str(cost):
                existing["global_conflicts"].append([
                    "跨文件成本冲突", "编码", key,
                    f"已有：{old_cost}（{old_file}）；新增：{cost}（{source_file}）。已保留已有成本。"
                ])
        else:
            existing["code_map"][key] = val

    for key, val in incoming["barcode_map"].items():
        cost, source_file = val
        if key in existing["barcode_map"]:
            old_cost, old_file = existing["barcode_map"][key]
            if str(old_cost) != str(cost):
                existing["global_conflicts"].append([
                    "跨文件成本冲突", "条码", key,
                    f"已有：{old_cost}（{old_file}）；新增：{cost}（{source_file}）。已保留已有成本。"
                ])
        else:
            existing["barcode_map"][key] = val

    existing["file_summaries"].append([
        incoming["source_file"],
        incoming["sheet_name"],
        incoming["header_index"] + 1,
        incoming["source_code_count"],
        incoming["source_barcode_count"],
        len(flatten_conflicts(incoming["conflicts"])),
    ])

    existing["source_rows"].extend(incoming["source_rows"])
    existing["global_abnormal"].extend(flatten_conflicts(incoming["conflicts"]))
    existing["global_abnormal"].extend(flatten_duplicates(incoming["duplicates"]))


class BatchCostSession:
    def __init__(self):
        self.reset()

    def reset(self):
        self.nc_path = ""
        self.nc_wb = None
        self.ws = None
        self.header_row = None
        self.code_col = None
        self.barcode_col = None
        self.cost_col = None

        self.source_pool = {
            "code_map": {},
            "barcode_map": {},
            "file_summaries": [],
            "source_rows": [],
            "global_conflicts": [],
            "global_abnormal": [],
        }

        self.matched_rows = []
        self.unmatched_rows = []
        self.skipped_existing = []
        self.processed_file_paths = set()
        self.has_applied = False

    def load_nc(self, nc_path):
        if os.path.splitext(nc_path)[1].lower() not in [".xlsx", ".xlsm"]:
            raise ValueError("NC存货成本表请使用 .xlsx 或 .xlsm 格式。")

        self.nc_path = nc_path
        self.nc_wb = load_workbook(nc_path)
        self.ws = self.nc_wb.active

        self.header_row, self.code_col, self.barcode_col, self.cost_col = find_nc_header_and_columns(self.ws)
        if not self.header_row:
            raise ValueError("NC存货成本表未识别到“存货编码”或“条码”列。")

        self.cost_col = get_or_create_cost_column(self.ws, self.header_row)

        self.matched_rows = []
        self.unmatched_rows = []
        self.skipped_existing = []
        self.has_applied = False

    def add_source_files(self, paths):
        added = []
        for path in paths:
            abs_path = os.path.abspath(path)
            if abs_path in self.processed_file_paths:
                continue

            info = build_source_maps(path)
            merge_source_info(self.source_pool, info)
            self.processed_file_paths.add(abs_path)
            added.append(info)

        return added

    def apply_to_nc(self, overwrite=True):
        if self.nc_wb is None:
            raise ValueError("请先选择NC成本表。")

        self.matched_rows = []
        self.unmatched_rows = []
        self.skipped_existing = []

        code_map = self.source_pool["code_map"]
        barcode_map = self.source_pool["barcode_map"]

        if not code_map and not barcode_map:
            raise ValueError("请先添加至少一个成本来源表。")

        for row in range(self.header_row + 1, self.ws.max_row + 1):
            nc_code = self.ws.cell(row=row, column=self.code_col).value if self.code_col else None
            nc_barcode = self.ws.cell(row=row, column=self.barcode_col).value if self.barcode_col else None

            if not clean_key(nc_code) and not clean_key(nc_barcode):
                continue

            matched = None

            # 编码优先
            if self.code_col and code_map:
                for key in split_keys(nc_code):
                    if key in code_map:
                        cost, source_file = code_map[key]
                        matched = ("编码匹配", key, cost, source_file)
                        break

            # 条码兜底
            if matched is None and self.barcode_col and barcode_map:
                for key in split_keys(nc_barcode):
                    if key in barcode_map:
                        cost, source_file = barcode_map[key]
                        matched = ("条码匹配", key, cost, source_file)
                        break

            if matched:
                match_type, matched_key, matched_cost, source_file = matched
                target_cell = self.ws.cell(row=row, column=self.cost_col)
                old_value = target_cell.value

                if overwrite or old_value in (None, ""):
                    target_cell.value = matched_cost
                    target_cell.number_format = "0.00"
                    self.matched_rows.append([row, match_type, matched_key, matched_cost, source_file])
                else:
                    self.skipped_existing.append([row, match_type, matched_key, old_value, matched_cost, source_file])
            else:
                self.unmatched_rows.append([
                    row,
                    clean_key(nc_code) if self.code_col else "",
                    clean_key(nc_barcode) if self.barcode_col else "",
                    "当前已导入的成本来源表中未找到对应编码或条码"
                ])

        self.has_applied = True

    def export(self, output_path):
        if self.nc_wb is None:
            raise ValueError("请先选择NC成本表。")
        if not self.has_applied:
            raise ValueError("请先点击“开始回填/加入暂存”，再导出。")

        abnormal_rows = (
            self.source_pool["global_abnormal"]
            + self.source_pool["global_conflicts"]
            + [["跳过已有成本", match_type, key, f"原成本：{old}；来源成本：{new}；来源文件：{source_file}"]
               for _, match_type, key, old, new, source_file in self.skipped_existing]
        )

        add_report_sheet(
            self.nc_wb,
            "回填统计",
            ["项目", "结果"],
            [
                ["NC来源文件", os.path.basename(self.nc_path)],
                ["已导入成本来源文件数", len(self.source_pool["file_summaries"])],
                ["来源编码数量", len(self.source_pool["code_map"])],
                ["来源条码数量", len(self.source_pool["barcode_map"])],
                ["NC表匹配回填行数", len(self.matched_rows)],
                ["NC表未匹配行数", len(self.unmatched_rows)],
                ["异常记录数", len(abnormal_rows)],
                ["跳过已有成本行数", len(self.skipped_existing)],
            ],
        )

        add_report_sheet(
            self.nc_wb,
            "来源文件统计",
            ["来源文件", "识别Sheet", "表头行", "编码数量", "条码数量", "文件内成本冲突数"],
            self.source_pool["file_summaries"],
        )

        add_report_sheet(
            self.nc_wb,
            "匹配明细",
            ["NC表行号", "匹配方式", "匹配值", "回填成本", "来源文件"],
            self.matched_rows,
        )

        add_report_sheet(
            self.nc_wb,
            "未匹配明细",
            ["NC表行号", "NC存货编码", "NC条码", "原因"],
            self.unmatched_rows,
        )

        add_report_sheet(
            self.nc_wb,
            "异常明细",
            ["异常类型", "匹配字段", "匹配值", "说明"],
            abnormal_rows,
        )

        self.nc_wb.save(output_path)
        return output_path


class InventoryCostApp:
    def __init__(self, root):
        self.root = root
        self.root.title("存货成本回填生成器 - 批量暂存版")
        self.root.geometry("920x620")
        self.root.resizable(False, False)

        self.session = BatchCostSession()

        self.nc_path = tk.StringVar()
        self.output_dir = tk.StringVar()
        self.overwrite = tk.BooleanVar(value=True)

        self.build_ui()

    def build_ui(self):
        tk.Label(
            self.root,
            text="存货成本回填生成器 - 批量暂存版",
            font=("Microsoft YaHei", 18, "bold")
        ).pack(pady=(16, 5))

        tk.Label(
            self.root,
            text="先选择NC表，再批量添加多个供应商成本表；回填结果暂存在程序内，最后手动导出Excel。",
            font=("Microsoft YaHei", 10)
        ).pack(pady=(0, 14))

        top = tk.Frame(self.root)
        top.pack(fill="x", padx=28)

        self.file_row(top, "NC成本表：", self.nc_path, self.choose_nc, 0)
        self.file_row(top, "输出文件夹：", self.output_dir, self.choose_output_dir, 1)

        option_frame = tk.Frame(self.root)
        option_frame.pack(fill="x", padx=30, pady=(8, 0))

        tk.Checkbutton(
            option_frame,
            text="覆盖NC表中已有成本；如不勾选，则只填空白成本",
            variable=self.overwrite,
            font=("Microsoft YaHei", 10)
        ).pack(anchor="w")

        btn_frame = tk.Frame(self.root)
        btn_frame.pack(pady=12)

        tk.Button(
            btn_frame,
            text="批量添加成本来源表",
            command=self.add_source_files,
            width=20,
            height=2,
            font=("Microsoft YaHei", 10, "bold")
        ).pack(side="left", padx=6)

        tk.Button(
            btn_frame,
            text="开始回填 / 加入暂存",
            command=self.apply_to_nc,
            width=20,
            height=2,
            font=("Microsoft YaHei", 10, "bold")
        ).pack(side="left", padx=6)

        tk.Button(
            btn_frame,
            text="导出结果",
            command=self.export_result,
            width=14,
            height=2,
            font=("Microsoft YaHei", 10, "bold")
        ).pack(side="left", padx=6)

        tk.Button(
            btn_frame,
            text="重置",
            command=self.reset_all,
            width=10,
            height=2,
            font=("Microsoft YaHei", 10)
        ).pack(side="left", padx=6)

        list_frame = tk.Frame(self.root)
        list_frame.pack(fill="x", padx=30, pady=(4, 8))

        tk.Label(list_frame, text="已添加成本来源表：", font=("Microsoft YaHei", 10, "bold")).pack(anchor="w")

        self.source_list = tk.Listbox(list_frame, height=7, width=125)
        self.source_list.pack(fill="x", pady=(4, 0))

        self.result_text = tk.Text(self.root, height=13, width=112, font=("Consolas", 10))
        self.result_text.pack(padx=30, pady=(6, 12))
        self.write_result(
            "操作流程：\n"
            "1. 选择NC成本表。\n"
            "2. 点击“批量添加成本来源表”，可一次选择多个绘强/征图/兴发文件。\n"
            "3. 点击“开始回填 / 加入暂存”，结果只保存在当前程序内，不会立即生成Excel。\n"
            "4. 继续添加更多文件也可以，之前回填结果会保留。\n"
            "5. 最后点击“导出结果”。"
        )

    def file_row(self, parent, label, var, command, row):
        tk.Label(parent, text=label, width=12, anchor="e", font=("Microsoft YaHei", 10)).grid(row=row, column=0, pady=7)
        tk.Entry(parent, textvariable=var, width=88).grid(row=row, column=1, padx=8, pady=7)
        tk.Button(parent, text="选择", command=command, width=8).grid(row=row, column=2, pady=7)

    def choose_nc(self):
        path = filedialog.askopenfilename(
            title="选择NC存货成本表",
            filetypes=[("Excel files", "*.xlsx *.xlsm"), ("All files", "*.*")]
        )
        if path:
            try:
                self.session.load_nc(path)
                self.nc_path.set(path)
                if not self.output_dir.get():
                    self.output_dir.set(os.path.dirname(path))
                self.write_result(
                    "NC表加载成功。\n\n"
                    f"NC文件：{path}\n"
                    f"表头行：{self.session.header_row}\n"
                    f"存货编码列：{self.session.code_col or '未识别'}\n"
                    f"条码列：{self.session.barcode_col or '未识别'}\n"
                    f"成本列：{self.session.cost_col}\n"
                )
            except Exception as e:
                messagebox.showerror("NC表加载失败", str(e))
                self.write_result("NC表加载失败：\n" + traceback.format_exc())

    def choose_output_dir(self):
        path = filedialog.askdirectory(title="选择输出文件夹")
        if path:
            self.output_dir.set(path)

    def add_source_files(self):
        if self.session.nc_wb is None:
            messagebox.showerror("错误", "请先选择NC成本表。")
            return

        paths = filedialog.askopenfilenames(
            title="批量选择成本来源表",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xls"), ("All files", "*.*")]
        )
        if not paths:
            return

        try:
            added = self.session.add_source_files(paths)

            for info in added:
                self.source_list.insert(
                    "end",
                    f"{info['source_file']} | Sheet:{info['sheet_name']} | 表头:{info['header_index'] + 1} | 编码:{info['source_code_count']} | 条码:{info['source_barcode_count']}"
                )

            msg = (
                f"本次新增来源表：{len(added)} 个。\n"
                f"累计来源表：{len(self.session.source_pool['file_summaries'])} 个。\n"
                f"累计编码键：{len(self.session.source_pool['code_map'])} 个。\n"
                f"累计条码键：{len(self.session.source_pool['barcode_map'])} 个。\n"
                f"累计跨文件冲突：{len(self.session.source_pool['global_conflicts'])} 条。\n\n"
                "说明：此时还没有导出Excel。点击“开始回填 / 加入暂存”后，回填结果会暂存在当前程序内。"
            )
            self.write_result(msg)

        except Exception as e:
            messagebox.showerror("添加来源表失败", str(e))
            self.write_result("添加来源表失败：\n\n" + str(e) + "\n\n详细错误：\n" + traceback.format_exc())

    def apply_to_nc(self):
        if self.session.nc_wb is None:
            messagebox.showerror("错误", "请先选择NC成本表。")
            return

        try:
            self.session.apply_to_nc(overwrite=self.overwrite.get())

            msg = (
                "已完成回填并暂存，暂未导出Excel。\n\n"
                f"累计来源表：{len(self.session.source_pool['file_summaries'])} 个\n"
                f"来源编码数量：{len(self.session.source_pool['code_map'])}\n"
                f"来源条码数量：{len(self.session.source_pool['barcode_map'])}\n"
                f"NC表匹配回填行数：{len(self.session.matched_rows)}\n"
                f"NC表未匹配行数：{len(self.session.unmatched_rows)}\n"
                f"跳过已有成本行数：{len(self.session.skipped_existing)}\n"
                f"跨文件成本冲突：{len(self.session.source_pool['global_conflicts'])}\n\n"
                "现在可以继续添加更多成本来源表；也可以点击“导出结果”生成Excel。"
            )
            self.write_result(msg)

        except Exception as e:
            messagebox.showerror("回填失败", str(e))
            self.write_result("回填失败：\n\n" + str(e) + "\n\n详细错误：\n" + traceback.format_exc())

    def export_result(self):
        if self.session.nc_wb is None:
            messagebox.showerror("错误", "请先选择NC成本表。")
            return

        if not self.session.has_applied:
            messagebox.showerror("错误", "请先点击“开始回填 / 加入暂存”，再导出。")
            return

        outdir = self.output_dir.get().strip() or os.path.dirname(self.session.nc_path)
        nc_name = os.path.splitext(os.path.basename(self.session.nc_path))[0]
        default_name = f"{nc_name}_批量回填成本结果.xlsx"

        output_path = filedialog.asksaveasfilename(
            title="导出结果",
            initialdir=outdir,
            initialfile=default_name,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )

        if not output_path:
            return

        try:
            self.session.export(output_path)
            self.write_result(
                "导出成功。\n\n"
                f"输出文件：{output_path}\n\n"
                f"匹配回填行数：{len(self.session.matched_rows)}\n"
                f"未匹配行数：{len(self.session.unmatched_rows)}\n"
                f"来源文件数：{len(self.session.source_pool['file_summaries'])}\n"
                f"异常记录数：{len(self.session.source_pool['global_abnormal']) + len(self.session.source_pool['global_conflicts']) + len(self.session.skipped_existing)}"
            )
            messagebox.showinfo("完成", "结果已导出。")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))
            self.write_result("导出失败：\n\n" + str(e) + "\n\n详细错误：\n" + traceback.format_exc())

    def reset_all(self):
        self.session.reset()
        self.nc_path.set("")
        self.output_dir.set("")
        self.source_list.delete(0, "end")
        self.write_result("已重置。请重新选择NC表和成本来源表。")

    def write_result(self, text):
        self.result_text.config(state="normal")
        self.result_text.delete("1.0", "end")
        self.result_text.insert("end", text)
        self.result_text.config(state="disabled")



# ===== 销量预测与备货建议生成器子模块（v1.6 优化版）=====
#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
销量预测与备货建议生成器 v1.6
=====================

功能说明：
    基于两年历史销售数据，预测未来需求并生成备货建议
    支持工作日效应计算（排除周末+法定假日）
    v1.5新增：SKU分层从帕累托ABC法改为6级日销量分类法（A+/A/B/C/D/E + 零销量）
    
GUI模式（默认）：
    双击运行或 python3 销量预测与备货建议生成器.py

输入要求：
    - 两个销售订单CSV文件（2024.5-2026.4，共约40万条）
    - 一个库存周转CSV文件（可选）
    
CSV编码：UTF-8 或 GBK（自动检测）

更新日志 v1.5:
    - SKU分层改为6级日销量分类法（按工作日日均销量）
    - A+类：日销 ≥ 100，超头部SKU，Holt-Winters精细预测
    - A类：日销 ≥ 50，头部SKU，Holt-Winters精细预测
    - B类：日销 ≥ 10，中频SKU，MA×趋势×同比×季节预测
    - C类：日销 ≥ 3，低频SKU，MA×趋势×同比×季节预测
    - D类：日销 ≥ 1，零星SKU，6月移动平均
    - E类：日销 < 1，极低频SKU，6月移动平均
    - 零销量：24个月总销量=0
    - 报表Sheet按预测方法重新分组（HW精细预测/因子预测/简单MA/零销量）

更新日志 v1.4:
    - 新增趋势因子：12个月线性回归斜率，限制0.7~1.3
    - 新增Holt-Winters指数平滑（A/B类），回退到趋势×MA×同比
    - 新增同比增长率修正（YoY），限制0.8~1.2，衰减系数0.3
    - A/B类预测取max(HW, MA×趋势×同比)保守估计
    - C类保持6月移动平均不变
    - 新增产品生命周期识别：上升期/成熟期/衰退期/休眠期/新品期
    - 生命周期调整：上升期×1.15，衰退期×0.7，休眠期×0
    - 新增预测偏差回测：用最近3个月验证预测准确率，计算MAPE

更新日志 v1.3:
    - 增加工作日效应计算，使用有效工作日替代日历日
    - 排除周末和法定假日，更准确反映实际销售日均

更新日志 v1.1:
    - 修复界面按钮隐藏问题（增大窗口高度）
    - 增加商品名称列
    - 精简月度销量为年度月均（2024.5-2025.5月均、2025.5-2026.5月均）
    - 增加备货金额统计（含税单价 × 建议备货量）
    - 增加版本迭代日志展示

作者：扬帆耗材集团兼容采购部
"""

import os
import sys
import threading
import platform
import subprocess
from datetime import datetime
from typing import Optional, Tuple, Dict, List
from collections import defaultdict

try:
    import pandas as pd
    import numpy as np
except ImportError:
    print("错误：缺少 pandas 或 numpy 库，请先安装：pip install pandas numpy")
    sys.exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：缺少 openpyxl 库，请先安装：pip install openpyxl")
    sys.exit(1)

try:
    import chardet
except ImportError:
    print("警告：chardet 未安装，将使用备用编码检测")
    chardet = None

# Holt-Winters支持（可选）
try:
    from statsmodels.tsa.holtwinters import ExponentialSmoothing
    HAS_STATSMODELS = True
except ImportError:
    print("警告：statsmodels 未安装，Holt-Winters指数平滑将不可用，将使用趋势×MA×同比预测")
    HAS_STATSMODELS = False


# ============================================================================
# 配置常量
# ============================================================================

# 版本信息
VERSION = "v1.6"
VERSION_LOG = """版本迭代日志:
v1.0 - 初始版本，支持销量预测与备货建议
v1.1 - 修复界面按钮隐藏问题；增加商品名称列；精简月度销量为年度月均；增加备货金额统计；增加版本迭代日志
v1.2 - 修复打开报告按钮路径问题；按钮固定底部不被遮挡；优化列宽适配；含税单价类型转换修复
v1.3 - 增加工作日效应计算；使用有效工作日替代日历日；排除周末和法定假日；更准确反映实际销售日均
v1.4 - 新增趋势因子（12月线性回归）；新增Holt-Winters指数平滑（A/B类）；新增同比增长率修正（YoY）；A/B类取max(HW,MA×趋势×同比)；新增客户集中度风险评估（Top1/Top3占比、高/中/低风险）；新增产品生命周期识别（上升期/成熟期/衰退期/休眠期/新品期）及备货调整；新增预测偏差回测（MAPE准确率验证）
v1.5 - SKU分层改为6级日销量分类法（A+/A/B/C/D/E+零销量）；按工作日日均销量划分；Holt-Winters仅用于A+/A类；B/C类用MA×趋势×同比×季节；D/E类用6月简单MA；报表按预测方法重新分组
v1.6 - 优化CSV/Excel自动读取；支持嵌入YaFo兼容采购助理作为功能子窗口；打开子窗口不再创建第二个Tk主循环；库存文件同步支持Excel
"""

# 样式颜色配置
COLOR_HEADER_BG = "4472C4"      # 表头背景色（蓝色）
COLOR_HEADER_FONT = "FFFFFF"    # 表头字体颜色（白色）
COLOR_GAP_RED = "FFC7CE"        # 备货量>0 红色底
COLOR_ALT_ROW = "F2F2F2"        # 交替行背景色

# SKU分类阈值（按工作日日均销量）
THRESHOLD_AP = 100              # A+类：日销 ≥ 100
THRESHOLD_A = 50                # A类：日销 ≥ 50
THRESHOLD_B = 10                # B类：日销 ≥ 10
THRESHOLD_C = 3                 # C类：日销 ≥ 3
THRESHOLD_D = 1                 # D类：日销 ≥ 1
# E类：日销 < 1
WORKING_DAYS_PER_MONTH = 22    # 月工作日数（用于月均→日均换算）

# 6级分类颜色配置（aRGB格式）
CLASS_COLORS = {
    'A+': 'FF4472C4',    # 深蓝色 - A+类超头部
    'A': 'FFD6EAF8',    # 浅蓝色 - A类头部
    'B': 'FFD5F5E3',    # 浅绿色 - B类中频
    'C': 'FFFFEB9C',    # 浅黄色 - C类低频
    'D': 'FFF2F2F2',    # 浅灰色 - D类零星
    'E': 'FFE0E0E0',    # 更浅灰色 - E类极低频
}

# 默认参数
DEFAULT_TARGET_DAYS = 45        # 目标库存周转天数
DEFAULT_LEAD_TIME = 7           # 交付周期天数
DEFAULT_SAFETY_DAYS = 14        # 安全库存天数（7天交付 + 7天缓冲）

# 预测月份数
FORECAST_MONTHS = 1             # 预测下1个月

# 数值精度
DECIMAL_PLACES = 2

# Holt-Winters参数（加法模型）
HW_ALPHA = 0.3    # 水平平滑系数
HW_BETA = 0.1     # 趋势平滑系数
HW_GAMMA = 0.2    # 季节平滑系数
HW_SEASONAL_PERIOD = 12  # 季节周期（月）

# 趋势因子边界
TREND_MIN = 0.7
TREND_MAX = 1.3

# 同比修正因子边界
YOY_MIN = 0.8
YOY_MAX = 1.2
YOY_DECAY = 0.3  # 衰减系数

# 客户集中度风险系数（影响安全库存展示值）
CONCENTRATION_HIGH_RISK_FACTOR = 0.6    # 高风险：收紧到货周期保障
CONCENTRATION_MEDIUM_RISK_FACTOR = 0.8   # 中风险：适度收紧

# 生命周期阶段颜色配置（aRGB格式）
LIFECYCLE_COLORS = {
    '上升期': 'FFD6EAF8',    # 浅蓝色
    '成熟期': None,          # 无特殊着色
    '衰退期': 'FFD5D5D5',    # 浅灰色
    '休眠期': 'FFA6A6A6',    # 深灰色
    '新品期': 'FFD5F5E3',    # 浅绿色
}

# 生命周期调整系数
LIFECYCLE_ADJUSTMENT = {
    '上升期': 1.15,   # 适度激进，防止缺货错失增长
    '成熟期': 1.0,    # 不调整
    '衰退期': 0.7,    # 保守，防止压货
    '休眠期': 0.0,    # 不备货
    '新品期': 1.0,    # 不调整，保守为上
}

# 回测MAPE颜色配置（aRGB格式）
BACKTEST_COLOR_HIGH = "FFC6E0B4"  # 绿色底 - MAPE < 30% 高准确
BACKTEST_COLOR_MEDIUM = "FFEB9C"  # 黄色底 - MAPE 30%-60% 中等
BACKTEST_COLOR_LOW = "FFC7CE"     # 红色底 - MAPE > 60% 低准确


# ============================================================================
# v1.4新增：产品生命周期识别
# ============================================================================

def identify_product_lifecycle(df: pd.DataFrame, months: List[str], logger=None) -> Tuple[pd.Series, dict]:
    """
    识别每个SKU的产品生命周期阶段
    
    生命周期阶段定义：
    1. 上升期（新品爬坡）：近6个月销量持续增长（后3月均值 > 前3月均值 × 1.15）
    2. 成熟期（平稳）：不满足上升期也不满足衰退期
    3. 衰退期（持续下滑）：近6个月销量持续下降（后3月均值 < 前3月均值 × 0.85）
    4. 休眠期：最近3个月销量=0
    5. 新品期：前12个月总销量<30（极低）且近6个月有销量
    
    Args:
        df: 包含月度销量数据的DataFrame，index为商品编码
        months: 月份列表（24个月）
        logger: 日志回调函数
    
    Returns:
        (lifecycle_series, stats_dict): 
        - lifecycle_series: Series，index为商品编码，value为生命周期阶段
        - stats_dict: dict，统计各阶段SKU数量 {'上升期': X, '成熟期': Y, '衰退期': Z, '休眠期': W, '新品期': N}
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在识别产品生命周期阶段...")
    
    lifecycles = {}
    stats = {'上升期': 0, '成熟期': 0, '衰退期': 0, '休眠期': 0, '新品期': 0}
    
    for sku in df.index:
        # 获取24个月销量数据
        monthly_sales_24 = []
        for m in months:
            if m in df.columns:
                monthly_sales_24.append(safe_value(df.loc[sku, m]))
            else:
                monthly_sales_24.append(0.0)
        
        # 切片
        recent_6 = monthly_sales_24[-6:]       # 近6个月
        recent_3 = monthly_sales_24[-3:]       # 近3个月
        first_12 = monthly_sales_24[:12]       # 前12个月
        last_12 = monthly_sales_24[-12:]       # 后12个月
        first_3_of_last6 = recent_6[:3]        # 近6个月的前3个月
        last_3_of_last6 = recent_6[3:]          # 近6个月的后3个月
        
        # 判断逻辑（按优先级）
        stage = None
        
        # 1. 休眠期：最近6个月总销量=0 或 最近3个月销量=0
        if sum(recent_6) == 0 or sum(recent_3) == 0:
            stage = '休眠期'
        
        # 2. 新品期：前12个月基本无销量（<30），近6个月有销量
        elif sum(first_12) < 30 and sum(recent_6) > 0:
            stage = '新品期'
        
        # 3. 上升期：后3月均值 > 前3月均值 × 1.15
        elif len(first_3_of_last6) == 3 and sum(first_3_of_last6) > 0:
            first_3_mean = sum(first_3_of_last6) / 3
            last_3_mean = sum(last_3_of_last6) / 3 if len(last_3_of_last6) == 3 else 0
            if last_3_mean > first_3_mean * 1.15:
                stage = '上升期'
        
        # 4. 衰退期：后3月均值 < 前3月均值 × 0.85
        if stage is None:
            if len(first_3_of_last6) == 3 and sum(first_3_of_last6) > 0:
                first_3_mean = sum(first_3_of_last6) / 3
                last_3_mean = sum(last_3_of_last6) / 3 if len(last_3_of_last6) == 3 else 0
                if last_3_mean < first_3_mean * 0.85:
                    stage = '衰退期'
        
        # 5. 成熟期：其他
        if stage is None:
            stage = '成熟期'
        
        lifecycles[sku] = stage
        stats[stage] += 1
    
    lifecycle_series = pd.Series(lifecycles)
    
    log(f"  产品生命周期识别完成：上升期 {stats['上升期']} 个，成熟期 {stats['成熟期']} 个，衰退期 {stats['衰退期']} 个，休眠期 {stats['休眠期']} 个，新品期 {stats['新品期']} 个")
    
    return lifecycle_series, stats


# ============================================================================
# v1.4新增：客户集中度风险评估
# ============================================================================

def calculate_customer_concentration(sales_df: pd.DataFrame, logger=None) -> pd.DataFrame:
    """
    计算客户集中度风险评估
    
    对每个SKU：
    1. 按客户编码聚合销量
    2. 计算Top1占比：最大客户销量 / 该SKU总销量
    3. 计算Top3占比：前3大客户销量之和 / 该SKU总销量
    4. 客户数：该SKU有多少个不同客户
    5. 风险等级：
       - 高风险：Top1占比 > 50%（单客户依赖严重）
       - 中风险：Top1占比 <= 50% 且 Top3占比 > 80%（少数客户主导）
       - 低风险：其他（客户分散）
    
    Args:
        sales_df: 清洗后的销售数据DataFrame
        logger: 日志回调函数
    
    Returns:
        DataFrame，index=商品编码，columns=[客户数, Top1占比, Top3占比, 集中度风险]
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在计算客户集中度风险...")
    
    # 只对有销量记录的SKU计算
    if '商品编码' not in sales_df.columns or '客户编码' not in sales_df.columns or '数量' not in sales_df.columns:
        log("  警告：缺少必要列，无法计算客户集中度")
        return pd.DataFrame(columns=['客户数', 'Top1占比', 'Top3占比', '集中度风险'])
    
    # 确保数量列是数值类型
    sales_df = sales_df.copy()
    sales_df['数量'] = pd.to_numeric(sales_df['数量'], errors='coerce').fillna(0)
    
    # 排除零销量记录
    sales_df = sales_df[sales_df['数量'] > 0]
    
    if len(sales_df) == 0:
        log("  警告：没有有效销量数据，跳过客户集中度计算")
        return pd.DataFrame(columns=['客户数', 'Top1占比', 'Top3占比', '集中度风险'])
    
    # 按 商品编码 + 客户编码 聚合销量
    customer_sales = sales_df.groupby(['商品编码', '客户编码'])['数量'].sum().reset_index()
    
    # 按商品编码计算总销量和客户数
    sku_total = customer_sales.groupby('商品编码')['数量'].sum()
    sku_customer_count = customer_sales.groupby('商品编码')['客户编码'].nunique()
    
    # 对每个SKU，按客户销量降序排序
    customer_sales = customer_sales.sort_values(['商品编码', '数量'], ascending=[True, False])
    
    # 计算Top1和Top3占比
    results = {}
    concentration_stats = {'高风险': 0, '中风险': 0, '低风险': 0}
    
    for sku in sku_total.index:
        sku_data = customer_sales[customer_sales['商品编码'] == sku].copy()
        
        total_qty = sku_total.get(sku, 0)
        if total_qty <= 0:
            continue
        
        customer_count = sku_customer_count.get(sku, 0)
        
        # Top1占比
        if len(sku_data) > 0:
            top1_qty = sku_data.iloc[0]['数量']
            top1_ratio = top1_qty / total_qty
        else:
            top1_ratio = 0.0
        
        # Top3占比：客户数<=3时直接设为1.0
        if customer_count <= 3:
            top3_ratio = 1.0
        else:
            top3_qty = sku_data.head(3)['数量'].sum()
            top3_ratio = top3_qty / total_qty
        
        # 风险等级划分
        if top1_ratio > 0.5:
            risk_level = '高风险'
        elif top1_ratio <= 0.5 and top3_ratio > 0.8:
            risk_level = '中风险'
        else:
            risk_level = '低风险'
        
        results[sku] = {
            '客户数': int(customer_count),
            'Top1占比': round(top1_ratio, 2),
            'Top3占比': round(top3_ratio, 2),
            '集中度风险': risk_level
        }
        
        concentration_stats[risk_level] += 1
    
    result_df = pd.DataFrame(results).T
    result_df.index.name = '商品编码'
    
    log(f"  客户集中度计算完成：共 {len(result_df)} 个SKU有客户分布数据")
    log(f"  风险分布：高风险 {concentration_stats['高风险']} 个，中风险 {concentration_stats['中风险']} 个，低风险 {concentration_stats['低风险']} 个")
    
    return result_df


# ============================================================================
# v1.4新增：预测偏差回测
# ============================================================================

def calculate_backtest(df: pd.DataFrame, months: List[str], seasonal_factors: Dict[int, float],
                       sku_names: dict = None, logger=None) -> Tuple[pd.DataFrame, dict]:
    """
    预测偏差回测：用历史数据回测预测准确率
    
    回测逻辑：
    - 使用最近3个月作为回测窗口
    - 对每个月的回测：
      - 用该月之前的数据做预测（避免数据泄露）
      - 预测方法：简化版，用3月移动平均×该月季节因子（与主预测逻辑一致但不引入趋势/同比）
      - 对比预测值和实际值
    
    Args:
        df: 包含月度销量数据的DataFrame，index为商品编码
        months: 月份列表（24个月）
        seasonal_factors: 季节因子dict {月份: 因子}
        sku_names: 商品名称dict {商品编码: 名称}
        logger: 日志回调函数
    
    Returns:
        backtest_df: DataFrame，每个SKU的回测结果
        backtest_summary: dict，汇总统计 {'A': X%, 'B': Y%, 'C': Z%, '整体': W%}
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在执行预测偏差回测...")
    
    # 用最近3个月回测
    recent_3_months = months[-3:]
    
    results = []
    for sku in df.index:
        sku_result = {'商品编码': sku}
        
        # 商品名称
        if sku_names and sku in sku_names:
            sku_result['商品名称'] = sku_names[sku]
        else:
            sku_result['商品名称'] = ''
        
        # 分类
        if '分类' in df.columns:
            sku_result['分类'] = df.loc[sku, '分类']
        else:
            sku_result['分类'] = ''
        
        errors = []
        
        for target_month in recent_3_months:
            target_idx = months.index(target_month)
            
            # 用目标月之前3个月做MA
            prev_3_months = months[max(0, target_idx-3):target_idx]
            if len(prev_3_months) == 0:
                continue
            
            # 获取实际值
            actual = safe_value(df.loc[sku, target_month]) if target_month in df.columns else 0
            
            # 计算MA预测
            valid_prev = [m for m in prev_3_months if m in df.columns]
            if len(valid_prev) > 0:
                ma_pred = df.loc[sku, valid_prev].mean()
            else:
                ma_pred = 0
            
            # 乘季节因子
            month_num = int(target_month.split('-')[1])
            season_f = seasonal_factors.get(month_num, 1.0)
            predicted = ma_pred * season_f
            
            # 转换月份名称（如 2026-03 -> 2026年3月）
            month_name = get_month_name(target_month)
            
            sku_result[f'{month_name}_实际'] = actual
            sku_result[f'{month_name}_预测'] = round(predicted, 0)
            
            if actual > 0:
                error_pct = abs(actual - predicted) / actual * 100
                sku_result[f'{month_name}_偏差%'] = round(error_pct, 1)
                errors.append(error_pct)
            else:
                sku_result[f'{month_name}_偏差%'] = None
        
        # 计算平均MAPE
        if errors:
            sku_result['MAPE'] = round(sum(errors) / len(errors), 1)
        else:
            sku_result['MAPE'] = None
        results.append(sku_result)
    
    backtest_df = pd.DataFrame(results)
    
    # v1.5修改：按分类统计MAPE（6级分类）
    mape_by_class = {}
    for cls in ['A+', 'A', 'B', 'C', 'D', 'E']:
        cls_df = backtest_df[backtest_df['分类'] == cls]
        cls_df = cls_df[cls_df['MAPE'].notna()]
        if len(cls_df) > 0:
            # 加权MAPE：按实际销量加权，避免低销量SKU的极端百分比拉偏结果
            # 简化：取中位数而非均值，更抗极值
            mape_by_class[cls] = round(cls_df['MAPE'].median(), 1)
        else:
            mape_by_class[cls] = None
    
    # 整体MAPE：取中位数
    overall = backtest_df['MAPE'].dropna()
    if len(overall) > 0:
        mape_by_class['整体'] = round(overall.median(), 1)
    else:
        mape_by_class['整体'] = None
    
    log(f"  回测完成（MAPE取中位数，抗极值）：")
    for cls, mape in mape_by_class.items():
        if mape is not None:
            log(f"    {cls}类MAPE = {mape}%")
    
    return backtest_df, mape_by_class


# ============================================================================
# 工作日计算
# ============================================================================

# 中国法定假日数据（放假天数，不含周末重叠部分）
# 2024-2026年实际放假天数
HOLIDAYS = {
    2024: {1: 1, 2: 9, 4: 1, 5: 2, 6: 1, 9: 1, 10: 5},
    2025: {1: 1, 2: 7, 4: 1, 5: 2, 6: 1, 9: 1, 10: 5},
    2026: {1: 1, 2: 7, 4: 1, 5: 2, 6: 1, 9: 1, 10: 5},
}


def get_working_days(year: int, month: int) -> int:
    """
    计算指定年月的中国工作日数（排除周末+法定假日）
    
    Args:
        year: 年份，如 2024
        month: 月份，如 6
    
    Returns:
        该月的工作日数
    """
    import calendar
    
    # 获取该月的总天数和第一天的星期几
    _, num_days = calendar.monthrange(year, month)
    
    # 计算周末天数（周六+周日）
    weekend_days = 0
    for day in range(1, num_days + 1):
        weekday = calendar.weekday(year, month, day)
        if weekday >= 5:  # 周六(5)或周日(6)
            weekend_days += 1
    
    # 获取该月法定假日天数
    holiday_days = HOLIDAYS.get(year, {}).get(month, 0)
    
    # 法定假日通常与周末有重叠，简化处理：
    # 假设假日中约有2天与周末重叠，所以实际额外休息 = 假日天数 - 2
    # 工作日 = 日历天数 - 周末天数 - 实际额外休息天数
    effective_holiday_days = max(0, holiday_days - 2)
    
    # 工作日 = 总天数 - 周末天数 - 有效假日天数
    working_days = num_days - weekend_days - effective_holiday_days
    
    return max(working_days, 15)  # 至少保证15个工作日（防止异常）


def get_target_month_working_days(logger=None) -> Tuple[int, str]:
    """
    获取预测目标月的工作日数
    
    Returns:
        (工作日数, 目标月份字符串 如 "2026年6月")
    """
    def log(msg):
        if logger:
            logger(msg)
    
    now = datetime.now()
    target_month = now + pd.DateOffset(months=1)
    target_year = target_month.year
    target_mon = target_month.month
    target_month_str = target_month.strftime("%Y年%m月")
    
    working_days = get_working_days(target_year, target_mon)
    
    log(f"  目标月 {target_month_str} 的工作日数: {working_days} 天")
    
    return working_days, target_month_str


# ============================================================================
# 工具函数
# ============================================================================

def detect_csv_encoding(file_path: str) -> str:
    """
    自动检测CSV文件编码
    """
    if chardet:
        try:
            with open(file_path, 'rb') as f:
                raw_data = f.read(10000)
                result = chardet.detect(raw_data)
                encoding = result['encoding'] or 'utf-8'
                if encoding.lower() in ['gb2312', 'gb18030']:
                    encoding = 'gbk'
                return encoding
        except:
            pass
    
    # 备用检测
    encodings = ['utf-8', 'utf-8-sig', 'gbk', 'gb2312', 'gb18030']
    for encoding in encodings:
        try:
            with open(file_path, 'r', encoding=encoding) as f:
                f.read(1024)
            return encoding
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    return 'utf-8'



def read_table_auto(file_path: str, logger=None, dtype=str) -> pd.DataFrame:
    """
    自动读取CSV或Excel文件。
    - CSV/TXT：自动检测编码后读取。
    - XLSX/XLSM/XLS：读取第一个工作表。
    这样GUI里选择Excel文件时不会再被强制按CSV读取。
    """
    def log(msg):
        if logger:
            logger(msg)

    ext = os.path.splitext(file_path)[1].lower()
    if ext in [".csv", ".txt"]:
        encoding = detect_csv_encoding(file_path)
        log(f"    检测编码: {encoding}")
        return pd.read_csv(file_path, encoding=encoding, dtype=dtype)

    if ext in [".xlsx", ".xlsm", ".xls"]:
        log("    读取方式: Excel第一个工作表")
        return pd.read_excel(file_path, dtype=dtype)

    # 未知扩展名：先尝试CSV，再尝试Excel
    try:
        encoding = detect_csv_encoding(file_path)
        log(f"    尝试按CSV读取，编码: {encoding}")
        return pd.read_csv(file_path, encoding=encoding, dtype=dtype)
    except Exception:
        log("    CSV读取失败，尝试按Excel读取")
        return pd.read_excel(file_path, dtype=dtype)


def clean_numeric_value(value) -> float:
    """
    清洗数值：处理None、空字符串、科学记数法
    """
    if value is None or value == '':
        return 0.0
    
    str_value = str(value).strip()
    
    if not str_value:
        return 0.0
    
    # 去除￥符号
    str_value = str_value.replace('￥', '').replace('¥', '')
    
    # 去除千分位逗号
    str_value = str_value.replace(',', '')
    
    # 处理科学记数法
    try:
        return float(str_value)
    except ValueError:
        return 0.0


def format_number(value: float) -> str:
    """
    格式化数值：保留2位小数，但整数不显示小数部分
    """
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return f"{value:.{DECIMAL_PLACES}f}"


def safe_value(value, default=0):
    """确保数值类型安全"""
    if isinstance(value, (int, float)):
        return value
    if value is None:
        return default
    try:
        return float(value)
    except (ValueError, TypeError):
        return default


def ceil_to_int(value: float) -> int:
    """向上取整（保守估计）"""
    return int(np.ceil(value))


def get_month_list() -> List[str]:
    """获取最近24个月份列表"""
    months = []
    now = datetime.now()
    for i in range(23, -1, -1):
        d = pd.DateOffset(months=i)
        month = (now - d).strftime("%Y-%m")
        months.append(month)
    return months


def get_yearly_month_range() -> Tuple[List[str], List[str]]:
    """获取两个年度的月份范围（各12个月）"""
    months = get_month_list()
    # 第一年：months[12:24] (24个月中的前12个月)
    year1_months = months[12:24]
    # 第二年：months[0:12] (最近的12个月)
    year2_months = months[0:12]
    return year1_months, year2_months


def calculate_yearly_avg_sales(df: pd.DataFrame, year_months: List[str]) -> pd.Series:
    """计算年度月均销量（排除零销量月份）"""
    # 只取指定月份的列
    available_cols = [m for m in year_months if m in df.columns]
    if not available_cols:
        return pd.Series(0, index=df.index)
    
    # 计算有销量的月份数
    positive_months = (df[available_cols] > 0).sum(axis=1)
    positive_months = positive_months.replace(0, 1)  # 避免除以0
    
    # 计算总销量
    total_sales = df[available_cols].sum(axis=1)
    
    # 月均销量 = 总销量 / 有销量的月份数
    return total_sales / positive_months


def get_month_name(month_str: str) -> str:
    """将 YYYY-MM 转换为 YYYY年MM月"""
    year, month = month_str.split('-')
    return f"{year}年{int(month)}月"


# ============================================================================
# v1.4新增：趋势因子计算
# ============================================================================

def calculate_trend_factor(df: pd.DataFrame, months: List[str], logger=None) -> pd.Series:
    """
    计算趋势因子（基于12个月线性回归）
    
    趋势因子 = 1 + (斜率 / 近12月均值)
    限制在 TREND_MIN(0.7) ~ TREND_MAX(1.3) 之间
    
    Args:
        df: 包含月度销量数据的DataFrame，index为商品编码
        months: 月份列表（24个月）
        logger: 日志回调函数
    
    Returns:
        趋势因子Series，index为商品编码
    """
    def log(msg):
        if logger:
            logger(msg)
    
    # 最近12个月
    recent_12_months = months[-12:]
    
    log("  正在计算趋势因子（12月线性回归）...")
    
    trend_factors = {}
    valid_count = 0
    
    for sku in df.index:
        # 获取最近12个月销量
        sales_data = []
        for m in recent_12_months:
            if m in df.columns:
                sales_data.append(safe_value(df.loc[sku, m]))
            else:
                sales_data.append(0.0)
        
        sales_arr = np.array(sales_data)
        
        # 计算均值
        mean_sales = np.mean(sales_arr)
        
        if mean_sales > 0:
            # 线性回归：y = a + b*x，x为月份序号(0~11)
            x = np.arange(12)
            
            # 计算斜率 b = Σ(xi-x̄)(yi-ȳ) / Σ(xi-x̄)²
            x_mean = np.mean(x)
            y_mean = mean_sales
            
            numerator = np.sum((x - x_mean) * (sales_arr - y_mean))
            denominator = np.sum((x - x_mean) ** 2)
            
            if denominator > 0:
                slope = numerator / denominator
                
                # 趋势因子 = 1 + (斜率 / 均值)
                trend_factor = 1.0 + (slope / y_mean)
                
                # 边界限制
                trend_factor = max(TREND_MIN, min(TREND_MAX, trend_factor))
            else:
                trend_factor = 1.0
        else:
            trend_factor = 1.0
        
        trend_factors[sku] = trend_factor
        if trend_factor != 1.0:
            valid_count += 1
    
    result = pd.Series(trend_factors)
    log(f"  趋势因子计算完成：{valid_count} 个SKU有趋势（!=1.0）")
    
    return result


# ============================================================================
# v1.4新增：同比增长率修正
# ============================================================================

def calculate_yoy_correction(df: pd.DataFrame, months: List[str], logger=None) -> pd.Series:
    """
    计算同比增长率修正因子
    
    YoY增长率 = (去年同月销量 - 前年同月销量) / 前年同月销量
    修正因子 = 1 + (YoY增长率 × 衰减系数0.3)
    限制在 YOY_MIN(0.8) ~ YOY_MAX(1.2) 之间
    
    Args:
        df: 包含月度销量数据的DataFrame
        months: 月份列表（24个月）
        logger: 日志回调函数
    
    Returns:
        同比修正因子Series
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在计算同比增长率修正因子...")
    
    # 获取目标月份（当前月份+1）
    now = datetime.now()
    target_month = (now + pd.DateOffset(months=1)).strftime("%Y-%m")
    target_month_num = int(target_month.split('-')[1])  # 月份数字 1-12
    
    # 去年同月 = 12个月前
    last_year_month = (pd.Timestamp(target_month) - pd.DateOffset(months=12)).strftime("%Y-%m")
    # 前年同月 = 24个月前
    two_years_month = (pd.Timestamp(target_month) - pd.DateOffset(months=24)).strftime("%Y-%m")
    
    yoy_factors = {}
    valid_count = 0
    
    for sku in df.index:
        # 获取去年同月和前年同月销量
        last_year_sales = safe_value(df.loc[sku, last_year_month]) if last_year_month in df.columns else 0
        two_years_sales = safe_value(df.loc[sku, two_years_month]) if two_years_month in df.columns else 0
        
        if two_years_sales > 0:
            # 计算YoY增长率
            yoy_rate = (last_year_sales - two_years_sales) / two_years_sales
            
            # 修正因子 = 1 + (YoY增长率 × 衰减系数)
            yoy_factor = 1.0 + (yoy_rate * YOY_DECAY)
            
            # 边界限制
            yoy_factor = max(YOY_MIN, min(YOY_MAX, yoy_factor))
        else:
            # 没有前年同月数据（新品），修正因子=1.0
            yoy_factor = 1.0
        
        yoy_factors[sku] = yoy_factor
        if yoy_factor != 1.0:
            valid_count += 1
    
    result = pd.Series(yoy_factors)
    log(f"  同比修正因子计算完成：{valid_count} 个SKU有同比修正（!=1.0）")
    
    return result


# ============================================================================
# v1.4新增：Holt-Winters指数平滑
# ============================================================================

def apply_holt_winters(sales_series: np.ndarray, forecast_periods: int = 1, 
                       seasonal_period: int = 12, logger=None) -> Tuple[Optional[float], bool]:
    """
    对单个SKU应用Holt-Winters三参数加法模型
    
    Args:
        sales_series: 销量时间序列（至少24个月）
        forecast_periods: 预测期数
        seasonal_period: 季节周期
        logger: 日志回调
    
    Returns:
        (预测值, 是否成功)
    """
    if not HAS_STATSMODELS:
        return None, False
    
    try:
        # 数据质量检查：至少需要 seasonal_period 个月的有效数据（非零）
        # 注：不要求24个月全部非零，兼容耗材间歇性采购很常见
        valid_data = np.where(np.isnan(sales_series), 0, sales_series)
        non_zero_count = np.sum(valid_data > 0)
        
        if non_zero_count < seasonal_period:
            return None, False
        
        # 拟合Holt-Winters加法模型
        model = ExponentialSmoothing(
            sales_series,
            seasonal_periods=seasonal_period,  # 注意：statsmodels参数名是seasonal_periods（复数）
            trend='add',
            seasonal='add',
            damped_trend=False,
            initialization_method='estimated'
        )
        
        # 使用固定参数（更稳定）
        fitted_model = model.fit(
            smoothing_level=HW_ALPHA,
            smoothing_trend=HW_BETA,
            smoothing_seasonal=HW_GAMMA,
            optimized=False
        )
        
        # 预测
        forecast = fitted_model.forecast(forecast_periods)
        # forecast可能是numpy数组或pandas Series，兼容两种情况
        pred_value = forecast[0] if isinstance(forecast, np.ndarray) else forecast.iloc[0]
        return float(pred_value), True
        
    except Exception as e:
        if logger:
            logger(f"    HW拟合失败: {str(e)[:50]}")
        return None, False


def calculate_holt_winters_forecast(df: pd.DataFrame, months: List[str], 
                                      seasonal_factors: Dict[int, float],
                                      logger=None) -> Tuple[pd.Series, pd.Series]:
    """
    为A+/A类SKU计算Holt-Winters预测（v1.5修改：仅A+/A类使用HW）
    
    Args:
        df: 包含月度销量数据的DataFrame
        months: 月份列表
        seasonal_factors: 季节因子字典
        logger: 日志回调
    
    Returns:
        (hw_forecasts, hw_success) Series: HW预测值和是否成功
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在计算Holt-Winters预测（A+/A类SKU）...")
    
    if not HAS_STATSMODELS:
        log("  statsmodels未安装，跳过Holt-Winters")
        return pd.Series(dtype=float), pd.Series(dtype=bool)
    
    # 获取最近24个月数据
    recent_24_months = months[-24:]
    
    hw_forecasts = {}
    hw_success = {}
    success_count = 0
    fail_count = 0
    
    # v1.5修改：只对A+/A类计算HW
    ap_a_df = df[df['分类'].isin(['A+', 'A'])]
    
    for sku in ap_a_df.index:
        # 构建24个月销量序列
        sales_data = []
        for m in recent_24_months:
            if m in df.columns:
                sales_data.append(safe_value(df.loc[sku, m]))
            else:
                sales_data.append(0.0)
        
        sales_arr = np.array(sales_data, dtype=float)
        
        # 应用Holt-Winters
        hw_pred, success = apply_holt_winters(
            sales_arr, 
            forecast_periods=1, 
            seasonal_period=HW_SEASONAL_PERIOD,
            logger=logger
        )
        
        if success and hw_pred is not None:
            hw_forecasts[sku] = max(0, hw_pred)  # 确保非负
            hw_success[sku] = True
            success_count += 1
        else:
            hw_forecasts[sku] = 0.0
            hw_success[sku] = False
            fail_count += 1
    
    log(f"  Holt-Winters完成：{success_count}个成功，{fail_count}个回退到MA×趋势×同比")
    if fail_count > 0 and fail_count <= 5:
        log(f"  提示：HW回退原因通常是数据中零值月过多或拟合数值不稳定")
    
    return pd.Series(hw_forecasts), pd.Series(hw_success)


# ============================================================================
# 数据清洗
# ============================================================================

def clean_sales_data(df: pd.DataFrame, logger=None) -> pd.DataFrame:
    """
    数据清洗规则：
    1. 排除赠品：是否赠品 = "Y"
    2. 只保留已完成订单：订单状态 = "完成"
    3. 排除数量 <= 0 的记录
    4. 排除客户编码为空或客户名称为空的记录
    5. 排除商品编码为空的记录
    6. 只保留大类="兼容"的记录
    """
    initial_count = len(df)
    
    def log(msg):
        if logger:
            logger(msg)
    
    # 1. 排除赠品
    if '是否赠品' in df.columns:
        before = len(df)
        df = df[df['是否赠品'] != 'Y']
        log(f"  排除赠品记录: {before - len(df)} 条")
    
    # 2. 只保留已完成订单
    if '订单状态' in df.columns:
        before = len(df)
        df = df[df['订单状态'] == '完成']
        log(f"  排除未完成订单: {before - len(df)} 条")
    
    # 3. 排除数量 <= 0
    if '数量' in df.columns:
        before = len(df)
        df['数量'] = pd.to_numeric(df['数量'], errors='coerce').fillna(0)
        df = df[df['数量'] > 0]
        log(f"  排除数量<=0记录: {before - len(df)} 条")
    
    # 4. 排除客户编码/名称为空
    if '客户编码' in df.columns:
        before = len(df)
        df = df[df['客户编码'].notna() & (df['客户编码'] != '')]
        log(f"  排除客户编码为空: {before - len(df)} 条")
    
    if '客户名称' in df.columns:
        before = len(df)
        df = df[df['客户名称'].notna() & (df['客户名称'] != '')]
        log(f"  排除客户名称为空: {before - len(df)} 条")
    
    # 5. 排除商品编码为空
    if '商品编码' in df.columns:
        before = len(df)
        df = df[df['商品编码'].notna() & (df['商品编码'] != '')]
        log(f"  排除商品编码为空: {before - len(df)} 条")
    
    # 6. 只保留大类="兼容"
    if '大类' in df.columns:
        before = len(df)
        df = df[df['大类'] == '兼容']
        log(f"  排除非兼容类记录: {before - len(df)} 条")
    
    log(f"  清洗完成: {initial_count} -> {len(df)} 条 (保留 {len(df)/initial_count*100:.1f}%)")
    
    return df


def detect_anomaly_orders(df: pd.DataFrame, logger=None) -> pd.DataFrame:
    """
    检测大单异常：
    单笔数量 > 50 且 该产品客户数 < 5 且 该产品订单数 < 10 → 标记但不删除
    """
    def log(msg):
        if logger:
            logger(msg)
    
    # 找出大单（单笔数量 > 50）
    df_large = df[df['数量'] > 50].copy()
    
    if len(df_large) == 0:
        return pd.DataFrame()
    
    # 计算每个产品的客户数和订单数
    product_stats = df.groupby('商品编码').agg({
        '客户编码': 'nunique',
        '订单号': 'count'
    }).rename(columns={'客户编码': '客户数', '订单号': '订单数'})
    
    # 筛选异常大单
    df_large = df_large.merge(product_stats, on='商品编码', how='left')
    anomaly = df_large[
        (df_large['客户数'] < 5) & 
        (df_large['订单数'] < 10)
    ]
    
    log(f"  检测到大单异常记录: {len(anomaly)} 条")
    
    return anomaly


# ============================================================================
# 预测逻辑
# ============================================================================

def aggregate_monthly_sales(df: pd.DataFrame, months: List[str], logger=None) -> pd.DataFrame:
    """
    按 商品编码 + 年月 聚合月度销量
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  开始按 SKU + 月份 聚合...")
    
    # 解析日期
    if '单据日期' not in df.columns:
        log("  错误：未找到单据日期列")
        return pd.DataFrame()
    
    df = df.copy()
    df['单据日期'] = pd.to_datetime(df['单据日期'], errors='coerce')
    df['年月'] = df['单据日期'].dt.strftime('%Y-%m')
    
    # 过滤在指定月份范围内的数据
    min_month = min(months)
    max_month = max(months)
    df = df[(df['年月'] >= min_month) & (df['年月'] <= max_month)]
    
    # 按商品编码和月份聚合
    agg_df = df.groupby(['商品编码', '年月'])['数量'].sum().reset_index()
    
    # 透视表，确保每个SKU有完整的24个月
    pivot_df = agg_df.pivot(index='商品编码', columns='年月', values='数量').fillna(0)
    
    # 补全缺失月份（某些SKU某些月没数据）
    for month in months:
        if month not in pivot_df.columns:
            pivot_df[month] = 0
    
    # 排序月份列
    pivot_df = pivot_df.reindex(columns=sorted(pivot_df.columns))
    
    log(f"  聚合完成: {len(pivot_df)} 个SKU, {len(pivot_df.columns)} 个月")
    
    return pivot_df


def classify_sku(df: pd.DataFrame, months: List[str]) -> pd.DataFrame:
    """
    v1.5 SKU分层（按工作日日均销量6级分类）：
    - A+类：日销 ≥ 100（超头部SKU）
    - A类：日销 ≥ 50（头部SKU）
    - B类：日销 ≥ 10（中频SKU）
    - C类：日销 ≥ 3（低频SKU）
    - D类：日销 ≥ 1（零星SKU）
    - E类：日销 < 1（极低频SKU）
    - 零销量：24个月总销量=0
    """
    df = df.copy()
    df['总销量'] = df[months].sum(axis=1)
    df['月均销量'] = df['总销量'] / len(months)
    df['日销数量'] = df['月均销量'] / WORKING_DAYS_PER_MONTH
    
    conditions = [
        df['日销数量'] >= THRESHOLD_AP,
        (df['日销数量'] >= THRESHOLD_A) & (df['日销数量'] < THRESHOLD_AP),
        (df['日销数量'] >= THRESHOLD_B) & (df['日销数量'] < THRESHOLD_A),
        (df['日销数量'] >= THRESHOLD_C) & (df['日销数量'] < THRESHOLD_B),
        (df['日销数量'] >= THRESHOLD_D) & (df['日销数量'] < THRESHOLD_C),
        (df['日销数量'] > 0) & (df['日销数量'] < THRESHOLD_D),
        df['总销量'] == 0
    ]
    choices = ['A+', 'A', 'B', 'C', 'D', 'E', '零销量']
    df['分类'] = np.select(conditions, choices, default='E')
    
    return df


def calculate_seasonal_factors(df: pd.DataFrame, months: List[str], logger=None) -> Dict[int, float]:
    """
    计算季节因子
    - 用最近12个月数据计算
    - 2月春节效应特殊处理，季节指数不低于0.3
    """
    def log(msg):
        if logger:
            logger(msg)
    
    # 最近12个月
    recent_months = months[-12:]
    
    # 按月份聚合所有SKU的销量
    monthly_totals = df[recent_months].sum()
    
    # 计算均值
    overall_mean = monthly_totals.mean()
    
    # 计算季节因子
    seasonal_factors = {}
    month_map = {'01': 1, '02': 2, '03': 3, '04': 4, '05': 5, '06': 6,
                 '07': 7, '08': 8, '09': 9, '10': 10, '11': 11, '12': 12}
    
    for month_str in recent_months:
        month_num = month_map[month_str.split('-')[1]]
        if overall_mean > 0:
            factor = monthly_totals[month_str] / overall_mean
        else:
            factor = 1.0
        
        # 2月特殊处理：春节效应，季节指数不低于0.3
        if month_num == 2:
            factor = max(factor, 0.3)
        
        seasonal_factors[month_num] = factor
    
    log(f"  季节因子计算完成（最近12个月）")
    
    return seasonal_factors


def forecast_sku(df: pd.DataFrame, seasonal_factors: Dict[int, float], 
                 months: List[str], target_month: int, 
                 trend_factors: pd.Series = None,
                 yoy_factors: pd.Series = None,
                 hw_forecasts: pd.Series = None,
                 hw_success: pd.Series = None,
                 logger=None) -> pd.DataFrame:
    """
    v1.5预测逻辑：
    - A+/A类：max(HW预测, MA×季节因子×趋势因子×同比修正因子)
    - B/C类：MA×季节因子×趋势因子×同比修正因子
    - D/E类：6月简单移动平均
    - 零销量：0
    """
    def log(msg):
        if logger:
            logger(msg)
    
    df = df.copy()
    
    # 预测目标月份（当前月份+1）
    now = datetime.now()
    target_month_str = (now + pd.DateOffset(months=1)).strftime("%m")
    target_month_num = int(target_month_str)
    
    # 获取季节因子
    season_factor = seasonal_factors.get(target_month_num, 1.0)
    
    # 最近3个月和6个月
    recent_3_months = months[-3:]
    recent_6_months = months[-6:]
    
    # 计算移动平均
    df['3月均值'] = df[recent_3_months].mean(axis=1)
    df['6月均值'] = df[recent_6_months].mean(axis=1)
    
    # 初始化新增列
    df['趋势因子'] = 1.0
    df['同比修正因子'] = 1.0
    df['预测方法'] = 'MA'
    
    # 如果有趋势因子数据
    if trend_factors is not None and len(trend_factors) > 0:
        df['趋势因子'] = df.index.map(trend_factors).fillna(1.0)
    
    # 如果有同比修正因子数据
    if yoy_factors is not None and len(yoy_factors) > 0:
        df['同比修正因子'] = df.index.map(yoy_factors).fillna(1.0)
    
    # 计算预测
    def calc_forecast(row):
        classification = row['分类']
        
        if classification == '零销量':
            return 0, 'MA'
        
        # 获取因子
        trend = row['趋势因子'] if pd.notna(row['趋势因子']) else 1.0
        yoy = row['同比修正因子'] if pd.notna(row['同比修正因子']) else 1.0
        
        if classification in ['A+', 'A']:
            # A+/A类：MA×季节×趋势×同比（回退方案）
            ma_forecast = row['3月均值'] * season_factor * trend * yoy
            
            # 如果有HW预测且成功
            # HW已内置趋势+季节，只乘同比修正补充宏观增速
            hw_pred = 0.0
            if hw_forecasts is not None and hw_success is not None:
                sku = row.name
                if sku in hw_forecasts.index and hw_success.get(sku, False):
                    hw_pred = hw_forecasts.get(sku, 0.0)
                    if hw_pred > 0:
                        hw_final = max(0, hw_pred * yoy)  # HW预测 × 同比修正
                        return max(ma_forecast, hw_final), 'HW'
            
            return max(0, ma_forecast), 'MA×T×YoY'
        
        elif classification in ['B', 'C']:
            # B/C类：MA×季节×趋势×同比
            return max(0, row['3月均值'] * season_factor * trend * yoy), 'MA×T×YoY'
        
        elif classification in ['D', 'E']:
            # D/E类：6月简单移动平均
            return max(0, row['6月均值']), '6月MA'
        
        return 0, 'MA'
    
    # 应用预测函数
    forecasts = df.apply(calc_forecast, axis=1)
    df['预测销量_raw'] = forecasts.apply(lambda x: x[0])
    df['预测方法'] = forecasts.apply(lambda x: x[1])
    
    # 向上取整
    df['预测销量'] = df['预测销量_raw'].apply(ceil_to_int)
    
    # 标记季节因子
    df['季节因子'] = season_factor
    
    # 统计各方法使用情况
    method_counts = df[df['分类'].isin(['A+', 'A'])]['预测方法'].value_counts()
    hw_count = method_counts.get('HW', 0)
    ma_count = method_counts.get('MA×T×YoY', 0)
    ma6_count = df[df['分类'].isin(['D', 'E'])]['预测方法'].value_counts().get('6月MA', 0)
    
    log(f"  预测方法统计：A+/A类 HW={hw_count}个, MA×T×YoY={ma_count}个；D/E类 6月MA={ma6_count}个")
    
    return df


def calculate_replenishment(df: pd.DataFrame, 
                           target_working_days: int,
                           inventory_df: Optional[pd.DataFrame] = None,
                           target_days: int = DEFAULT_TARGET_DAYS,
                           lead_time: int = DEFAULT_LEAD_TIME,
                           lifecycle_series: pd.Series = None,
                           logger=None) -> pd.DataFrame:
    """
    计算备货量（v1.5修订逻辑）
    - 工作日日均预测销量 = 预测月销量 / 目标月工作日数
    - 安全库存 = 工作日日均 × 到货周期天数（覆盖到货期消耗）
    - 目标库存 = 工作日日均 × 目标周转天数 + 安全库存
    - 建议备货量 = (目标库存 + 安全库存) - 可用库存 - 采购在途
    - 根据产品生命周期阶段调整建议备货量
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("  正在计算工作日效应...")
    
    df = df.copy()
    
    # 工作日日均预测销量（v1.3: 使用目标月工作日数替代固定30天）
    df['工作日日均销量'] = df['预测销量'] / target_working_days
    df['日均销量'] = df['工作日日均销量']  # 保持列名兼容
    
    log(f"  工作日日均 = 预测销量 / {target_working_days} 工作日")
    
    # 安全库存 = 日均 × 到货周期（覆盖到货期消耗）
    df['安全库存'] = df['日均销量'] * lead_time
    
    # 目标库存 = 日均 × 目标周转天数 + 安全库存
    # 目标周转天数覆盖正常销售周期，安全库存覆盖到货周期，两者叠加
    df['目标库存量'] = df['日均销量'] * target_days + df['安全库存']
    
    # 如果有库存数据，匹配库存信息
    cost_price_map = {}  # 成本单价 map
    if inventory_df is not None and len(inventory_df) > 0:
        log("  正在匹配库存数据...")
        
        # 找到关键列
        sku_col = None
        available_col = None
        on_the_way_col = None
        stock_qty_col = None  # 结存数量
        stock_amount_col = None  # 库存金额
        
        for col in inventory_df.columns:
            if '存货编码' in col or '商品编码' in col:
                sku_col = col
            if '可用库存' in col:
                available_col = col
            elif '可用数量' in col:
                available_col = col
            elif '库存数量' in col:
                # 库存数量作为备选，优先用可用库存/可用数量
                if available_col is None:
                    available_col = col
            if '结存数量' in col:
                stock_qty_col = col
            if '采购在途' in col:
                on_the_way_col = col
            if '库存金额' in col:
                stock_amount_col = col
        
        if sku_col and available_col:
            # 同一SKU可能有多个仓库记录，按SKU聚合（求和）
            log(f"  库存列匹配: SKU={sku_col}, 可用={available_col}, 在途={on_the_way_col}")
            
            # 数值转换
            inventory_df[available_col] = pd.to_numeric(
                inventory_df[available_col].astype(str).str.replace(',', '').str.replace('￥', '').str.replace('¥', ''),
                errors='coerce').fillna(0)
            
            # 按SKU聚合（多仓库求和）
            inv_grouped = inventory_df.groupby(sku_col).agg({available_col: 'sum'}).reset_index()
            available_map = dict(zip(inv_grouped[sku_col].astype(str), 
                                    inv_grouped[available_col]))
            
            on_the_way_map = {}
            if on_the_way_col:
                inventory_df[on_the_way_col] = pd.to_numeric(
                    inventory_df[on_the_way_col].astype(str).str.replace(',', '').str.replace('￥', '').str.replace('¥', ''),
                    errors='coerce').fillna(0)
                otw_grouped = inventory_df.groupby(sku_col).agg({on_the_way_col: 'sum'}).reset_index()
                on_the_way_map = dict(zip(otw_grouped[sku_col].astype(str),
                                          otw_grouped[on_the_way_col]))
            
            # 计算成本单价 = 库存金额 / 结存数量
            if stock_amount_col and stock_qty_col:
                inventory_df[stock_amount_col] = pd.to_numeric(
                    inventory_df[stock_amount_col].astype(str).str.replace(',', '').str.replace('￥', '').str.replace('¥', ''),
                    errors='coerce').fillna(0)
                inventory_df[stock_qty_col] = pd.to_numeric(
                    inventory_df[stock_qty_col].astype(str).str.replace(',', '').str.replace('￥', '').str.replace('¥', ''),
                    errors='coerce').fillna(0)
                price_grouped = inventory_df.groupby(sku_col).agg({
                    stock_amount_col: 'sum',
                    stock_qty_col: 'sum'
                }).reset_index()
                for _, row in price_grouped.iterrows():
                    qty = row[stock_qty_col]
                    amt = row[stock_amount_col]
                    if qty > 0 and amt > 0:
                        cost_price_map[str(row[sku_col])] = amt / qty
                log(f"  匹配到成本单价的SKU: {len(cost_price_map)} 个")
            
            df['可用库存'] = df.index.astype(str).map(available_map).fillna(0)
            df['采购在途'] = df.index.astype(str).map(on_the_way_map).fillna(0)
            log(f"  匹配完成: {len(available_map)} 个SKU有库存数据")
        else:
            log("  警告: 未找到匹配的库存列，可用库存设为0")
            df['可用库存'] = 0
            df['采购在途'] = 0
    else:
        df['可用库存'] = 0
        df['采购在途'] = 0
    
    # 建议备货量 = 目标库存 - 可用库存 - 采购在途
    df['建议备货量'] = df['目标库存量'] - df['可用库存'] - df['采购在途']
    df['建议备货量'] = df['建议备货量'].apply(lambda x: max(0, ceil_to_int(x)))
    
    # 总需求 = 目标库存
    df['总需求'] = df['目标库存量']
    
    # 备货后库存周转天数 = (可用库存 + 采购在途 + 建议备货量) / 日均销量
    df['备货后周转天数'] = (df['可用库存'] + df['采购在途'] + df['建议备货量']) / df['日均销量']
    df['备货后周转天数'] = df['备货后周转天数'].apply(lambda x: round(x, 1) if x > 0 else 0)
    
    # v1.4新增：根据产品生命周期阶段调整建议备货量
    if lifecycle_series is not None and len(lifecycle_series) > 0:
        def get_lifecycle_adjustment(sku):
            if sku in lifecycle_series.index:
                stage = lifecycle_series.get(sku, '成熟期')
                return LIFECYCLE_ADJUSTMENT.get(stage, 1.0)
            return 1.0
        
        df['生命周期'] = df.index.map(lifecycle_series).fillna('成熟期')
        df['生命周期调整系数'] = df.index.map(get_lifecycle_adjustment)
        
        # 调整建议备货量
        df['建议备货量_调整前'] = df['建议备货量'].copy()
        df['建议备货量'] = (df['建议备货量'] * df['生命周期调整系数']).apply(lambda x: max(0, ceil_to_int(x)))
        
        # 休眠期：直接设建议备货量=0
        df.loc[df['生命周期'] == '休眠期', '建议备货量'] = 0
        
        # 重新计算备货后周转天数
        df['备货后周转天数'] = (df['可用库存'] + df['采购在途'] + df['建议备货量']) / df['日均销量']
        df['备货后周转天数'] = df['备货后周转天数'].apply(lambda x: round(x, 1) if x > 0 else 0)
        
        log("  生命周期调整已完成")
    else:
        df['生命周期'] = '成熟期'
        df['生命周期调整系数'] = 1.0
    
    return df, cost_price_map


# ============================================================================

def create_workbook():
    """创建工作簿"""
    wb = openpyxl.Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    return wb


def set_header_style(cell):
    """设置表头样式"""
    cell.font = Font(bold=True, color=COLOR_HEADER_FONT, size=10)
    cell.fill = PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type="solid")
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def set_data_style(cell):
    """设置数据单元格样式"""
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    cell.alignment = Alignment(horizontal='center', vertical='center')


def set_alt_row_style(cell, row_idx):
    """设置交替行样式"""
    set_data_style(cell)
    if row_idx % 2 == 0:
        cell.fill = PatternFill(start_color=COLOR_ALT_ROW, end_color=COLOR_ALT_ROW, fill_type="solid")


def adjust_column_width(ws, max_width=25):
    """自动调整列宽"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        # 只检查前100行，避免大表太慢
        for cell in column[:100]:
            try:
                if cell.value:
                    length = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        
        # 商品名称列放宽一点，编码列和数值列收窄
        header_val = str(column[0].value) if column[0].value else ''
        if '名称' in header_val:
            adjusted_width = min(max_length + 2, 35)
        elif any(kw in header_val for kw in ['编码', '分类', '品牌', '品类']):
            adjusted_width = min(max_length + 2, 15)
        else:
            adjusted_width = min(max_length + 2, max_width)
        
        ws.column_dimensions[column_letter].width = adjusted_width


def write_summary_sheet(wb, summary: dict, forecast_df: pd.DataFrame, 
                        target_month: str, months: List[str], target_working_days: int = None,
                        trend_stats: dict = None, yoy_stats: dict = None,
                        concentration_stats: dict = None,
                        lifecycle_stats: dict = None,
                        backtest_mape: dict = None):
    """生成Sheet1: 预测总览"""
    ws = wb.create_sheet("预测总览", 0)
    
    # 标题
    ws['A1'] = f"销量预测总览 - {target_month}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    # v1.4: 新功能说明
    ws['A2'] = "v1.4新增：趋势因子 + Holt-Winters + 同比增长率修正 + 客户集中度风险评估 + 预测偏差回测"
    ws['A2'].font = Font(italic=True, size=9, color="FF0066CC")
    ws.merge_cells('A2:G2')
    
    # v1.3: 工作日效应说明
    if target_working_days:
        ws['A3'] = f"工作日效应：目标月工作日 {target_working_days} 天"
        ws['A3'].font = Font(italic=True, size=9, color="FF666666")
        ws.merge_cells('A3:D3')
        start_row = 5
    else:
        start_row = 4
    
    # 汇总信息
    row = start_row
    ws.cell(row, 1, "总SKU数").font = Font(bold=True)
    ws.cell(row, 2, int(summary['total_sku']))
    set_data_style(ws.cell(row, 2))
    
    row += 1
    ws.cell(row, 1, "有历史数据SKU").font = Font(bold=True)
    ws.cell(row, 2, int(summary['with_data_sku']))
    set_data_style(ws.cell(row, 2))
    
    row += 1
    ws.cell(row, 1, "无历史数据SKU").font = Font(bold=True)
    ws.cell(row, 2, int(summary['zero_sku']))
    set_data_style(ws.cell(row, 2))
    
    row += 2
    ws.cell(row, 1, "【SKU分层统计】").font = Font(bold=True, size=11)
    row += 1
    for cls, count in summary['class_counts'].items():
        ws.cell(row, 1, f"{cls}类SKU")
        ws.cell(row, 2, int(count))
        set_data_style(ws.cell(row, 1))
        set_data_style(ws.cell(row, 2))
        row += 1
    
    row += 1
    ws.cell(row, 1, "【预测汇总】").font = Font(bold=True, size=11)
    row += 1
    ws.cell(row, 1, "预测总销量")
    ws.cell(row, 2, int(summary['total_forecast']))
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    row += 1
    ws.cell(row, 1, "建议总备货量")
    ws.cell(row, 2, int(summary['total_replenishment']))
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    row += 1
    ws.cell(row, 1, "总目标库存")
    ws.cell(row, 2, int(summary['total_target']))
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    # v1.1: 增加总备货金额
    row += 1
    ws.cell(row, 1, "总备货金额")
    ws.cell(row, 2, round(summary.get('total_replenishment_amount', 0), 2))
    ws.cell(row, 2).number_format = '#,##0.00'
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    # v1.5: 备货后库存金额
    row += 1
    ws.cell(row, 1, "备货后库存金额")
    ws.cell(row, 2, round(summary.get('total_stock_after_amount', 0), 2))
    ws.cell(row, 2).number_format = '#,##0.00'
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    # v1.4新增: 回测MAPE统计
    if backtest_mape:
        row += 2
        ws.cell(row, 1, "【预测回测MAPE】").font = Font(bold=True, size=11, color="FF0066CC")
        row += 1
        
        # v1.5修改：6级分类MAPE统计
        for cls in ['A+', 'A', 'B', 'C', 'D', 'E', '整体']:
            mape = backtest_mape.get(cls)
            if mape is not None:
                ws.cell(row, 1, f"{cls}类MAPE")
                ws.cell(row, 2, f"{mape}%")
                set_data_style(ws.cell(row, 1))
                set_data_style(ws.cell(row, 2))
                
                # MAPE着色
                if cls == '整体':
                    if mape < 30:
                        ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_HIGH, end_color=BACKTEST_COLOR_HIGH, fill_type="solid")
                        ws.cell(row, 2).font = Font(color="FF375623", bold=True)
                    elif mape < 60:
                        ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_MEDIUM, end_color=BACKTEST_COLOR_MEDIUM, fill_type="solid")
                        ws.cell(row, 2).font = Font(color="FFBF8F00", bold=True)
                    else:
                        ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_LOW, end_color=BACKTEST_COLOR_LOW, fill_type="solid")
                        ws.cell(row, 2).font = Font(color="FFCC0000", bold=True)
                row += 1
    
    # v1.4新增: 因子统计信息
    if trend_stats or yoy_stats:
        row += 1
        ws.cell(row, 1, "【v1.4新增因子统计】").font = Font(bold=True, size=11, color="FF0066CC")
        row += 1
        
        if trend_stats:
            ws.cell(row, 1, "趋势因子>1（上升）")
            ws.cell(row, 2, trend_stats.get('rising', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            row += 1
            
            ws.cell(row, 1, "趋势因子<1（下降）")
            ws.cell(row, 2, trend_stats.get('falling', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            row += 1
        
        if yoy_stats:
            ws.cell(row, 1, "YoY修正>1（增长）")
            ws.cell(row, 2, yoy_stats.get('growing', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            row += 1
            
            ws.cell(row, 1, "YoY修正<1（下降）")
            ws.cell(row, 2, yoy_stats.get('shrinking', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            row += 1
        
        # v1.4新增: 客户集中度风险统计
        if concentration_stats:
            ws.cell(row, 1, "客户集中度风险（高风险）").font = Font(color="FFCC0000")
            ws.cell(row, 2, concentration_stats.get('高风险', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            row += 1
            
            ws.cell(row, 1, "客户集中度风险（中风险）").font = Font(color="FFBF8F00")
            ws.cell(row, 2, concentration_stats.get('中风险', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            row += 1
            
            ws.cell(row, 1, "客户集中度风险（低风险）").font = Font(color="FF375623")
            ws.cell(row, 2, concentration_stats.get('低风险', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")
            row += 1
        
        # v1.4新增: 产品生命周期统计
        if lifecycle_stats:
            row += 1
            ws.cell(row, 1, "产品生命周期（上升期）").font = Font(color="FF0070C0")
            ws.cell(row, 2, lifecycle_stats.get('上升期', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFD6EAF8", end_color="FFD6EAF8", fill_type="solid")
            row += 1
            
            ws.cell(row, 1, "产品生命周期（成熟期）")
            ws.cell(row, 2, lifecycle_stats.get('成熟期', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            row += 1
            
            ws.cell(row, 1, "产品生命周期（衰退期）").font = Font(color="FF666666")
            ws.cell(row, 2, lifecycle_stats.get('衰退期', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFD5D5D5", end_color="FFD5D5D5", fill_type="solid")
            row += 1
            
            ws.cell(row, 1, "产品生命周期（休眠期）").font = Font(color="FF333333", bold=True)
            ws.cell(row, 2, lifecycle_stats.get('休眠期', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")
            row += 1
            
            ws.cell(row, 1, "产品生命周期（新品期）").font = Font(color="FF375623")
            ws.cell(row, 2, lifecycle_stats.get('新品期', 0))
            set_data_style(ws.cell(row, 1))
            set_data_style(ws.cell(row, 2))
            ws.cell(row, 2).fill = PatternFill(start_color="FFD5F5E3", end_color="FFD5F5E3", fill_type="solid")
    
    # 说明文档
    row += 2
    ws.cell(row, 1, "【分类与预测说明】").font = Font(bold=True, size=11)
    row += 1
    
    explanations = [
        # v1.5修改：6级日销量分类说明
        ("SKU分层规则（按工作日日均销量6级分类）", ""),
        ("A+类", f"日销数量 ≥ {THRESHOLD_AP}，超头部SKU，Holt-Winters精细预测"),
        ("A类", f"日销数量 ≥ {THRESHOLD_A}，头部SKU，Holt-Winters精细预测"),
        ("B类", f"日销数量 ≥ {THRESHOLD_B}，中频SKU，MA×趋势×同比×季节预测"),
        ("C类", f"日销数量 ≥ {THRESHOLD_C}，低频SKU，MA×趋势×同比×季节预测"),
        ("D类", f"日销数量 ≥ {THRESHOLD_D}，零星SKU，6月移动平均"),
        ("E类", f"日销数量 < {THRESHOLD_D}，极低频SKU，6月移动平均"),
        ("零销量", "24个月总销量=0，不备货"),
        ("", ""),
        ("预测方法说明", ""),
        ("Holt-Winters", "三参数指数平滑（水平+趋势+季节），仅用于A+/A类"),
        ("MA×T×YoY", "3月均值×季节因子×趋势因子×同比修正，用于B/C类"),
        ("6月MA", "6月简单移动平均，用于D/E类"),
        ("", ""),
        ("预测因子说明", ""),
        ("季节因子", "最近12个月该月销量 / 月均销量，反映季节性波动"),
        ("趋势因子", "12个月线性回归斜率/均值+1，0.7~1.3区间，反映涨跌趋势"),
        ("同比修正因子", "1 + (去年同月vs前年同月增速 × 0.3)，0.8~1.2区间，反映宏观增速"),
        ("", ""),
        ("备货计算逻辑", ""),
        ("目标库存", "日均销量 × 目标周转天数 + 安全库存（到货周期影响备货量）"),
        ("安全库存", "日均销量 × 到货周期天数（覆盖到货期消耗，影响备货量计算）"),
        ("建议备货量", "(目标库存 + 安全库存) - 可用库存 - 采购在途"),
        ("备货后周转天数", "(可用库存 + 采购在途 + 建议备货量) / 日均销量"),
        ("备货后库存金额", "(可用库存 + 采购在途 + 建议备货量) × 成本单价（库存金额/结存数量）"),
        ("", ""),
        ("客户集中度风险", ""),
        ("高风险", "Top1客户销量占比 > 50%，单客户依赖严重，预测极不稳定"),
        ("中风险", "Top1占比 ≤ 50% 且 Top3占比 > 80%，少数客户主导"),
        ("低风险", "客户分散，预测相对稳定"),
        ("", ""),
        ("产品生命周期", ""),
        ("上升期", "后3月均值 > 前3月均值×1.15，销量持续增长，备货量×1.15（适度激进）"),
        ("成熟期", "不满足上升期也不满足衰退期，备货量×1.0（正常备货）"),
        ("衰退期", "后3月均值 < 前3月均值×0.85，销量持续下滑，备货量×0.7（保守防压货）"),
        ("休眠期", "最近3个月销量=0，备货量=0（不备货）"),
        ("新品期", "前12月销量<30且近6月有销量，备货量×1.0（谨慎备货）"),
        ("", ""),
        ("预测偏差回测（v1.4新增）", ""),
        ("回测方法", "用最近3个月验证预测准确率，用该月之前3个月MA×季节因子预测，对比实际值"),
        ("MAPE", "平均绝对百分比误差 = mean(|实际-预测|/|实际|) × 100%"),
        ("高准确", "MAPE < 30%，绿色底"),
        ("中等准确", "MAPE 30%-60%，黄色底"),
        ("低准确", "MAPE > 60%，红色底"),
    ]
    
    for label, desc in explanations:
        if label == "" and desc == "":
            row += 1
            continue
        if desc == "":
            # 小标题
            ws.cell(row, 1, label).font = Font(bold=True, size=10)
        else:
            ws.cell(row, 1, label)
            ws.cell(row, 2, desc)
            ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical='center')
        row += 1
    
    adjust_column_width(ws)


def write_forecast_sheet(wb, df: pd.DataFrame, sheet_name: str, 
                         target_classes: list, months: List[str]):
    """
    v1.5: 生成预测Sheet（按分类分组）
    - HW精细预测（A+/A类）：Holt-Winters或回退到MA×趋势×同比×季节
    - 因子预测（B/C类）：MA×趋势×同比×季节
    - 简单MA预测（D/E类）：6月简单移动平均
    """
    ws = wb.create_sheet(sheet_name)
    
    # 筛选对应分类
    filtered = df[df['分类'].isin(target_classes)].copy()
    
    if len(filtered) == 0:
        ws['A1'] = f"无{'/'.join(target_classes)}类产品"
        adjust_column_width(ws)
        return
    
    # 按预测销量降序
    filtered = filtered.sort_values('预测销量', ascending=False)
    
    # v1.5: 统一表头 - 含分类列（用于颜色区分）、趋势因子、同比修正因子、预测方法、客户集中度列、生命周期列
    headers = [
        '商品编码', '商品名称', 
        '2024.5-2025.5月均', '2025.5-2026.5月均',
        '分类', '预测销量', '季节因子',
        '趋势因子', '同比修正因子', '预测方法',
        '客户数', 'Top1占比', 'Top3占比', '集中度风险',
        '生命周期', '生命周期调整系数',
        '含税单价', '成本单价', '备货金额', '备货后库存金额',
        '工作日日均销量', '日均销量', '目标库存量', '安全库存', '可用库存', '采购在途', '建议备货量', '备货后周转天数'
    ]
    
    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        set_header_style(cell)
    
    # 写数据
    for row_idx, (idx, row) in enumerate(filtered.iterrows(), 2):
        cls_value = row.get('分类', '')
        
        # v1.5: 根据分类设置行背景色
        row_bg_color = CLASS_COLORS.get(cls_value, COLOR_ALT_ROW)
        
        values = [
            idx,  # 商品编码
            row.get('商品名称', ''),  # 商品名称
            safe_value(row.get('2024.5-2025.5月均', 0)),  # 年度月均1
            safe_value(row.get('2025.5-2026.5月均', 0)),  # 年度月均2
            cls_value,
            safe_value(row.get('预测销量', 0)),
            safe_value(row.get('季节因子', 1.0)),
            safe_value(row.get('趋势因子', 1.0)),
            safe_value(row.get('同比修正因子', 1.0)),
            row.get('预测方法', 'MA'),
            int(row.get('客户数', 0)) if pd.notna(row.get('客户数')) else 0,  # 客户数
            safe_value(row.get('Top1占比', 0)),  # Top1占比
            safe_value(row.get('Top3占比', 0)),  # Top3占比
            row.get('集中度风险', ''),  # 集中度风险
            row.get('生命周期', '成熟期'),  # 生命周期
            safe_value(row.get('生命周期调整系数', 1.0)),  # 生命周期调整系数
            safe_value(row.get('含税单价', 0)),  # 含税单价
            safe_value(row.get('成本单价', 0)),  # 成本单价
            safe_value(row.get('备货金额', 0)),  # 备货金额
            safe_value(row.get('备货后库存金额', 0)),  # 备货后库存金额
            safe_value(row.get('工作日日均销量', 0)),
            safe_value(row.get('日均销量', 0)),
            safe_value(row.get('目标库存量', 0)),
            safe_value(row.get('安全库存', 0)),
            safe_value(row.get('可用库存', 0)),
            safe_value(row.get('采购在途', 0)),
            safe_value(row.get('建议备货量', 0)),
            safe_value(row.get('备货后周转天数', 0)),
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col)
            cell.value = value
            
            # v1.5: 使用分类背景色
            cell.fill = PatternFill(start_color=row_bg_color, end_color=row_bg_color, fill_type="solid")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # 数值格式化
            header = headers[col - 1]
            if header in ['2024.5-2025.5月均', '2025.5-2026.5月均', '预测销量', 
                         '工作日日均销量', '日均销量',
                         '目标库存量', '安全库存', '可用库存', '采购在途', '含税单价', '成本单价', '备货金额', '备货后库存金额']:
                cell.number_format = '#,##0.00'
            elif header == '备货后周转天数':
                cell.number_format = '0.0'
            elif header in ['季节因子', '趋势因子', '同比修正因子', 'Top1占比', 'Top3占比']:
                cell.number_format = '0.00'
            
            # 建议备货量 > 0 标红
            if header == '建议备货量' and isinstance(value, (int, float)) and value > 0:
                cell.fill = PatternFill(start_color=COLOR_GAP_RED, end_color=COLOR_GAP_RED, fill_type="solid")
            
            # 备货后周转天数着色：低于目标天数一半标红，低于目标天数标黄
            if header == '备货后周转天数' and isinstance(value, (int, float)):
                if value > 0 and value < 15:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色
                elif value > 0 and value < 30:
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色
            
            # 预测方法列特殊颜色
            if header == '预测方法':
                if value == 'HW':
                    cell.font = Font(bold=True, color="FF0066CC")  # Holt-Winters蓝色
                elif value == 'MA×T×YoY':
                    cell.font = Font(color="FF375623")  # 绿色
                elif value == '6月MA':
                    cell.font = Font(color="FF666666")  # 灰色
            
            # Top1占比>50%标红色字体
            if header == 'Top1占比' and isinstance(value, (int, float)) and value > 0.5:
                cell.font = Font(color="FFCC0000", bold=True)
            
            # 集中度风险列着色
            if header == '集中度风险':
                if value == '高风险':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色底
                    cell.font = Font(color="FFCC0000", bold=True)
                elif value == '中风险':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色底
                    cell.font = Font(color="FFBF8F00", bold=True)
                elif value == '低风险':
                    cell.fill = PatternFill(start_color="FFC6E0B4", end_color="FFC6E0B4", fill_type="solid")  # 绿色底
                    cell.font = Font(color="FF375623", bold=True)
            
            # 生命周期列着色
            if header == '生命周期':
                stage = value if isinstance(value, str) else '成熟期'
                if stage == '上升期':
                    cell.fill = PatternFill(start_color="FFD6EAF8", end_color="FFD6EAF8", fill_type="solid")  # 浅蓝色
                    cell.font = Font(color="FF0070C0", bold=True)
                elif stage == '衰退期':
                    cell.fill = PatternFill(start_color="FFD5D5D5", end_color="FFD5D5D5", fill_type="solid")  # 浅灰色
                    cell.font = Font(color="FF666666")
                elif stage == '休眠期':
                    cell.fill = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")  # 深灰色
                    cell.font = Font(color="FFFFFFFF", bold=True)  # 白色字体
                elif stage == '新品期':
                    cell.fill = PatternFill(start_color="FFD5F5E3", end_color="FFD5F5E3", fill_type="solid")  # 浅绿色
                    cell.font = Font(color="FF375623", bold=True)
    
    adjust_column_width(ws)


def write_zero_sku_sheet(wb, df: pd.DataFrame, months: List[str]):
    """生成Sheet: 零销量SKU"""
    ws = wb.create_sheet("零销量SKU")
    
    # 筛选零销量
    filtered = df[df['分类'] == '零销量'].copy()
    
    if len(filtered) == 0:
        ws['A1'] = "无零销量SKU"
        adjust_column_width(ws)
        return
    
    # 写表头 - v1.1: 添加商品名称
    headers = ['商品编码', '商品名称', '24个月总销量', '月均销量', '分类']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        set_header_style(cell)
    
    # 写数据
    for row_idx, (idx, row) in enumerate(filtered.iterrows(), 2):
        values = [
            idx,
            row.get('商品名称', ''),
            safe_value(row.get('总销量', 0)),
            safe_value(row.get('月均销量', 0)),
            row.get('分类', ''),
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col)
            set_alt_row_style(cell, row_idx)
            cell.value = value
    
    adjust_column_width(ws)


def write_anomaly_sheet(wb, anomaly_df: pd.DataFrame):
    """生成Sheet: 大单异常记录"""
    ws = wb.create_sheet("大单异常记录")
    
    if len(anomaly_df) == 0:
        ws['A1'] = "无大单异常记录"
        adjust_column_width(ws)
        return
    
    # 写表头
    headers = ['商品编码', '商品名称', '订单号', '单据日期', '数量', '客户编码', '客户名称', '客户数', '订单数']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(1, col, header)
        set_header_style(cell)
    
    # 写数据
    for row_idx, (_, row) in enumerate(anomaly_df.iterrows(), 2):
        values = [
            row.get('商品编码', ''),
            row.get('商品名称', ''),
            row.get('订单号', ''),
            str(row.get('单据日期', '')),
            safe_value(row.get('数量', 0)),
            row.get('客户编码', ''),
            row.get('客户名称', ''),
            safe_value(row.get('客户数', 0)),
            safe_value(row.get('订单数', 0)),
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row_idx, col)
            set_alt_row_style(cell, row_idx)
            cell.value = value
    
    adjust_column_width(ws)


def write_seasonal_sheet(wb, seasonal_factors: Dict[int, float]):
    """生成Sheet: 季节因子表"""
    ws = wb.create_sheet("季节因子表")
    
    # 标题
    ws['A1'] = "月度季节因子"
    ws['A1'].font = Font(bold=True, size=12)
    
    # 表头
    headers = ['月份', '季节因子', '说明']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(2, col, header)
        set_header_style(cell)
    
    # 月份名称
    month_names = ['1月', '2月', '3月', '4月', '5月', '6月',
                   '7月', '8月', '9月', '10月', '11月', '12月']
    
    # 写数据
    for i, month_name in enumerate(month_names, 1):
        row = i + 2
        factor = seasonal_factors.get(i, 1.0)
        
        ws.cell(row, 1, month_name)
        ws.cell(row, 2, round(factor, 4))
        ws.cell(row, 2).number_format = '0.0000'
        
        # 说明
        if factor > 1.1:
            desc = "旺季"
        elif factor < 0.9:
            desc = "淡季"
        else:
            desc = "正常"
        ws.cell(row, 3, desc)
        
        for col in range(1, 4):
            set_data_style(ws.cell(row, col))
    
    adjust_column_width(ws)


def write_backtest_sheet(wb, backtest_df: pd.DataFrame, backtest_mape: dict, 
                          logger=None):
    """生成Sheet: 预测回测"""
    def log(msg):
        if logger:
            logger(msg)
    
    ws = wb.create_sheet("预测回测")
    
    # 按MAPE降序排列（最不准确的排最前面）
    backtest_df = backtest_df.copy()
    backtest_df = backtest_df.sort_values('MAPE', ascending=False, na_position='last')
    
    # 获取回测月份列名（用于表头）
    month_cols = [col for col in backtest_df.columns if col.endswith('_实际')]
    
    # ===== 顶部汇总区域 =====
    ws['A1'] = "预测偏差回测汇总"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')
    
    ws['A2'] = "回测方法：用该月之前3个月MA×季节因子预测，对比实际值"
    ws['A2'].font = Font(italic=True, size=9, color="FF666666")
    ws.merge_cells('A2:G2')
    
    row = 4
    ws.cell(row, 1, "【MAPE统计（6级分类）】").font = Font(bold=True, size=11)
    row += 1
    
    # v1.5修改：6级分类MAPE
    for cls in ['A+', 'A', 'B', 'C', 'D', 'E', '整体']:
        mape = backtest_mape.get(cls)
        ws.cell(row, 1, f"{cls}类MAPE")
        if mape is not None:
            ws.cell(row, 2, f"{mape}%")
        else:
            ws.cell(row, 2, "N/A")
        set_data_style(ws.cell(row, 1))
        set_data_style(ws.cell(row, 2))
        
        # MAPE着色
        if mape is not None:
            if mape < 30:
                ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_HIGH, end_color=BACKTEST_COLOR_HIGH, fill_type="solid")
                ws.cell(row, 2).font = Font(color="FF375623", bold=True)
            elif mape < 60:
                ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_MEDIUM, end_color=BACKTEST_COLOR_MEDIUM, fill_type="solid")
                ws.cell(row, 2).font = Font(color="FFBF8F00", bold=True)
            else:
                ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_LOW, end_color=BACKTEST_COLOR_LOW, fill_type="solid")
                ws.cell(row, 2).font = Font(color="FFCC0000", bold=True)
        row += 1
    
    # 着色说明
    row += 1
    ws.cell(row, 1, "【着色说明】").font = Font(bold=True, size=10)
    row += 1
    
    ws.cell(row, 1, "MAPE < 30%")
    ws.cell(row, 2, "高准确")
    ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_HIGH, end_color=BACKTEST_COLOR_HIGH, fill_type="solid")
    ws.cell(row, 2).font = Font(color="FF375623", bold=True)
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    row += 1
    
    ws.cell(row, 1, "MAPE 30%-60%")
    ws.cell(row, 2, "中等准确")
    ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_MEDIUM, end_color=BACKTEST_COLOR_MEDIUM, fill_type="solid")
    ws.cell(row, 2).font = Font(color="FFBF8F00", bold=True)
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    row += 1
    
    ws.cell(row, 1, "MAPE > 60%")
    ws.cell(row, 2, "低准确")
    ws.cell(row, 2).fill = PatternFill(start_color=BACKTEST_COLOR_LOW, end_color=BACKTEST_COLOR_LOW, fill_type="solid")
    ws.cell(row, 2).font = Font(color="FFCC0000", bold=True)
    set_data_style(ws.cell(row, 1))
    set_data_style(ws.cell(row, 2))
    
    # ===== 回测明细表格 =====
    start_row = row + 2
    headers = ['商品编码', '商品名称', '分类']
    
    # 添加3个月的列
    for month_col in month_cols:
        month_name = month_col.replace('_实际', '')
        headers.extend([
            f'{month_name}_实际',
            f'{month_name}_预测',
            f'{month_name}_偏差%'
        ])
    headers.append('MAPE')
    
    # 写表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        set_header_style(cell)
    
    # 写数据
    for row_idx, (_, data_row) in enumerate(backtest_df.iterrows(), start_row + 1):
        col_idx = 1
        
        # 商品编码
        cell = ws.cell(row_idx, col_idx, data_row.get('商品编码', ''))
        set_alt_row_style(cell, row_idx)
        col_idx += 1
        
        # 商品名称
        cell = ws.cell(row_idx, col_idx, data_row.get('商品名称', ''))
        set_alt_row_style(cell, row_idx)
        col_idx += 1
        
        # 分类
        cell = ws.cell(row_idx, col_idx, data_row.get('分类', ''))
        set_alt_row_style(cell, row_idx)
        col_idx += 1
        
        # 3个月的回测数据
        for month_col in month_cols:
            month_name = month_col.replace('_实际', '')
            
            # 实际值
            cell = ws.cell(row_idx, col_idx, data_row.get(f'{month_name}_实际', 0))
            set_alt_row_style(cell, row_idx)
            col_idx += 1
            
            # 预测值
            cell = ws.cell(row_idx, col_idx, data_row.get(f'{month_name}_预测', 0))
            set_alt_row_style(cell, row_idx)
            col_idx += 1
            
            # 偏差%
            error_val = data_row.get(f'{month_name}_偏差%')
            cell = ws.cell(row_idx, col_idx, error_val)
            set_alt_row_style(cell, row_idx)
            if error_val is not None:
                if error_val < 30:
                    cell.fill = PatternFill(start_color=BACKTEST_COLOR_HIGH, end_color=BACKTEST_COLOR_HIGH, fill_type="solid")
                elif error_val < 60:
                    cell.fill = PatternFill(start_color=BACKTEST_COLOR_MEDIUM, end_color=BACKTEST_COLOR_MEDIUM, fill_type="solid")
                else:
                    cell.fill = PatternFill(start_color=BACKTEST_COLOR_LOW, end_color=BACKTEST_COLOR_LOW, fill_type="solid")
            col_idx += 1
        
        # MAPE
        mape_val = data_row.get('MAPE')
        cell = ws.cell(row_idx, col_idx, mape_val)
        set_alt_row_style(cell, row_idx)
        if mape_val is not None:
            if mape_val < 30:
                cell.fill = PatternFill(start_color=BACKTEST_COLOR_HIGH, end_color=BACKTEST_COLOR_HIGH, fill_type="solid")
                cell.font = Font(color="FF375623", bold=True)
            elif mape_val < 60:
                cell.fill = PatternFill(start_color=BACKTEST_COLOR_MEDIUM, end_color=BACKTEST_COLOR_MEDIUM, fill_type="solid")
                cell.font = Font(color="FFBF8F00", bold=True)
            else:
                cell.fill = PatternFill(start_color=BACKTEST_COLOR_LOW, end_color=BACKTEST_COLOR_LOW, fill_type="solid")
                cell.font = Font(color="FFCC0000", bold=True)
    
    log("  ✓ Sheet: 预测回测")
    
    adjust_column_width(ws)


# ============================================================================
# 主程序
# ============================================================================

def run_prediction(sales_files: List[str], inventory_file: Optional[str] = None,
                  target_days: int = DEFAULT_TARGET_DAYS,
                  lead_time: int = DEFAULT_LEAD_TIME,
                  logger=None):
    """
    执行销量预测主流程 v1.4
    """
    def log(msg):
        if logger:
            logger(msg)
    
    log("=" * 50)
    log("销量预测与备货建议生成器 v1.6")
    log("=" * 50)
    
    # 检查statsmodels
    if HAS_STATSMODELS:
        log("✓ statsmodels已安装，支持Holt-Winters指数平滑")
    else:
        log("⚠ statsmodels未安装，将使用MA×趋势×同比预测")
    
    # 步骤1: 读取并合并销售数据
    log("\n【步骤1】读取销售数据...")
    all_sales = []
    for i, file_path in enumerate(sales_files, 1):
        if not os.path.exists(file_path):
            log(f"  文件不存在: {file_path}")
            continue
        
        log(f"  读取文件 {i}: {os.path.basename(file_path)}")
        try:
            df = read_table_auto(file_path, logger=logger, dtype=str)
            log(f"    读取 {len(df)} 条记录")
            all_sales.append(df)
        except Exception as e:
            log(f"    读取失败: {e}")
            continue
    
    if not all_sales:
        log("错误：没有成功读取任何销售数据文件")
        return None, None
    
    # 合并数据
    sales_df = pd.concat(all_sales, ignore_index=True)
    log(f"  合并后总计: {len(sales_df)} 条记录")
    
    # 步骤2: 数据清洗
    log("\n【步骤2】数据清洗...")
    sales_df = clean_sales_data(sales_df, logger)
    
    # 步骤3: 检测大单异常
    log("\n【步骤3】检测大单异常...")
    anomaly_df = detect_anomaly_orders(sales_df.copy(), logger)
    
    # 步骤4: 计算客户集中度风险（新增）
    log("\n【步骤4】计算客户集中度风险...")
    concentration_df = calculate_customer_concentration(sales_df.copy(), logger)
    
    # 步骤5: 按SKU聚合月度销量
    log("\n【步骤5】按SKU聚合月度销量...")
    months = get_month_list()
    monthly_df = aggregate_monthly_sales(sales_df, months, logger)
    
    # 步骤6: SKU分层
    log("\n【步骤6】SKU分层（6级日销量分类）...")
    classified_df = classify_sku(monthly_df, months)
    
    class_counts = classified_df['分类'].value_counts().to_dict()
    for cls in ['A+', 'A', 'B', 'C', 'D', 'E', '零销量']:
        count = class_counts.get(cls, 0)
        log(f"  {cls}类SKU: {count} 个")
    
    # 步骤7: 计算季节因子
    log("\n【步骤7】计算季节因子...")
    seasonal_factors = calculate_seasonal_factors(classified_df, months, logger)
    
    # v1.4新增步骤: 计算趋势因子
    log("\n【步骤8】计算趋势因子（12月线性回归）...")
    trend_factors = calculate_trend_factor(classified_df, months, logger)
    
    # v1.4新增步骤: 计算同比增长率修正
    log("\n【步骤9】计算同比增长率修正...")
    yoy_factors = calculate_yoy_correction(classified_df, months, logger)
    
    # v1.4新增步骤: 识别产品生命周期
    log("\n【步骤10】识别产品生命周期阶段...")
    lifecycle_series, lifecycle_stats = identify_product_lifecycle(classified_df, months, logger)
    
    # v1.5修改步骤: Holt-Winters预测（仅A+/A类）
    log("\n【步骤11】Holt-Winters指数平滑预测（A+/A类）...")
    hw_forecasts, hw_success = calculate_holt_winters_forecast(
        classified_df, months, seasonal_factors, logger
    )
    
    # 步骤12: 预测
    log("\n【步骤12】预测下月销量...")
    forecast_df = forecast_sku(
        classified_df, seasonal_factors, months, 1,
        trend_factors, yoy_factors, hw_forecasts, hw_success,
        logger
    )
    
    # 步骤13: 读取库存数据（如有）
    inventory_df = None
    if inventory_file and os.path.exists(inventory_file):
        log("\n【步骤13】读取库存数据...")
        try:
            inventory_df = read_table_auto(inventory_file, logger=logger, dtype=str)
            log(f"  读取 {len(inventory_df)} 条库存记录")
        except Exception as e:
            log(f"  读取失败: {e}")
    
    # 步骤14: 获取目标月工作日数
    log("\n【步骤14】计算工作日效应...")
    target_working_days, target_month = get_target_month_working_days(logger)
    
    # 步骤15: 计算备货量（传入生命周期数据）
    log("\n【步骤15】计算备货量（含生命周期调整）...")
    result_df, cost_price_map = calculate_replenishment(forecast_df, target_working_days, inventory_df, target_days, lead_time, lifecycle_series, logger)
    
    # 将客户集中度数据合并到result_df（新增步骤）
    log("\n【步骤16】合并客户集中度数据...")
    if len(concentration_df) > 0:
        concentration_df = concentration_df.reindex(result_df.index)
        result_df['客户数'] = concentration_df['客户数'].astype('Int64')  # 兼容NaN
        result_df['Top1占比'] = concentration_df['Top1占比']
        result_df['Top3占比'] = concentration_df['Top3占比']
        result_df['集中度风险'] = concentration_df['集中度风险'].fillna('')
        log(f"  合并完成：{result_df['集中度风险'].notna().sum()} 个SKU有集中度数据")
    else:
        result_df['客户数'] = 0
        result_df['Top1占比'] = 0.0
        result_df['Top3占比'] = 0.0
        result_df['集中度风险'] = ''
    
    # 添加商品名称（从原始销售数据）
    log("\n【步骤17】匹配商品名称...")
    sku_names = {}
    if '商品名称' in sales_df.columns:
        sku_names_series = sales_df.groupby('商品编码')['商品名称'].first()
        sku_names = sku_names_series.to_dict()
        result_df['商品名称'] = result_df.index.map(sku_names)
    else:
        result_df['商品名称'] = result_df.index
    
    # 添加含税单价（从销售数据取最近一次成交的含税单价）
    log("\n【步骤18】匹配含税单价...")
    if '含税单价' in sales_df.columns and '单据日期' in sales_df.columns:
        # 转换日期并按时间排序，取每个商品编码最后一次的含税单价
        sales_df['单据日期'] = pd.to_datetime(sales_df['单据日期'], errors='coerce')
        sales_df['含税单价'] = pd.to_numeric(sales_df['含税单价'].astype(str).str.replace(',', '').str.replace('￥', '').str.replace('¥', ''), errors='coerce').fillna(0)
        latest_prices = sales_df[sales_df['含税单价'] > 0].dropna(subset=['单据日期']).sort_values('单据日期')
        latest_prices = latest_prices.groupby('商品编码')['含税单价'].last()
        result_df['含税单价'] = result_df.index.map(latest_prices).fillna(0)
        log(f"  匹配到含税单价的SKU: {(result_df['含税单价'] > 0).sum()} 个")
    else:
        result_df['含税单价'] = 0
    
    # 计算两个年度月均销量
    log("\n【步骤19】计算年度月均销量...")
    year1_months, year2_months = get_yearly_month_range()
    result_df['2024.5-2025.5月均'] = calculate_yearly_avg_sales(monthly_df, year1_months)
    result_df['2025.5-2026.5月均'] = calculate_yearly_avg_sales(monthly_df, year2_months)
    
    # 计算备货金额（用成本单价）
    result_df['成本单价'] = result_df.index.astype(str).map(cost_price_map).fillna(0)
    result_df['备货金额'] = result_df['建议备货量'] * result_df['成本单价']
    
    # 备货后库存金额 = (可用库存 + 采购在途 + 建议备货量) × 成本单价
    result_df['备货后库存量'] = result_df['可用库存'] + result_df['采购在途'] + result_df['建议备货量']
    result_df['备货后库存金额'] = result_df['备货后库存量'] * result_df['成本单价']
    
    # v1.5修改: 统计趋势因子和同比修正因子信息（A+/A类）
    ap_a_df = result_df[result_df['分类'].isin(['A+', 'A'])]
    trend_stats = {
        'rising': int((ap_a_df['趋势因子'] > 1.0).sum()),
        'falling': int((ap_a_df['趋势因子'] < 1.0).sum()),
    }
    yoy_stats = {
        'growing': int((ap_a_df['同比修正因子'] > 1.0).sum()),
        'shrinking': int((ap_a_df['同比修正因子'] < 1.0).sum()),
    }
    
    log(f"  趋势因子统计（A+/A类）：上升{trend_stats['rising']}个，下降{trend_stats['falling']}个")
    log(f"  同比修正统计（A+/A类）：增长{yoy_stats['growing']}个，下降{yoy_stats['shrinking']}个")
    
    # v1.5修改: 统计客户集中度风险（A+/A类SKU）
    concentration_stats = {'高风险': 0, '中风险': 0, '低风险': 0}
    if len(ap_a_df) > 0 and '集中度风险' in ap_a_df.columns:
        for risk in concentration_stats.keys():
            concentration_stats[risk] = int((ap_a_df['集中度风险'] == risk).sum())
    log(f"  客户集中度风险：高风险 {concentration_stats['高风险']} 个，中风险 {concentration_stats['中风险']} 个，低风险 {concentration_stats['低风险']} 个")
    
    # v1.4新增: 预测偏差回测
    log("\n【步骤20】执行预测偏差回测...")
    # 使用result_df（已有分类列）而非monthly_df
    # 构建带分类的月度数据供回测使用
    backtest_input = result_df.copy()
    backtest_df, backtest_mape = calculate_backtest(
        backtest_input, months, seasonal_factors, sku_names, logger
    )
    
    # 输出回测结果
    log(f"  回测结果：A+类MAPE={backtest_mape.get('A+', 'N/A')}%，A类MAPE={backtest_mape.get('A', 'N/A')}%，B类MAPE={backtest_mape.get('B', 'N/A')}%，C类MAPE={backtest_mape.get('C', 'N/A')}%，D类MAPE={backtest_mape.get('D', 'N/A')}%，E类MAPE={backtest_mape.get('E', 'N/A')}%，整体MAPE={backtest_mape.get('整体', 'N/A')}%")
    
    # 汇总统计
    summary = {
        'total_sku': len(result_df),
        'with_data_sku': len(result_df[result_df['分类'] != '零销量']),
        'zero_sku': len(result_df[result_df['分类'] == '零销量']),
        'class_counts': class_counts,
        'total_forecast': int(result_df['预测销量'].sum()),
        'total_replenishment': int(result_df['建议备货量'].sum()),
        'total_target': int(result_df['目标库存量'].sum()),
        'total_replenishment_amount': float(result_df['备货金额'].sum()),
        'total_stock_after_amount': float(result_df['备货后库存金额'].sum()),
    }
    
    # 生成Excel
    log("\n【步骤21】生成Excel报告...")
    wb = create_workbook()
    
    # 每个Sheet独立try/except，一个失败不影响其他
    sheet_tasks = [
        ("预测总览", lambda: write_summary_sheet(wb, summary, result_df, target_month, months, target_working_days,
                           trend_stats, yoy_stats, concentration_stats, lifecycle_stats, backtest_mape)),
        ("HW精细预测(A+-A)", lambda: write_forecast_sheet(wb, result_df, "HW精细预测(A+-A)", ['A+', 'A'], months)),
        ("因子预测(B-C)", lambda: write_forecast_sheet(wb, result_df, "因子预测(B-C)", ['B', 'C'], months)),
        ("简单MA预测(D-E)", lambda: write_forecast_sheet(wb, result_df, "简单MA预测(D-E)", ['D', 'E'], months)),
        ("零销量SKU", lambda: write_zero_sku_sheet(wb, result_df, months)),
        ("大单异常记录", lambda: write_anomaly_sheet(wb, anomaly_df)),
        ("季节因子表", lambda: write_seasonal_sheet(wb, seasonal_factors)),
        ("预测回测", lambda: write_backtest_sheet(wb, backtest_df, backtest_mape, logger)),
    ]
    
    for i, (sheet_name, sheet_func) in enumerate(sheet_tasks, 1):
        try:
            sheet_func()
            log(f"  ✓ Sheet{i}: {sheet_name}")
        except Exception as e:
            log(f"  ✗ Sheet{i}: {sheet_name} 生成失败: {e}")
            import traceback
            for line in traceback.format_exc().splitlines():
                log(f"    {line}")
    
    return wb, summary, result_df, anomaly_df, seasonal_factors, months, trend_stats, yoy_stats, concentration_stats, lifecycle_stats, backtest_df, backtest_mape


# ============================================================================
# GUI界面
# ============================================================================

def run_sales_forecast_gui(parent=None, embedded=False):
    """运行销量预测GUI界面；parent存在时作为YaFo主程序子窗口打开，不启动第二个mainloop。"""
    try:
        import tkinter as tk
        from tkinter import ttk, filedialog, messagebox
    except ImportError:
        print("错误：tkinter 不可用")
        sys.exit(1)
    
    # 创建窗口：独立运行时创建Tk；嵌入主生成器时创建Toplevel
    use_mainloop = parent is None
    root = tk.Tk() if use_mainloop else tk.Toplevel(parent)
    if parent is not None:
        try:
            root.transient(parent)
        except Exception:
            pass
    root.title(f"销量预测与备货建议生成器 {VERSION}")
    root.geometry("720x800")
    root.resizable(True, True)
    root.minsize(680, 700)
    
    # 居中显示
    root.update_idletasks()
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - 720) // 2
    y = (screen_height - 800) // 2
    root.geometry(f"720x800+{x}+{y}")
    
    # 变量
    sales_file_list = []  # 动态销售文件列表
    selected_files = {'inventory': None}
    output_path = {'path': None}
    is_running = {'flag': False}
    
    # ===== 顶部区域 =====
    top_frame = tk.Frame(root)
    top_frame.pack(fill=tk.X, padx=20, pady=(15, 5))
    
    # 标题 - v1.4: 包含版本号
    title_label = tk.Label(
        top_frame,
        text=f"销量预测与备货建议生成器 {VERSION}",
        font=("Microsoft YaHei", 16, "bold"),
        fg="#4472C4"
    )
    title_label.pack(pady=(0, 5))
    
    # 版本迭代日志（按钮查看）
    version_log_btn = tk.Button(
        top_frame,
        text="📋 查看版本迭代日志",
        command=lambda: show_version_log(),
        font=("Microsoft YaHei", 8),
        bg="#F0F0F0",
        relief=tk.FLAT,
        cursor="hand2"
    )
    version_log_btn.pack(anchor='w', pady=(2, 5))
    
    def show_version_log():
        log_win = tk.Toplevel(root)
        log_win.title("版本迭代日志")
        log_win.geometry("600x400")
        
        vl_text = tk.Text(
            log_win,
            font=("Microsoft YaHei", 9),
            wrap=tk.WORD,
            state='normal',
            bg="#F8F8F8",
            fg="#333333",
            padx=10,
            pady=8
        )
        vl_scrollbar = tk.Scrollbar(log_win, orient=tk.VERTICAL, command=vl_text.yview)
        vl_text.config(yscrollcommand=vl_scrollbar.set)
        vl_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        vl_text.pack(fill=tk.BOTH, expand=True)
        
        vl_text.insert('1.0', VERSION_LOG)
        vl_text.config(state='disabled')
        
        btn_frame = tk.Frame(log_win)
        btn_frame.pack(fill=tk.X, pady=5)
        tk.Button(btn_frame, text="关闭", command=log_win.destroy, width=10).pack()
    
    # ===== 文件选择区域 =====
    file_frame = tk.LabelFrame(root, text="数据文件导入", font=("Microsoft YaHei", 9, "bold"))
    file_frame.pack(fill=tk.X, padx=20, pady=5)
    
    # 销售文件区域（支持多文件）
    sales_label_frame = tk.Frame(file_frame)
    sales_label_frame.pack(fill=tk.X, padx=10, pady=(5, 0))
    tk.Label(sales_label_frame, text="销售订单文件*:", anchor='w').pack(side=tk.LEFT)
    sales_count_label = tk.Label(sales_label_frame, text="(已选0个)", fg="#999999", 
                                  font=("Microsoft YaHei", 8))
    sales_count_label.pack(side=tk.LEFT, padx=(5, 0))
    
    sales_btn_frame = tk.Frame(file_frame)
    sales_btn_frame.pack(fill=tk.X, padx=10, pady=2)
    tk.Button(sales_btn_frame, text="添加文件", command=lambda: add_sales_files(),
              width=10).pack(side=tk.LEFT)
    tk.Button(sales_btn_frame, text="移除选中", command=lambda: remove_sales_file(),
              width=10).pack(side=tk.LEFT, padx=(5, 0))
    tk.Button(sales_btn_frame, text="清空全部", command=lambda: clear_sales_files(),
              width=10).pack(side=tk.LEFT, padx=(5, 0))
    
    # 文件列表Listbox
    sales_list_frame = tk.Frame(file_frame)
    sales_list_frame.pack(fill=tk.X, padx=10, pady=2)
    sales_listbox = tk.Listbox(sales_list_frame, height=4, font=("Microsoft YaHei", 8),
                                selectmode=tk.EXTENDED, bg="#FAFAFA", relief=tk.GROOVE)
    sales_list_scroll = tk.Scrollbar(sales_list_frame, orient=tk.VERTICAL, 
                                      command=sales_listbox.yview)
    sales_listbox.config(yscrollcommand=sales_list_scroll.set)
    sales_list_scroll.pack(side=tk.RIGHT, fill=tk.Y)
    sales_listbox.pack(fill=tk.X, expand=True)
    
    # 库存文件（可选）
    row3 = tk.Frame(file_frame)
    row3.pack(fill=tk.X, padx=10, pady=2)
    tk.Label(row3, text="库存信息文件:", width=14, anchor='w').pack(side=tk.LEFT)
    inv_var = tk.StringVar(value="未选择（可选）")
    tk.Entry(row3, textvariable=inv_var, font=("Microsoft YaHei", 9), state='readonly',
             fg="#666666").pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    tk.Button(row3, text="浏览...", command=lambda: select_inventory_file(inv_var),
              width=8).pack(side=tk.RIGHT)
    tk.Label(row3, text="（可选）", fg="#999999", font=("Microsoft YaHei", 8)).pack(side=tk.RIGHT, padx=(5, 0))
    
    # ===== 参数设置区域 =====
    param_frame = tk.LabelFrame(root, text="参数设置", font=("Microsoft YaHei", 9, "bold"))
    param_frame.pack(fill=tk.X, padx=20, pady=5)
    
    param_row = tk.Frame(param_frame)
    param_row.pack(fill=tk.X, padx=10, pady=6)
    
    tk.Label(param_row, text="目标周转天数:", anchor='w').pack(side=tk.LEFT)
    target_days_var = tk.StringVar(value=str(DEFAULT_TARGET_DAYS))
    tk.Entry(param_row, textvariable=target_days_var, width=8).pack(side=tk.LEFT, padx=(5, 5))
    tk.Label(param_row, text="天").pack(side=tk.LEFT)
    
    tk.Label(param_row, text="到货周期:", anchor='w').pack(side=tk.LEFT, padx=(20, 0))
    lead_time_var = tk.StringVar(value=str(DEFAULT_LEAD_TIME))
    tk.Entry(param_row, textvariable=lead_time_var, width=8).pack(side=tk.LEFT, padx=(5, 5))
    tk.Label(param_row, text="天（安全库存=日均×到货周期）").pack(side=tk.LEFT)
    
    # ===== 日志区域 =====
    log_frame = tk.Frame(root)
    log_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=(5, 2))
    
    log_scrollbar = tk.Scrollbar(log_frame, orient=tk.VERTICAL)
    log_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
    
    log_text = tk.Text(
        log_frame,
        font=("Consolas", 9),
        wrap=tk.WORD,
        state='disabled',
        bg="#F8F8F8",
        fg="#333333",
        relief=tk.FLAT,
        padx=8,
        pady=5,
        yscrollcommand=log_scrollbar.set
    )
    log_text.pack(fill=tk.BOTH, expand=True)
    log_scrollbar.config(command=log_text.yview)
    
    # 日志颜色标签
    log_text.tag_config("title", foreground="#4472C4", font=("Consolas", 10, "bold"))
    log_text.tag_config("success", foreground="#006600")
    log_text.tag_config("error", foreground="#CC0000")
    log_text.tag_config("info", foreground="#333333")
    log_text.tag_config("step", foreground="#4472C4")
    
    def append_log(message, tag="info"):
        log_text.config(state='normal')
        log_text.insert(tk.END, message + "\n", tag)
        log_text.see(tk.END)
        log_text.config(state='disabled')
        # 同步更新状态栏（显示最新步骤）
        if '步骤' in message or '完成' in message or '分析' in message:
            try:
                status_var.set(message[:50])
                if tag == 'error':
                    status_label.config(fg="#CC0000")
                elif tag == 'success':
                    status_label.config(fg="#006600")
                else:
                    status_label.config(fg="#4472C4")
            except:
                pass
    
    # ===== 底部按钮区域（固定在底部，不被日志区挤掉）=====
    bottom_frame = tk.Frame(root)
    bottom_frame.pack(fill=tk.X, padx=20, pady=(2, 5), side=tk.BOTTOM)
    
    # 状态显示
    status_var = tk.StringVar(value="等待选择文件")
    status_label = tk.Label(
        bottom_frame,
        textvariable=status_var,
        font=("Microsoft YaHei", 10),
        fg="#333333"
    )
    status_label.pack(pady=(0, 3))
    
    # 按钮行
    btn_row = tk.Frame(bottom_frame)
    btn_row.pack()
    
    # 开始分析按钮
    analyze_btn = tk.Button(
        btn_row,
        text="开始分析",
        state='disabled',
        command=lambda: start_analysis(),
        font=("Microsoft YaHei", 11, "bold"),
        width=12,
        height=1,
        bg="#28A745",
        fg="white",
        activebackground="#218838"
    )
    analyze_btn.pack(side=tk.LEFT, padx=5)
    
    # 导出报告按钮
    export_btn = tk.Button(
        btn_row,
        text="导出报告",
        state='disabled',
        command=lambda: export_report(),
        font=("Microsoft YaHei", 11),
        width=12,
        height=1,
        bg="#6C757D",
        fg="white",
        activebackground="#5A6268"
    )
    export_btn.pack(side=tk.LEFT, padx=5)
    
    # 打开报告/文件夹按钮（初始隐藏）
    def open_file():
        if output_path['path'] and os.path.exists(output_path['path']):
            try:
                if platform.system() == 'Windows':
                    os.startfile(output_path['path'])
                elif platform.system() == 'Darwin':
                    subprocess.run(['open', output_path['path']])
                else:
                    subprocess.run(['xdg-open', output_path['path']])
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件：{e}")
        else:
            messagebox.showwarning("提示", "报告文件未找到，请先导出报告！")
    
    def open_folder():
        if output_path['path'] and os.path.exists(output_path['path']):
            folder = os.path.dirname(os.path.abspath(output_path['path']))
            try:
                if platform.system() == 'Windows':
                    os.startfile(folder)
                elif platform.system() == 'Darwin':
                    subprocess.run(['open', folder])
                else:
                    subprocess.run(['xdg-open', folder])
            except Exception as e:
                messagebox.showerror("错误", f"无法打开文件夹：{e}")
        else:
            messagebox.showwarning("提示", "报告文件未找到，请先导出报告！")
    
    open_file_btn = tk.Button(
        btn_row,
        text="打开报告",
        command=open_file,
        font=("Microsoft YaHei", 10),
        width=10
    )
    
    open_folder_btn = tk.Button(
        btn_row,
        text="打开文件夹",
        command=open_folder,
        font=("Microsoft YaHei", 10),
        width=10
    )
    
    # 导出日志按钮
    def export_log():
        log_text.config(state='normal')
        content = log_text.get('1.0', tk.END).strip()
        log_text.config(state='disabled')
        
        if not content:
            messagebox.showwarning("提示", "日志为空，无法导出")
            return
        
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_name = f"预测日志_{timestamp}.txt"
        
        save_path = filedialog.asksaveasfilename(
            defaultextension=".txt",
            filetypes=[("文本文件", "*.txt"), ("所有文件", "*.*")],
            initialfile=default_name,
            title="导出运行日志"
        )
        
        if save_path:
            try:
                with open(save_path, 'w', encoding='utf-8') as f:
                    f.write(f"销量预测与备货建议生成器 v1.6 运行日志\n")
                    f.write(f"导出时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                    f.write("=" * 50 + "\n\n")
                    f.write(content)
                messagebox.showinfo("成功", f"日志已导出到:\n{save_path}")
            except Exception as e:
                messagebox.showerror("错误", f"导出日志失败：{e}")
    
    export_log_btn = tk.Button(
        btn_row,
        text="导出日志",
        command=export_log,
        font=("Microsoft YaHei", 10),
        width=10
    )
    export_log_btn.pack(side=tk.LEFT, padx=5)
    
    # v1.4新增: 查看回测按钮
    def show_backtest_window():
        """显示回测结果弹窗"""
        backtest_mape = analysis_result.get('backtest_mape', {})
        backtest_df = analysis_result.get('backtest_df', pd.DataFrame())
        
        if not backtest_mape:
            messagebox.showinfo("提示", "暂无回测数据")
            return
        
        # 创建弹窗
        backtest_win = tk.Toplevel(root)
        backtest_win.title("预测偏差回测结果")
        backtest_win.geometry("600x500")
        backtest_win.transient(root)  # 关联主窗口
        backtest_win.grab_set()  # 模态
        
        # 居中显示
        backtest_win.update_idletasks()
        win_width = 600
        win_height = 500
        x = (backtest_win.winfo_screenwidth() - win_width) // 2
        y = (backtest_win.winfo_screenheight() - win_height) // 2
        backtest_win.geometry(f"{win_width}x{win_height}+{x}+{y}")
        
        # 标题
        title = tk.Label(backtest_win, text="预测偏差回测结果", font=("Microsoft YaHei", 14, "bold"), fg="#4472C4")
        title.pack(pady=10)
        
        # MAPE汇总
        summary_frame = tk.Frame(backtest_win)
        summary_frame.pack(fill=tk.X, padx=20, pady=10)
        
        tk.Label(summary_frame, text="【MAPE统计（6级分类）】", font=("Microsoft YaHei", 10, "bold")).pack(anchor='w')
        
        # v1.5修改：6级分类MAPE
        for cls in ['A+', 'A', 'B', 'C', 'D', 'E', '整体']:
            mape = backtest_mape.get(cls)
            label_text = f"{cls}类MAPE: {mape}%" if mape is not None else f"{cls}类MAPE: N/A"
            label = tk.Label(summary_frame, text=label_text, font=("Microsoft YaHei", 9))
            label.pack(anchor='w', pady=2)
            
            # 着色
            if mape is not None:
                if mape < 30:
                    label.config(fg="#375623", font=("Microsoft YaHei", 9, "bold"))
                elif mape < 60:
                    label.config(fg="#BF8F00", font=("Microsoft YaHei", 9, "bold"))
                else:
                    label.config(fg="#CC0000", font=("Microsoft YaHei", 9, "bold"))
        
        # 准确率最高/最低
        list_frame = tk.Frame(backtest_win)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=10)
        
        # 准确率最高Top10
        top_frame = tk.LabelFrame(list_frame, text="准确率最高的Top10 SKU（MAPE最低）", font=("Microsoft YaHei", 9))
        top_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        top_text = tk.Text(top_frame, font=("Consolas", 8), height=15, wrap=tk.WORD)
        top_scroll = tk.Scrollbar(top_frame, orient=tk.VERTICAL, command=top_text.yview)
        top_text.config(yscrollcommand=top_scroll.set)
        top_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        top_text.pack(fill=tk.BOTH, expand=True)
        
        # 准确率最低Top10
        bottom_frame = tk.LabelFrame(list_frame, text="准确率最低的Top10 SKU（MAPE最高）", font=("Microsoft YaHei", 9))
        bottom_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        bottom_text = tk.Text(bottom_frame, font=("Consolas", 8), height=15, wrap=tk.WORD)
        bottom_scroll = tk.Scrollbar(bottom_frame, orient=tk.VERTICAL, command=bottom_text.yview)
        bottom_text.config(yscrollcommand=bottom_scroll.set)
        bottom_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        bottom_text.pack(fill=tk.BOTH, expand=True)
        
        # 填充数据
        if len(backtest_df) > 0:
            # 有效MAPE数据
            valid_df = backtest_df[backtest_df['MAPE'].notna()].copy()
            
            # 按MAPE升序（准确率最高在前）
            top10 = valid_df.nsmallest(10, 'MAPE')
            for _, row in top10.iterrows():
                sku = row.get('商品编码', '')
                name = str(row.get('商品名称', ''))[:15]
                mape = row.get('MAPE', 'N/A')
                top_text.insert(tk.END, f"{sku}\n  {name}... MAPE={mape}%\n")
            
            # 按MAPE降序（准确率最低在前）
            bottom10 = valid_df.nlargest(10, 'MAPE')
            for _, row in bottom10.iterrows():
                sku = row.get('商品编码', '')
                name = str(row.get('商品名称', ''))[:15]
                mape = row.get('MAPE', 'N/A')
                bottom_text.insert(tk.END, f"{sku}\n  {name}... MAPE={mape}%\n")
        
        # 关闭按钮
        close_btn = tk.Button(backtest_win, text="关闭", command=backtest_win.destroy, width=10)
        close_btn.pack(pady=10)
    
    view_backtest_btn = tk.Button(
        btn_row,
        text="查看回测",
        command=show_backtest_window,
        font=("Microsoft YaHei", 10),
        width=10,
        state='disabled'
    )
    view_backtest_btn.pack(side=tk.LEFT, padx=5)
    
    result_btns_visible = [False]
    
    def show_result_btns():
        if not result_btns_visible[0]:
            open_file_btn.pack(side=tk.LEFT, padx=5)
            open_folder_btn.pack(side=tk.LEFT, padx=5)
            result_btns_visible[0] = True
    
    def hide_result_btns():
        if result_btns_visible[0]:
            open_file_btn.pack_forget()
            open_folder_btn.pack_forget()
            result_btns_visible[0] = False
    
    # 销售文件操作函数
    def add_sales_files():
        """添加销售文件（支持多选）"""
        file_paths = filedialog.askopenfilenames(
            title="选择销售订单文件（CSV/Excel，可多选）",
            filetypes=[("CSV/Excel文件", "*.csv *.xlsx *.xlsm *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")]
        )
        if file_paths:
            for fp in file_paths:
                if fp not in sales_file_list:
                    sales_file_list.append(fp)
                    sales_listbox.insert(tk.END, os.path.basename(fp))
            update_sales_count()
            check_can_analyze()
    
    def remove_sales_file():
        """移除选中的销售文件"""
        selections = list(sales_listbox.curselection())
        selections.sort(reverse=True)  # 从后往前删，避免索引错乱
        for idx in selections:
            sales_file_list.pop(idx)
            sales_listbox.delete(idx)
        update_sales_count()
        check_can_analyze()
    
    def clear_sales_files():
        """清空所有销售文件"""
        sales_file_list.clear()
        sales_listbox.delete(0, tk.END)
        update_sales_count()
        check_can_analyze()
    
    def update_sales_count():
        """更新已选文件数量显示"""
        sales_count_label.config(text=f"(已选{len(sales_file_list)}个)")
    
    # 库存文件选择函数
    def select_inventory_file(var):
        file_path = filedialog.askopenfilename(
            title="选择库存信息文件（CSV/Excel，可选）",
            filetypes=[("CSV/Excel文件", "*.csv *.xlsx *.xlsm *.xls"), ("CSV文件", "*.csv"), ("Excel文件", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")]
        )
        if file_path:
            selected_files['inventory'] = file_path
            var.set(file_path)
    
    def check_can_analyze():
        """检查是否可以开始分析"""
        can_analyze = len(sales_file_list) >= 1
        analyze_btn.config(state='normal' if can_analyze else 'disabled')
    
    # 分析结果存储
    analysis_result = {'wb': None, 'summary': None, 'result_df': None, 
                      'anomaly_df': None, 'seasonal_factors': None, 'months': None,
                      'trend_stats': None, 'yoy_stats': None, 'concentration_stats': None,
                      'lifecycle_stats': None, 'backtest_df': None, 'backtest_mape': None}  # v1.4新增
    
    def start_analysis():
        """开始分析"""
        if is_running['flag']:
            return
        
        # 关闭窗口确认
        if is_running['flag']:
            if not messagebox.askyesno("确认", "分析正在进行中，确定要取消吗？"):
                return
        
        # 清空日志
        log_text.config(state='normal')
        log_text.delete('1.0', tk.END)
        log_text.config(state='disabled')
        
        # 禁用按钮
        analyze_btn.config(state='disabled')
        export_btn.config(state='disabled')
        view_backtest_btn.config(state='disabled')
        status_var.set("正在分析...")
        status_label.config(fg="#4472C4")
        hide_result_btns()
        
        is_running['flag'] = True
        
        def analyze():
            try:
                # 获取参数
                target_days = int(target_days_var.get()) if target_days_var.get().strip() else DEFAULT_TARGET_DAYS
                lead_time = int(lead_time_var.get()) if lead_time_var.get().strip() else DEFAULT_LEAD_TIME
                
                # 运行预测
                result = run_prediction(
                    sales_file_list,
                    selected_files['inventory'],
                    target_days,
                    lead_time,
                    logger=append_log
                )
                
                if result[0] is not None:
                    analysis_result['wb'] = result[0]
                    analysis_result['summary'] = result[1]
                    analysis_result['result_df'] = result[2]
                    analysis_result['anomaly_df'] = result[3]
                    analysis_result['seasonal_factors'] = result[4]
                    analysis_result['months'] = result[5]
                    analysis_result['trend_stats'] = result[6]
                    analysis_result['yoy_stats'] = result[7]
                    analysis_result['concentration_stats'] = result[8]
                    analysis_result['lifecycle_stats'] = result[9]
                    analysis_result['backtest_df'] = result[10]  # v1.4新增
                    analysis_result['backtest_mape'] = result[11]  # v1.4新增
                    
                    # 生成默认输出路径
                    date_str = datetime.now().strftime("%Y%m%d")
                    default_path = os.path.join(os.getcwd(), f"销量预测报告_{date_str}.xlsx")
                    output_path['path'] = default_path
                    
                    root.after(0, on_analysis_success)
                else:
                    root.after(0, on_analysis_error, "分析失败")
                
            except Exception as e:
                import traceback
                error_msg = str(e)
                for line in traceback.format_exc().splitlines():
                    root.after(0, lambda l=line: append_log(l, "error"))
                root.after(0, on_analysis_error, error_msg)
            finally:
                is_running['flag'] = False
        
        thread = threading.Thread(target=analyze)
        thread.daemon = True
        thread.start()
    
    def on_analysis_success():
        status_var.set("✓ 分析完成")
        status_label.config(fg="#006600")
        analyze_btn.config(state='normal')
        export_btn.config(state='normal')
        view_backtest_btn.config(state='normal')  # v1.4新增：启用查看回测按钮
        show_result_btns()
        
        # 显示汇总 - v1.5: 6级分类统计
        summary = analysis_result['summary']
        trend_stats = analysis_result.get('trend_stats', {})
        yoy_stats = analysis_result.get('yoy_stats', {})
        concentration_stats = analysis_result.get('concentration_stats', {})
        backtest_mape = analysis_result.get('backtest_mape', {})
        
        append_log("", "info")
        append_log("=" * 50, "title")
        append_log(f"分析完成！v1.6", "success")
        append_log("=" * 50, "title")
        append_log(f"SKU总数: {summary['total_sku']}", "info")
        # v1.5修改：6级分类统计
        append_log(f"A+/A/B/C/D/E类SKU: {summary['class_counts'].get('A+', 0)}/{summary['class_counts'].get('A', 0)}/{summary['class_counts'].get('B', 0)}/{summary['class_counts'].get('C', 0)}/{summary['class_counts'].get('D', 0)}/{summary['class_counts'].get('E', 0)}", "info")
        append_log(f"零销量SKU: {summary['zero_sku']}", "info")
        append_log(f"预测总销量: {summary['total_forecast']}", "info")
        append_log(f"建议总备货量: {summary['total_replenishment']}", "info")
        append_log(f"总备货金额: ¥{summary.get('total_replenishment_amount', 0):,.2f}", "info")
        
        # v1.5修改：统计输出
        if trend_stats:
            append_log(f"趋势上升/下降: {trend_stats.get('rising', 0)}/{trend_stats.get('falling', 0)} 个", "info")
        if yoy_stats:
            append_log(f"YoY增长/下降: {yoy_stats.get('growing', 0)}/{yoy_stats.get('shrinking', 0)} 个", "info")
        if concentration_stats:
            append_log(f"客户集中度风险: 高{concentration_stats.get('高风险', 0)}/中{concentration_stats.get('中风险', 0)}/低{concentration_stats.get('低风险', 0)} 个", "info")
        lifecycle_stats = analysis_result.get('lifecycle_stats', {})
        if lifecycle_stats:
            append_log(f"产品生命周期: 上升期{lifecycle_stats.get('上升期', 0)}/成熟期{lifecycle_stats.get('成熟期', 0)}/衰退期{lifecycle_stats.get('衰退期', 0)}/休眠期{lifecycle_stats.get('休眠期', 0)}/新品期{lifecycle_stats.get('新品期', 0)} 个", "info")
        
        # v1.5修改: 回测MAPE输出（6级分类）
        if backtest_mape:
            mape_str = f"A+类MAPE={backtest_mape.get('A+', 'N/A')}%，A类MAPE={backtest_mape.get('A', 'N/A')}%，B类MAPE={backtest_mape.get('B', 'N/A')}%，C类MAPE={backtest_mape.get('C', 'N/A')}%，D类MAPE={backtest_mape.get('D', 'N/A')}%，E类MAPE={backtest_mape.get('E', 'N/A')}%，整体MAPE={backtest_mape.get('整体', 'N/A')}%"
            append_log(f"预测偏差回测: {mape_str}", "info")
        
        messagebox.showinfo("完成", f"分析完成！\n点击「导出报告」保存Excel文件。")
    
    def on_analysis_error(error_msg):
        status_var.set(f"出错：{error_msg[:30]}...")
        status_label.config(fg="#CC0000")
        analyze_btn.config(state='normal')
        is_running['flag'] = False
    
    def export_report():
        """导出报告"""
        if analysis_result['wb'] is None:
            messagebox.showwarning("警告", "请先执行分析！")
            return
        
        # 保存文件对话框
        filepath = filedialog.asksaveasfilename(
            title="保存销量预测报告",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx")],
            initialfile=f"销量预测报告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )
        
        if not filepath:
            return
        
        try:
            analysis_result['wb'].save(filepath)
            output_path['path'] = filepath
            
            append_log("", "info")
            append_log(f"报告已保存: {filepath}", "success")
            
            messagebox.showinfo("成功", f"报告已保存！\n{filepath}")
        except Exception as e:
            append_log(f"保存失败: {e}", "error")
            if "Permission denied" in str(e):
                messagebox.showerror("保存失败", f"文件被占用（可能在Excel中打开了）\n请先关闭Excel中的文件，再重新导出。\n\n或换一个文件名保存。")
            else:
                messagebox.showerror("错误", f"保存失败：{e}")
    
    # 窗口关闭处理
    def on_closing():
        if is_running['flag']:
            if not messagebox.askyesno("确认", "分析正在进行中，确定要退出吗？"):
                return
        root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    
    # 独立运行时启动mainloop；嵌入主程序时交由主程序mainloop接管
    if use_mainloop:
        root.mainloop()
    return root

# ===== 销量预测与备货建议生成器子模块结束 =====


def _open_pricing_module(self):
    """打开新品上新定价生成器子模块"""
    existing = getattr(self, 'pricing_window', None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.lift()
                existing.focus_force()
                return
        except Exception:
            pass
    self.pricing_window = PricingToolWindow(self)
    try:
        self.log("✓ 已打开新品上新定价子模块")
    except Exception:
        pass



def _open_inventory_cost_module(self):
    """打开存货成本回填生成器子模块"""
    existing = getattr(self, 'cost_fill_window', None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.lift()
                existing.focus_force()
                return
        except Exception:
            pass

    win = tk.Toplevel(self.root)
    self.cost_fill_window = win
    self.cost_fill_app = InventoryCostApp(win)

    def _on_close():
        try:
            win.destroy()
        finally:
            self.cost_fill_window = None
            self.cost_fill_app = None

    win.protocol("WM_DELETE_WINDOW", _on_close)
    try:
        self.log("✓ 已打开存货成本回填子模块")
    except Exception:
        pass



def _open_sales_forecast_module(self):
    """打开销量预测与备货建议生成器子模块"""
    existing = getattr(self, 'sales_forecast_window', None)
    if existing is not None:
        try:
            if existing.winfo_exists():
                existing.lift()
                existing.focus_force()
                return
        except Exception:
            pass

    try:
        win = run_sales_forecast_gui(parent=self.root, embedded=True)
        self.sales_forecast_window = win

        def _mark_closed(event=None):
            try:
                if event is None or event.widget == win:
                    self.sales_forecast_window = None
            except Exception:
                self.sales_forecast_window = None

        # 不覆盖子模块自身的关闭协议，保留“分析进行中”退出确认；仅监听销毁事件以清理引用。
        try:
            win.bind("<Destroy>", _mark_closed, add="+")
        except TypeError:
            win.bind("<Destroy>", _mark_closed)
        try:
            self.log("✓ 已打开销量预测与备货建议子模块")
        except Exception:
            pass
    except Exception as e:
        messagebox.showerror("打开失败", f"销量预测与备货建议子模块打开失败：\n{e}")
        try:
            self.log(f"❌ 销量预测与备货建议子模块打开失败：{e}")
        except Exception:
            pass

PurchaseOrderGenerator.open_supplier_export_module = _open_supplier_export_module
PurchaseOrderGenerator.add_factory_files_from_submodule = _add_factory_files_from_submodule
PurchaseOrderGenerator.open_pricing_module = _open_pricing_module
PurchaseOrderGenerator.open_inventory_cost_module = _open_inventory_cost_module
PurchaseOrderGenerator.open_sales_forecast_module = _open_sales_forecast_module

def main():
    root = tk.Tk()
    app = PurchaseOrderGenerator(root)
    root.mainloop()


if __name__ == "__main__":
    main()

